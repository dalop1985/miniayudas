import os
import re
import decimal
import threading
import secrets
import hashlib
import uuid
import time
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import SpooledTemporaryFile
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pyodbc
import orjson
from dotenv import load_dotenv
from fastapi import FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, ORJSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
 
BASE_DIR = Path(__file__).resolve().parent.parent

dotenv_path = os.getenv("DOTENV_PATH")
load_dotenv(dotenv_path=dotenv_path or (BASE_DIR / ".env"), override=False)
 
app = FastAPI()

_cors_origins_raw = os.getenv("CORS_ORIGINS") or ""
_cors_origins = [s.strip() for s in _cors_origins_raw.split(",") if s.strip()]
if _cors_origins:
  app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["*"],
  )
 
UMA_MXN_BY_VIGENCIA_YEAR: Dict[int, decimal.Decimal] = {}
_UMA_STORE_LOCK = threading.Lock()
_UMA_STORE_PATH = Path(os.getenv("UMA_STORE_PATH") or (BASE_DIR / "umas.json"))

_INPC_STORE_LOCK = threading.Lock()
_INPC_STORE_DIR = Path(os.getenv("INPC_STORE_DIR") or (BASE_DIR / "data"))
_INPC_FILE_NAME_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.-]{0,120}\.json$")

RECARGOS_MORA_BY_YEAR: Dict[int, decimal.Decimal] = {}
_RECARGOS_STORE_LOCK = threading.Lock()
_RECARGOS_STORE_PATH = Path(os.getenv("RECARGOS_STORE_PATH") or (_INPC_STORE_DIR / "recargos_historico.json"))

_UPLOAD_JSON_MAX_BYTES = int(os.getenv("UPLOAD_JSON_MAX_BYTES") or "1048576")

_LOGIN_RATE_LOCK = threading.Lock()
_LOGIN_RATE_BUCKETS: Dict[str, List[float]] = {}
_LOGIN_RATE_WINDOW_SECONDS = int(os.getenv("LOGIN_RATE_WINDOW_SECONDS") or "300")
_LOGIN_RATE_MAX = int(os.getenv("LOGIN_RATE_MAX") or "20")

_API_RATE_LOCK = threading.Lock()
_API_RATE_BUCKETS: Dict[str, List[float]] = {}
_API_RATE_WINDOW_SECONDS = int(os.getenv("API_RATE_WINDOW_SECONDS") or "60")
_API_RATE_MAX = int(os.getenv("API_RATE_MAX") or "600")

_REPORT_RATE_LOCK = threading.Lock()
_REPORT_RATE_BUCKETS: Dict[str, List[float]] = {}
_REPORT_RATE_WINDOW_SECONDS = int(os.getenv("REPORT_RATE_WINDOW_SECONDS") or "60")
_REPORT_RATE_MAX = int(os.getenv("REPORT_RATE_MAX") or "60")

_REPORT_ASYNC_THRESHOLD_ROWS = int(os.getenv("REPORT_ASYNC_THRESHOLD_ROWS") or "10000")
_REPORT_JOBS_TTL_SECONDS = int(os.getenv("REPORT_JOBS_TTL_SECONDS") or str(24 * 60 * 60))
_REPORT_JOBS_DIR = Path(os.getenv("REPORT_JOBS_DIR") or (BASE_DIR / "tmp_report_jobs"))

_DEADLOCK_RETRIES = int(os.getenv("DEADLOCK_RETRIES") or "3")


def _boolish(value: Any, default: bool = False) -> bool:
  if value in (None, "", "null"):
    return bool(default)
  return str(value).strip().lower() in {"1", "true", "yes", "y", "si", "sí"}


_AUDIT_ENABLED = _boolish(os.getenv("AUDIT_ENABLED"), True)


def _client_ip(request: Request) -> str:
  try:
    xff = request.headers.get("x-forwarded-for") or ""
    if xff:
      return xff.split(",")[0].strip() or "unknown"
    if request.client and request.client.host:
      return str(request.client.host)
  except Exception:
    pass
  return "unknown"


def _is_secure_request(request: Request) -> bool:
  try:
    xf_proto = (request.headers.get("x-forwarded-proto") or "").strip().lower()
    if xf_proto == "https":
      return True
    if xf_proto == "http":
      return False
  except Exception:
    pass
  return (request.url.scheme == "https")


def _reject_if_too_large(request: Request, max_bytes: int) -> None:
  try:
    raw = request.headers.get("content-length")
    if raw is None:
      return
    n = int(str(raw).strip())
    if n > int(max_bytes):
      raise HTTPException(status_code=413, detail="Contenido demasiado grande")
  except HTTPException:
    raise
  except Exception:
    return


def _check_login_rate_limit(request: Request) -> None:
  ip = _client_ip(request)
  now = datetime.utcnow().timestamp()
  with _LOGIN_RATE_LOCK:
    bucket = _LOGIN_RATE_BUCKETS.get(ip)
    if not bucket:
      bucket = []
      _LOGIN_RATE_BUCKETS[ip] = bucket
    cutoff = now - float(_LOGIN_RATE_WINDOW_SECONDS)
    while bucket and bucket[0] < cutoff:
      bucket.pop(0)
    if len(bucket) >= int(_LOGIN_RATE_MAX):
      raise HTTPException(status_code=429, detail="Demasiados intentos, espera unos minutos")
    bucket.append(now)


def _check_ip_rate_limit(
  request: Request,
  buckets: Dict[str, List[float]],
  lock: threading.Lock,
  window_seconds: int,
  max_requests: int,
  scope: str,
) -> None:
  ip = _client_ip(request)
  now = datetime.utcnow().timestamp()
  key = f"{scope}:{ip}"
  with lock:
    bucket = buckets.get(key)
    if not bucket:
      bucket = []
      buckets[key] = bucket
    cutoff = now - float(window_seconds)
    while bucket and bucket[0] < cutoff:
      bucket.pop(0)
    if len(bucket) >= int(max_requests):
      raise HTTPException(status_code=429, detail="Demasiadas solicitudes, intenta más tarde")
    bucket.append(now)


def _is_deadlock_error(exc: Exception) -> bool:
  try:
    msg = str(exc).lower()
    if "deadlock" in msg or "1205" in msg:
      return True
    if "40001" in msg:
      return True
  except Exception:
    return False
  return False

@app.middleware("http")
async def _security_headers(request: Request, call_next):
  resp = await call_next(request)
  try:
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("Referrer-Policy", "no-referrer")
    resp.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
  except Exception:
    pass
  return resp


def _load_umas_from_disk() -> None:
  try:
    path = _UMA_STORE_PATH
    if not path.exists():
      return
    raw = path.read_bytes()
    data = orjson.loads(raw) if raw else {}
    items = data.get("items") if isinstance(data, dict) else None
    if not isinstance(items, list):
      return
    UMA_MXN_BY_VIGENCIA_YEAR.clear()
    for item in items:
      if not isinstance(item, dict):
        continue
      year_raw = item.get("vigenciaYear")
      uma_raw = item.get("umaMxn")
      try:
        year = int(str(year_raw).strip())
      except Exception:
        continue
      if uma_raw in (None, "", "null"):
        continue
      try:
        uma = decimal.Decimal(str(uma_raw).strip())
      except Exception:
        continue
      UMA_MXN_BY_VIGENCIA_YEAR[year] = uma
  except Exception:
    return


def _save_umas_to_disk() -> None:
  path = _UMA_STORE_PATH
  path.parent.mkdir(parents=True, exist_ok=True)
  items = []
  for year in sorted(UMA_MXN_BY_VIGENCIA_YEAR.keys()):
    v = UMA_MXN_BY_VIGENCIA_YEAR.get(year)
    if v is None:
      continue
    items.append({"vigenciaYear": year, "umaMxn": str(v)})
  payload = {"version": 1, "items": items, "count": len(items)}
  data = orjson.dumps(payload, option=orjson.OPT_INDENT_2)
  tmp = path.with_suffix(path.suffix + ".tmp")
  tmp.write_bytes(data)
  os.replace(tmp, path)


def _safe_inpc_filename(name: Any) -> str:
  base = str(name or "").strip()
  base = base.replace("\\", "/").split("/")[-1].strip()
  base = re.sub(r"\s+", "_", base)
  if not base.lower().endswith(".json"):
    base = base + ".json"
  if not _INPC_FILE_NAME_RE.match(base):
    raise HTTPException(status_code=400, detail="Nombre de archivo INPC inválido")
  return base


def _flatten_inpc_payload(payload: Any) -> Dict[str, Any]:
  if payload is None:
    return {}
  if not isinstance(payload, dict):
    raise HTTPException(status_code=400, detail="INPC inválido")
  if "inpc_historico" in payload and isinstance(payload.get("inpc_historico"), dict):
    out: Dict[str, Any] = {}
    for _year, months in (payload.get("inpc_historico") or {}).items():
      if not isinstance(months, dict):
        continue
      for k, v in months.items():
        out[str(k)] = v
    return out
  return payload


def _list_inpc_files() -> List[Path]:
  try:
    base = _INPC_STORE_DIR
    if not base.exists() or not base.is_dir():
      return []
    files = [p for p in base.glob("inpc*.json") if p.is_file()]
    files.sort(key=lambda p: p.name.lower())
    return files
  except Exception:
    return []


def _load_inpc_table_from_disk() -> Tuple[Dict[int, decimal.Decimal], List[str]]:
  flat: Dict[str, Any] = {}
  sources: List[str] = []
  for p in _list_inpc_files():
    try:
      raw = p.read_bytes()
      data = orjson.loads(raw) if raw else {}
      part = _flatten_inpc_payload(data)
      if not isinstance(part, dict):
        continue
      flat.update(part)
      sources.append(p.name)
    except Exception:
      continue
  return _parse_inpc_table(flat), sources


def _load_recargos_from_disk() -> None:
  try:
    path = _RECARGOS_STORE_PATH
    if not path.exists():
      return
    raw = path.read_bytes()
    data = orjson.loads(raw) if raw else {}
    if not isinstance(data, dict):
      return
    items = data.get("items") if isinstance(data.get("items"), list) else None
    tasas = data.get("tasas") if isinstance(data.get("tasas"), dict) else None
    RECARGOS_MORA_BY_YEAR.clear()
    if items:
      for item in items:
        if not isinstance(item, dict):
          continue
        year_raw = item.get("year") if item.get("year") is not None else item.get("ejercicioYear")
        tasa_raw = item.get("tasaMora") if item.get("tasaMora") is not None else item.get("tasaMensual")
        try:
          year = int(str(year_raw).strip())
        except Exception:
          continue
        if tasa_raw in (None, "", "null"):
          continue
        try:
          tasa = decimal.Decimal(str(tasa_raw).strip())
        except Exception:
          continue
        if tasa <= 0:
          continue
        RECARGOS_MORA_BY_YEAR[year] = tasa
      return
    if tasas:
      for k, v in tasas.items():
        try:
          year = int(str(k).strip())
        except Exception:
          continue
        try:
          tasa = decimal.Decimal(str(v).strip())
        except Exception:
          continue
        if tasa <= 0:
          continue
        RECARGOS_MORA_BY_YEAR[year] = tasa
  except Exception:
    return


def _save_recargos_to_disk() -> None:
  path = _RECARGOS_STORE_PATH
  path.parent.mkdir(parents=True, exist_ok=True)
  items = []
  for year in sorted(RECARGOS_MORA_BY_YEAR.keys()):
    v = RECARGOS_MORA_BY_YEAR.get(year)
    if v is None:
      continue
    items.append({"year": year, "tasaMora": str(v)})
  payload = {"version": 1, "tipo": "mora", "items": items, "count": len(items)}
  data = orjson.dumps(payload, option=orjson.OPT_INDENT_2)
  tmp = path.with_suffix(path.suffix + ".tmp")
  tmp.write_bytes(data)
  os.replace(tmp, path)


def _recargo_mora_rate_for_year(year: int, table: Dict[int, decimal.Decimal]) -> decimal.Decimal:
  if not table:
    return decimal.Decimal("0")
  y = int(year)
  if y in table:
    return table[y]
  candidates = [k for k in table.keys() if k <= y]
  if candidates:
    return table[max(candidates)]
  return table[min(table.keys())]
 
 
def _uma_vigencia_year_for_date(value: datetime) -> int:
  return value.year if value.month >= 2 else value.year - 1
 
 
def get_uma_mxn_for_date(value: datetime) -> Optional[decimal.Decimal]:
  return UMA_MXN_BY_VIGENCIA_YEAR.get(_uma_vigencia_year_for_date(value))
 

_load_umas_from_disk()
_load_recargos_from_disk()

 
def _normalize_bool(value: Any, fallback: bool = False) -> bool:
  if value is None or value == "":
    return fallback
  return str(value).strip().lower() in {"true", "1", "yes", "y", "si", "sí"}
 
 
def _parse_extra_params(raw: str) -> Dict[str, str]:
  params: Dict[str, str] = {}
  if not raw:
    return params
  for item in str(raw).split(";"):
    chunk = item.strip()
    if not chunk:
      continue
    parts = chunk.split("=")
    key = parts[0].strip().lower()
    val = "=".join(parts[1:]).strip() if len(parts) > 1 else ""
    params[key] = val
  return params


def _parse_factus_input(raw: Any) -> List[Dict[str, Any]]:
  items: List[Dict[str, Any]] = []
  text = str(raw or "")
  for line in text.splitlines():
    s = line.strip()
    if not s:
      continue
    serie_match = re.search(r"\bserie\b\s*([A-Za-z]{1,10})", s, flags=re.IGNORECASE)
    serie = (serie_match.group(1) if serie_match else "")
    if not serie:
      serie_start = re.match(r"\s*([A-Za-z]{1,10})\b", s)
      serie = (serie_start.group(1) if serie_start else "")
    if not serie:
      continue
    folios = re.findall(r"\d+", s)
    for folio in folios:
      try:
        items.append({"serie": serie.upper(), "folio": int(folio)})
      except Exception:
        continue
  return items


def _parse_factus_range(raw: Any) -> Optional[Dict[str, Any]]:
  if raw is None:
    return None
  if isinstance(raw, dict):
    serie = str(raw.get("serie") or "").strip().upper()
    folio_from_raw = raw.get("folioFrom")
    folio_to_raw = raw.get("folioTo")
    if not serie and (folio_from_raw is None and folio_to_raw is None):
      return None
    try:
      folio_from = int(str(folio_from_raw).strip())
      folio_to = int(str(folio_to_raw).strip())
    except Exception:
      raise HTTPException(status_code=400, detail="Rango de folios inválido")
    if not serie:
      raise HTTPException(status_code=400, detail="Serie requerida para rango")
    if folio_from <= 0 or folio_to <= 0:
      raise HTTPException(status_code=400, detail="Rango de folios inválido")
    if folio_to < folio_from:
      folio_from, folio_to = folio_to, folio_from
    return {"serie": serie, "folioFrom": folio_from, "folioTo": folio_to}
  return None
 
 
def get_db_target() -> Dict[str, Any]:
  server = os.getenv("DB_HOST") or os.getenv("DB_SERVER")
  database = os.getenv("DB_NAME") or os.getenv("DB_DATABASE") or "Tulum"
  user = os.getenv("DB_USER")
  password = os.getenv("DB_PASSWORD")
  port_raw = os.getenv("DB_PORT")
  port = int(port_raw) if port_raw and port_raw.isdigit() else None
  extra = _parse_extra_params(os.getenv("DB_EXTRA_PARAMS") or "")
 
  if not server:
    raise ValueError("Falta DB_HOST o DB_SERVER")
  if not user:
    raise ValueError("Falta DB_USER")
  if not password:
    raise ValueError("Falta DB_PASSWORD")
 
  encrypt = _normalize_bool(extra.get("encrypt"), _normalize_bool(os.getenv("DB_ENCRYPT"), True))
  trust_server_certificate = _normalize_bool(
    extra.get("trustservercertificate"),
    _normalize_bool(os.getenv("DB_TRUST_SERVER_CERT"), True),
  )
  tls_server_name = os.getenv("DB_TLS_SERVER_NAME") or ""
 
  return {
    "server": server,
    "port": port,
    "database": database,
    "user": user,
    "password": password,
    "encrypt": encrypt,
    "trustServerCertificate": trust_server_certificate,
    "serverName": tls_server_name or None,
  }
 
 
def _connection_string() -> str:
  return ""
 
 
def get_conn():
  cfg = get_db_target()
  driver = os.getenv("DB_ODBC_DRIVER") or "ODBC Driver 17 for SQL Server"
  
  # For pyodbc, server port must be specified with comma if using default SQL Server ODBC
  server_addr = f"{cfg['server']},{cfg['port']}" if cfg['port'] else cfg['server']
  
  conn_str = (
      f"DRIVER={{{driver}}};"
      f"SERVER={server_addr};"
      f"DATABASE={cfg['database']};"
      f"UID={cfg['user']};"
      f"PWD={cfg['password']};"
  )
  
  if cfg["encrypt"]:
      conn_str += "Encrypt=yes;"
  else:
      conn_str += "Encrypt=no;"
      
  if cfg["trustServerCertificate"]:
      conn_str += "TrustServerCertificate=yes;"
  else:
      conn_str += "TrustServerCertificate=no;"
      
  # pyodbc handles connection pooling natively at the ODBC driver manager level by default
  return pyodbc.connect(conn_str, autocommit=False, timeout=60)
 

def get_conn_for_database(database: str):
  cfg = get_db_target()
  driver = os.getenv("DB_ODBC_DRIVER") or "ODBC Driver 17 for SQL Server"
  server_addr = f"{cfg['server']},{cfg['port']}" if cfg['port'] else cfg['server']
  conn_str = (
      f"DRIVER={{{driver}}};"
      f"SERVER={server_addr};"
      f"DATABASE={database};"
      f"UID={cfg['user']};"
      f"PWD={cfg['password']};"
  )
  
  if cfg["encrypt"]:
      conn_str += "Encrypt=yes;"
  else:
      conn_str += "Encrypt=no;"
      
  if cfg["trustServerCertificate"]:
      conn_str += "TrustServerCertificate=yes;"
  else:
      conn_str += "TrustServerCertificate=no;"
      
  return pyodbc.connect(conn_str, autocommit=False, timeout=60)

 
def _rows(cursor) -> List[Dict[str, Any]]:
  while cursor.description is None:
    if not cursor.nextset():
      return []
  cols = [c[0] for c in cursor.description]
  
  rows = []
  for row in cursor.fetchall():
    d = {}
    for k, v in zip(cols, row):
      # orjson does not serialize Decimal natively by default, convert to float
      if isinstance(v, decimal.Decimal):
        d[k] = float(v)
      else:
        d[k] = v
    rows.append(d)
  return rows
 
 
def _csv_escape(value: Any) -> str:
  if value is None:
    return ""
  if isinstance(value, datetime):
    s = value.isoformat()
  else:
    s = str(value)
  if any(ch in s for ch in [",", "\"", "\n", "\r"]):
    return "\"" + s.replace("\"", "\"\"") + "\""
  return s
 
 
def _require_admin(x_admin_key: Optional[str]) -> None:
  expected = str(os.getenv("ADMIN_KEY") or "").strip()
  if not expected:
    raise HTTPException(status_code=503, detail="ADMIN_KEY no configurada en el servidor")
  provided = str(x_admin_key or "").strip()
  if not provided or not secrets.compare_digest(provided, expected):
    raise HTTPException(status_code=401, detail="No autorizado")


def _has_valid_admin_key(request: Request) -> bool:
  expected = str(os.getenv("ADMIN_KEY") or "").strip()
  if not expected:
    return False
  provided = str(request.headers.get("x-admin-key") or "").strip()
  if not provided:
    return False
  try:
    return secrets.compare_digest(provided, expected)
  except Exception:
    return False
 

def _auth_secret() -> Optional[bytes]:
  raw = os.getenv("AUTH_SECRET") or os.getenv("ADMIN_KEY") or ""
  raw = str(raw).strip()
  if not raw:
    return None
  return raw.encode("utf-8")


def _sha256_bytes(value: str) -> bytes:
  return hashlib.sha256(value.encode("utf-8")).digest()


def _pbkdf2_hash(password: str, salt: bytes, iterations: int) -> bytes:
  return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(iterations), dklen=32)


def _ensure_auth_schema_microservicios() -> None:
  ddl: List[str] = []
  ddl.append(
    """
    IF OBJECT_ID('dbo.AppUsers', 'U') IS NULL
    BEGIN
      CREATE TABLE dbo.AppUsers (
        UserId            INT IDENTITY(1,1) NOT NULL CONSTRAINT PK_AppUsers PRIMARY KEY,
        Username          NVARCHAR(120) NOT NULL,
        DisplayName       NVARCHAR(200) NULL,
        Role              NVARCHAR(50) NOT NULL CONSTRAINT DF_AppUsers_Role DEFAULT ('cajero'),
        PasswordSalt      VARBINARY(32) NOT NULL,
        PasswordHash      VARBINARY(32) NOT NULL,
        PasswordIterations INT NOT NULL,
        IsActive          BIT NOT NULL CONSTRAINT DF_AppUsers_IsActive DEFAULT (1),
        FailedAttempts    INT NOT NULL CONSTRAINT DF_AppUsers_FailedAttempts DEFAULT (0),
        LockedUntil       DATETIME2(0) NULL,
        CreatedAt         DATETIME2(0) NOT NULL CONSTRAINT DF_AppUsers_CreatedAt DEFAULT (SYSUTCDATETIME()),
        UpdatedAt         DATETIME2(0) NOT NULL CONSTRAINT DF_AppUsers_UpdatedAt DEFAULT (SYSUTCDATETIME()),
        CONSTRAINT UQ_AppUsers_Username UNIQUE (Username)
      );
      CREATE INDEX IX_AppUsers_Role ON dbo.AppUsers(Role);
      CREATE INDEX IX_AppUsers_IsActive ON dbo.AppUsers(IsActive);
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('dbo.AppSessions', 'U') IS NULL
    BEGIN
      CREATE TABLE dbo.AppSessions (
        SessionId   BIGINT IDENTITY(1,1) NOT NULL CONSTRAINT PK_AppSessions PRIMARY KEY,
        TokenHash   VARBINARY(32) NOT NULL,
        UserId      INT NOT NULL,
        CreatedAt   DATETIME2(0) NOT NULL CONSTRAINT DF_AppSessions_CreatedAt DEFAULT (SYSUTCDATETIME()),
        ExpiresAt   DATETIME2(0) NOT NULL,
        Ip          NVARCHAR(45) NULL,
        UserAgent   NVARCHAR(512) NULL,
        RevokedAt   DATETIME2(0) NULL,
        CONSTRAINT UQ_AppSessions_TokenHash UNIQUE (TokenHash),
        CONSTRAINT FK_AppSessions_UserId FOREIGN KEY (UserId) REFERENCES dbo.AppUsers(UserId)
      );
      CREATE INDEX IX_AppSessions_UserId ON dbo.AppSessions(UserId);
      CREATE INDEX IX_AppSessions_ExpiresAt ON dbo.AppSessions(ExpiresAt) INCLUDE (RevokedAt);
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('dbo.AppAuthAudit', 'U') IS NULL
    BEGIN
      CREATE TABLE dbo.AppAuthAudit (
        AuditId     BIGINT IDENTITY(1,1) NOT NULL CONSTRAINT PK_AppAuthAudit PRIMARY KEY,
        CreatedAt   DATETIME2(0) NOT NULL CONSTRAINT DF_AppAuthAudit_CreatedAt DEFAULT (SYSUTCDATETIME()),
        Event       NVARCHAR(50) NOT NULL,
        Username    NVARCHAR(120) NULL,
        UserId      INT NULL,
        Ip          NVARCHAR(45) NULL,
        UserAgent   NVARCHAR(512) NULL,
        Detail      NVARCHAR(400) NULL
      );
      CREATE INDEX IX_AppAuthAudit_Event ON dbo.AppAuthAudit(Event, CreatedAt);
      CREATE INDEX IX_AppAuthAudit_UserId ON dbo.AppAuthAudit(UserId, CreatedAt);
    END
    """
  )

  ddl.append(
    """
    IF OBJECT_ID('dbo.AppLocks', 'U') IS NULL
    BEGIN
      CREATE TABLE dbo.AppLocks (
        LockKey       NVARCHAR(200) NOT NULL CONSTRAINT PK_AppLocks PRIMARY KEY,
        OwnerUserId   INT NULL,
        OwnerUsername NVARCHAR(120) NULL,
        AcquiredAt    DATETIME2(0) NOT NULL CONSTRAINT DF_AppLocks_AcquiredAt DEFAULT (SYSUTCDATETIME()),
        ExpiresAt     DATETIME2(0) NOT NULL,
        UpdatedAt     DATETIME2(0) NOT NULL CONSTRAINT DF_AppLocks_UpdatedAt DEFAULT (SYSUTCDATETIME())
      );
      CREATE INDEX IX_AppLocks_ExpiresAt ON dbo.AppLocks(ExpiresAt);
    END
    """
  )

  ddl.append(
    """
    IF OBJECT_ID('dbo.AppIdempotency', 'U') IS NULL
    BEGIN
      CREATE TABLE dbo.AppIdempotency (
        IdempotencyKey NVARCHAR(80) NOT NULL,
        Path           NVARCHAR(200) NOT NULL,
        UserId         INT NOT NULL CONSTRAINT DF_AppIdempotency_UserId DEFAULT (0),
        CreatedAt      DATETIME2(0) NOT NULL CONSTRAINT DF_AppIdempotency_CreatedAt DEFAULT (SYSUTCDATETIME()),
        RequestHash    VARBINARY(32) NULL,
        StatusCode     INT NOT NULL,
        ResponseJson   NVARCHAR(MAX) NOT NULL,
        CONSTRAINT PK_AppIdempotency PRIMARY KEY (IdempotencyKey, Path, UserId)
      );
      CREATE INDEX IX_AppIdempotency_CreatedAt ON dbo.AppIdempotency(CreatedAt);
    END
    """
  )

  ddl.append(
    """
    IF OBJECT_ID('dbo.AppApiAudit', 'U') IS NULL
    BEGIN
      CREATE TABLE dbo.AppApiAudit (
        AuditId     BIGINT IDENTITY(1,1) NOT NULL CONSTRAINT PK_AppApiAudit PRIMARY KEY,
        CreatedAt   DATETIME2(0) NOT NULL CONSTRAINT DF_AppApiAudit_CreatedAt DEFAULT (SYSUTCDATETIME()),
        RequestId   NVARCHAR(60) NULL,
        UserId      INT NULL,
        Username    NVARCHAR(120) NULL,
        Role        NVARCHAR(50) NULL,
        Ip          NVARCHAR(45) NULL,
        UserAgent   NVARCHAR(512) NULL,
        Method      NVARCHAR(10) NOT NULL,
        Path        NVARCHAR(200) NOT NULL,
        StatusCode  INT NOT NULL,
        DurationMs  INT NOT NULL
      );
      CREATE INDEX IX_AppApiAudit_CreatedAt ON dbo.AppApiAudit(CreatedAt);
      CREATE INDEX IX_AppApiAudit_Path ON dbo.AppApiAudit(Path, CreatedAt);
      CREATE INDEX IX_AppApiAudit_UserId ON dbo.AppApiAudit(UserId, CreatedAt);
    END
    """
  )

  ddl.append(
    """
    IF OBJECT_ID('dbo.AppReportJobs', 'U') IS NULL
    BEGIN
      CREATE TABLE dbo.AppReportJobs (
        JobId            UNIQUEIDENTIFIER NOT NULL CONSTRAINT PK_AppReportJobs PRIMARY KEY,
        Kind             NVARCHAR(80) NOT NULL,
        ParamsJson       NVARCHAR(MAX) NOT NULL,
        Status           NVARCHAR(20) NOT NULL,
        CreatedByUserId  INT NULL,
        CreatedByUsername NVARCHAR(120) NULL,
        CreatedAt        DATETIME2(0) NOT NULL CONSTRAINT DF_AppReportJobs_CreatedAt DEFAULT (SYSUTCDATETIME()),
        StartedAt        DATETIME2(0) NULL,
        FinishedAt       DATETIME2(0) NULL,
        ExpiresAt        DATETIME2(0) NOT NULL,
        RowsCount        INT NULL,
        FileName         NVARCHAR(200) NULL,
        ErrorMessage     NVARCHAR(400) NULL
      );
      CREATE INDEX IX_AppReportJobs_Status ON dbo.AppReportJobs(Status, CreatedAt);
      CREATE INDEX IX_AppReportJobs_ExpiresAt ON dbo.AppReportJobs(ExpiresAt);
      CREATE INDEX IX_AppReportJobs_CreatedBy ON dbo.AppReportJobs(CreatedByUserId, CreatedAt);
    END
    """
  )

  conn = get_conn_for_database("MicroServicios")
  try:
    cur = conn.cursor()
    cur.execute("SET NOCOUNT ON; SET XACT_ABORT ON;")
    for stmt in ddl:
      cur.execute(stmt)
    conn.commit()
  except Exception:
    conn.rollback()
    raise
  finally:
    conn.close()


def _audit_auth_event(event: str, username: Optional[str], user_id: Optional[int], request: Request, detail: Optional[str] = None) -> None:
  try:
    _ensure_auth_schema_microservicios()
    ip = (request.client.host if request.client else None) or None
    ua = request.headers.get("user-agent") or None
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        INSERT INTO dbo.AppAuthAudit (Event, Username, UserId, Ip, UserAgent, Detail)
        VALUES (?, ?, ?, ?, ?, ?);
        """,
        (str(event)[:50], (str(username)[:120] if username else None), user_id, ip, (str(ua)[:512] if ua else None), (str(detail)[:400] if detail else None)),
      )
      conn.commit()
  except Exception:
    return


def _audit_api_event(request: Request, user: Optional[Dict[str, Any]], request_id: str, status_code: int, duration_ms: int) -> None:
  if not _AUDIT_ENABLED:
    return
  try:
    _ensure_auth_schema_microservicios()
    ip = (request.client.host if request.client else None) or None
    ua = request.headers.get("user-agent") or None
    u = user or {}
    user_id = int(u.get("id") or 0) if str(u.get("id") or "").strip() else None
    username = (str(u.get("username") or "")[:120] or None) if u else None
    role = (str(u.get("role") or "")[:50] or None) if u else None
    method = str(request.method or "")[:10]
    path = str(request.url.path or "")[:200]
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        INSERT INTO dbo.AppApiAudit (RequestId, UserId, Username, Role, Ip, UserAgent, Method, Path, StatusCode, DurationMs)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        (
          (str(request_id)[:60] if request_id else None),
          user_id,
          username,
          role,
          ip,
          (str(ua)[:512] if ua else None),
          method,
          path,
          int(status_code),
          int(duration_ms),
        ),
      )
      conn.commit()
  except Exception:
    return


def _lock_key(*parts: Any) -> str:
  return ":".join([str(p) for p in parts if p is not None])[:200]


def _acquire_lock(lock_key: str, user: Dict[str, Any], ttl_seconds: int) -> Dict[str, Any]:
  _ensure_auth_schema_microservicios()
  owner_id = int(user.get("id") or 0) if str(user.get("id") or "").strip() else None
  owner_username = str(user.get("username") or "")[:120] or None
  ttl = int(ttl_seconds) if int(ttl_seconds) > 0 else 60
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute("SET NOCOUNT ON; SET XACT_ABORT ON;")
    cur.execute(
      """
      UPDATE dbo.AppLocks
      SET
        OwnerUserId = ?,
        OwnerUsername = ?,
        AcquiredAt = SYSUTCDATETIME(),
        ExpiresAt = DATEADD(SECOND, ?, SYSUTCDATETIME()),
        UpdatedAt = SYSUTCDATETIME()
      WHERE LockKey = ?
        AND ExpiresAt <= SYSUTCDATETIME();
      """,
      (owner_id, owner_username, ttl, lock_key),
    )
    if int(cur.rowcount or 0) > 0:
      conn.commit()
      return {"lockKey": lock_key, "ownerUserId": owner_id, "ownerUsername": owner_username, "ttlSeconds": ttl}

    try:
      cur.execute(
        """
        INSERT INTO dbo.AppLocks (LockKey, OwnerUserId, OwnerUsername, ExpiresAt)
        VALUES (?, ?, ?, DATEADD(SECOND, ?, SYSUTCDATETIME()));
        """,
        (lock_key, owner_id, owner_username, ttl),
      )
      conn.commit()
      return {"lockKey": lock_key, "ownerUserId": owner_id, "ownerUsername": owner_username, "ttlSeconds": ttl}
    except Exception:
      try:
        conn.rollback()
      except Exception:
        pass

    cur.execute(
      """
      SELECT TOP 1
        LockKey AS lockKey,
        OwnerUserId AS ownerUserId,
        OwnerUsername AS ownerUsername,
        ExpiresAt AS expiresAt
      FROM dbo.AppLocks
      WHERE LockKey = ?;
      """,
      (lock_key,),
    )
    rows = _rows(cur)
    info = rows[0] if rows else {"lockKey": lock_key}
    raise HTTPException(status_code=409, detail={"code": "E_LOCKED", "message": "Trámite en proceso", "lock": info})


def _release_lock(lock_key: str, user: Dict[str, Any]) -> None:
  try:
    _ensure_auth_schema_microservicios()
    owner_id = int(user.get("id") or 0) if str(user.get("id") or "").strip() else None
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        DELETE FROM dbo.AppLocks
        WHERE LockKey = ?
          AND (? IS NULL OR OwnerUserId = ?);
        """,
        (lock_key, owner_id, owner_id),
      )
      conn.commit()
  except Exception:
    return


def _idempotency_lookup(request: Request, user: Dict[str, Any], body: Any) -> Optional[ORJSONResponse]:
  key = (request.headers.get("Idempotency-Key") or "").strip()
  if not key or len(key) > 80:
    return None
  _ensure_auth_schema_microservicios()
  try:
    raw = orjson.dumps(body or {}, option=orjson.OPT_SORT_KEYS)
  except Exception:
    raw = orjson.dumps({})
  req_hash = _sha256_bytes(raw)
  path = str(request.url.path or "")[:200]
  user_id = int(user.get("id") or 0) if str(user.get("id") or "").strip() else 0
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      SELECT TOP 1 StatusCode, ResponseJson, RequestHash
      FROM dbo.AppIdempotency
      WHERE IdempotencyKey = ?
        AND Path = ?
        AND UserId = ?;
      """,
      (key, path, user_id),
    )
    rows = cur.fetchall()
    if not rows:
      return None
    status_code = int(rows[0][0] or 200)
    resp_json = str(rows[0][1] or "{}")
    stored_hash = bytes(rows[0][2]) if rows[0][2] is not None else None
    if stored_hash is not None and stored_hash != req_hash:
      raise HTTPException(status_code=409, detail="Idempotency-Key ya fue usado con otro payload")
    try:
      payload = orjson.loads(resp_json.encode("utf-8"))
    except Exception:
      payload = {"ok": False, "error": "Respuesta idempotente inválida"}
    return ORJSONResponse(status_code=status_code, content=payload)


def _idempotency_store(request: Request, user: Dict[str, Any], body: Any, response: ORJSONResponse) -> None:
  try:
    key = (request.headers.get("Idempotency-Key") or "").strip()
    if not key or len(key) > 80:
      return
    _ensure_auth_schema_microservicios()
    try:
      raw = orjson.dumps(body or {}, option=orjson.OPT_SORT_KEYS)
    except Exception:
      raw = orjson.dumps({})
    req_hash = _sha256_bytes(raw)
    path = str(request.url.path or "")[:200]
    user_id = int(user.get("id") or 0) if str(user.get("id") or "").strip() else 0
    status_code = int(getattr(response, "status_code", 200) or 200)
    resp_raw = bytes(getattr(response, "body", b"{}") or b"{}")
    resp_text = resp_raw.decode("utf-8", errors="replace")
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        IF NOT EXISTS (
          SELECT 1 FROM dbo.AppIdempotency
          WHERE IdempotencyKey = ?
            AND Path = ?
            AND UserId = ?
        )
        BEGIN
          INSERT INTO dbo.AppIdempotency (IdempotencyKey, Path, UserId, RequestHash, StatusCode, ResponseJson)
          VALUES (?, ?, ?, ?, ?, ?);
        END
        """,
        (key, path, user_id, key, path, user_id, req_hash, status_code, resp_text),
      )
      conn.commit()
  except Exception:
    return


def _get_current_user(request: Request) -> Optional[Dict[str, Any]]:
  token = request.cookies.get("ms_session") or ""
  token = str(token).strip()
  if not token:
    return None
  token_hash = _sha256_bytes(token)
  _ensure_auth_schema_microservicios()
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      SELECT TOP 1
        u.UserId AS id,
        u.Username AS username,
        COALESCE(u.DisplayName, u.Username) AS displayName,
        u.Role AS role
      FROM dbo.AppSessions s
      INNER JOIN dbo.AppUsers u ON u.UserId = s.UserId
      WHERE s.TokenHash = ?
        AND s.RevokedAt IS NULL
        AND s.ExpiresAt > SYSUTCDATETIME()
        AND u.IsActive = 1;
      """,
      (token_hash,),
    )
    rows = _rows(cur)
    return rows[0] if rows else None


def _role_allows_api_path(role: str, path: str) -> bool:
  r = str(role or "").strip().lower()
  p = str(path or "")

  if r == "admin":
    return True

  if r == "dir_ingresos":
    if p.startswith("/api/reportes/") or p.startswith("/api/analitica/") or p.startswith("/api/fuentes"):
      return True
    return False

  if r == "cajero":
    if p.startswith("/api/cajas/"):
      return True
    return False

  return False


@app.middleware("http")
async def _auth_guard_api(request: Request, call_next):
  path = request.url.path or ""
  if not path.startswith("/api/"):
    return await call_next(request)

  if path == "/api/health":
    return await call_next(request)

  if path.startswith("/api/auth/"):
    return await call_next(request)

  if path.startswith("/api/auth/admin/") or path == "/api/admin/bootstrap-microservicios":
    return await call_next(request)

  user = _get_current_user(request)
  if not user:
    return ORJSONResponse(status_code=401, content={"ok": False, "detail": "No autenticado"})

  if path in {"/api/activaciones", "/api/consolidar"} and _has_valid_admin_key(request):
    return await call_next(request)

  if not _role_allows_api_path(str(user.get("role") or ""), path):
    return ORJSONResponse(status_code=403, content={"ok": False, "detail": "No autorizado"})

  return await call_next(request)


@app.middleware("http")
async def _rate_limit_and_audit(request: Request, call_next):
  path = request.url.path or ""
  request_id = (request.headers.get("x-request-id") or "").strip() or uuid.uuid4().hex
  started = datetime.utcnow()

  if path.startswith("/api/"):
    if path.startswith("/api/reportes/") or path.startswith("/api/analitica/"):
      _check_ip_rate_limit(
        request,
        _REPORT_RATE_BUCKETS,
        _REPORT_RATE_LOCK,
        _REPORT_RATE_WINDOW_SECONDS,
        _REPORT_RATE_MAX,
        "report",
      )
    else:
      _check_ip_rate_limit(
        request,
        _API_RATE_BUCKETS,
        _API_RATE_LOCK,
        _API_RATE_WINDOW_SECONDS,
        _API_RATE_MAX,
        "api",
      )

  user = None
  try:
    if path.startswith("/api/") and not path.startswith("/api/auth/"):
      user = _get_current_user(request)
  except Exception:
    user = None

  status_code = 500
  try:
    resp = await call_next(request)
    status_code = int(getattr(resp, "status_code", 200) or 200)
    try:
      resp.headers.setdefault("X-Request-Id", request_id)
    except Exception:
      pass
    return resp
  except HTTPException as e:
    status_code = int(getattr(e, "status_code", 500) or 500)
    raise
  except Exception:
    status_code = 500
    raise
  finally:
    try:
      duration_ms = int((datetime.utcnow() - started).total_seconds() * 1000)
      _audit_api_event(request, user, request_id, status_code, duration_ms)
    except Exception:
      pass

@app.get("/api/auth/me")
def auth_me(request: Request) -> ORJSONResponse:
  user = _get_current_user(request)
  if not user:
    raise HTTPException(status_code=401, detail="No autenticado")
  return ORJSONResponse({"ok": True, "user": user})


@app.post("/api/auth/login")
async def auth_login(request: Request) -> ORJSONResponse:
  secret = _auth_secret()
  if not secret:
    raise HTTPException(status_code=503, detail="AUTH_SECRET no configurada en el servidor")

  _reject_if_too_large(request, 16384)
  _check_login_rate_limit(request)

  body = await request.json()
  username = str((body or {}).get("username") or "").strip()
  password = str((body or {}).get("password") or "")
  if not username or len(username) > 120:
    raise HTTPException(status_code=400, detail="Usuario inválido")
  if not password or len(password) > 256:
    raise HTTPException(status_code=400, detail="Contraseña inválida")

  _ensure_auth_schema_microservicios()
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      SELECT TOP 1
        UserId,
        Username,
        COALESCE(DisplayName, Username) AS DisplayName,
        Role,
        PasswordSalt,
        PasswordHash,
        PasswordIterations,
        IsActive,
        FailedAttempts,
        LockedUntil
      FROM dbo.AppUsers
      WHERE Username = ?;
      """,
      (username,),
    )
    rows = cur.fetchall()
    if not rows:
      try:
        cur.execute("SELECT COUNT(1) AS cnt FROM dbo.AppUsers;")
        cnt_row = cur.fetchone()
        total_users = int(cnt_row[0]) if cnt_row else 0
      except Exception:
        total_users = 0

      bootstrap_username = str(os.getenv("BOOTSTRAP_ADMIN_USERNAME") or "admin@tulum.gob.mx").strip()
      bootstrap_password = str(os.getenv("BOOTSTRAP_ADMIN_PASSWORD") or (os.getenv("ADMIN_KEY") or "")).strip()

      if total_users == 0:
        if bootstrap_username and username.lower() != bootstrap_username.lower():
          raise HTTPException(
            status_code=401,
            detail=f"Sistema sin usuarios. Inicia sesión con {bootstrap_username}.",
          )
        if bootstrap_username and bootstrap_password and username.lower() == bootstrap_username.lower() and password != bootstrap_password:
          raise HTTPException(
            status_code=401,
            detail="Sistema sin usuarios. Usa la ADMIN_KEY como contraseña temporal (o configura BOOTSTRAP_ADMIN_PASSWORD).",
          )

      if (
        total_users == 0
        and bootstrap_username
        and bootstrap_password
        and username.lower() == bootstrap_username.lower()
        and password == bootstrap_password
      ):
        salt = secrets.token_bytes(16)
        iterations = 210000
        pw_hash = _pbkdf2_hash(password, salt, iterations)
        try:
          cur.execute(
            """
            INSERT INTO dbo.AppUsers (Username, DisplayName, Role, PasswordSalt, PasswordHash, PasswordIterations, IsActive)
            VALUES (?, ?, 'admin', ?, ?, ?, 1);
            """,
            (bootstrap_username, "Administrador", salt, pw_hash, iterations),
          )
          conn.commit()
          _audit_auth_event("bootstrap_admin_created", bootstrap_username, None, request, None)
        except Exception:
          try:
            conn.rollback()
          except Exception:
            pass

        cur.execute(
          """
          SELECT TOP 1
            UserId,
            Username,
            COALESCE(DisplayName, Username) AS DisplayName,
            Role,
            PasswordSalt,
            PasswordHash,
            PasswordIterations,
            IsActive,
            FailedAttempts,
            LockedUntil
          FROM dbo.AppUsers
          WHERE Username = ?;
          """,
          (bootstrap_username,),
        )
        rows = cur.fetchall()

      if not rows:
        _audit_auth_event("login_fail", username, None, request, "user_not_found")
        raise HTTPException(status_code=401, detail="Credenciales inválidas")

    row = rows[0]
    user_id = int(row[0])
    is_active = bool(row[7])
    failed_attempts = int(row[8] or 0)
    locked_until = row[9]

    if not is_active:
      _audit_auth_event("login_fail", username, user_id, request, "inactive")
      raise HTTPException(status_code=403, detail="Usuario inactivo")

    if locked_until and isinstance(locked_until, datetime) and locked_until > datetime.utcnow():
      _audit_auth_event("login_fail", username, user_id, request, "locked")
      raise HTTPException(status_code=403, detail="Usuario bloqueado temporalmente")

    salt = bytes(row[4])
    stored_hash = bytes(row[5])
    iterations = int(row[6] or 0)
    if iterations < 10000:
      iterations = 210000
    computed = _pbkdf2_hash(password, salt, iterations)

    if computed != stored_hash:
      failed_attempts += 1
      lock_until = None
      if failed_attempts >= 5:
        lock_until = datetime.utcnow() + timedelta(minutes=10)
      cur.execute(
        """
        UPDATE dbo.AppUsers
        SET FailedAttempts = ?, LockedUntil = ?, UpdatedAt = SYSUTCDATETIME()
        WHERE UserId = ?;
        """,
        (failed_attempts, lock_until, user_id),
      )
      conn.commit()
      _audit_auth_event("login_fail", username, user_id, request, "bad_password")
      raise HTTPException(status_code=401, detail="Credenciales inválidas")

    cur.execute(
      """
      UPDATE dbo.AppUsers
      SET FailedAttempts = 0, LockedUntil = NULL, UpdatedAt = SYSUTCDATETIME()
      WHERE UserId = ?;
      """,
      (user_id,),
    )

    token = secrets.token_urlsafe(48)
    token_hash = _sha256_bytes(token)
    expires_at = datetime.utcnow() + timedelta(hours=12)
    ip = (request.client.host if request.client else None) or None
    ua = request.headers.get("user-agent") or None
    cur.execute(
      """
      INSERT INTO dbo.AppSessions (TokenHash, UserId, ExpiresAt, Ip, UserAgent)
      VALUES (?, ?, ?, ?, ?);
      """,
      (token_hash, user_id, expires_at, ip, (str(ua)[:512] if ua else None)),
    )
    conn.commit()

    user = {"id": user_id, "username": str(row[1]), "displayName": str(row[2]), "role": str(row[3])}
    _audit_auth_event("login_ok", username, user_id, request, None)

  resp = ORJSONResponse({"ok": True, "user": user})
  resp.set_cookie(
    key="ms_session",
    value=token,
    httponly=True,
    samesite="strict",
    secure=_is_secure_request(request),
    max_age=12 * 60 * 60,
    path="/",
  )
  return resp


@app.post("/api/auth/logout")
def auth_logout(request: Request) -> ORJSONResponse:
  token = request.cookies.get("ms_session") or ""
  token = str(token).strip()
  if token:
    try:
      _ensure_auth_schema_microservicios()
      token_hash = _sha256_bytes(token)
      with get_conn_for_database("MicroServicios") as conn:
        cur = conn.cursor()
        cur.execute(
          """
          UPDATE dbo.AppSessions
          SET RevokedAt = SYSUTCDATETIME()
          WHERE TokenHash = ? AND RevokedAt IS NULL;
          """,
          (token_hash,),
        )
        conn.commit()
    except Exception:
      pass

  resp = ORJSONResponse({"ok": True})
  resp.delete_cookie("ms_session", path="/")
  return resp


def _require_login(request: Request) -> Dict[str, Any]:
  user = _get_current_user(request)
  if not user:
    raise HTTPException(status_code=401, detail="No autenticado")
  return user


def _require_role(user: Dict[str, Any], role: str) -> None:
  current = str(user.get("role") or "").strip().lower()
  if current != str(role).strip().lower():
    raise HTTPException(status_code=403, detail="No autorizado")


@app.get("/api/users")
def list_users(request: Request) -> ORJSONResponse:
  user = _require_login(request)
  _require_role(user, "admin")
  _ensure_auth_schema_microservicios()
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      SELECT
        UserId AS id,
        Username AS username,
        COALESCE(DisplayName, Username) AS displayName,
        Role AS role,
        IsActive AS isActive,
        CreatedAt AS createdAt,
        UpdatedAt AS updatedAt
      FROM dbo.AppUsers
      ORDER BY Username ASC;
      """
    )
    return ORJSONResponse({"ok": True, "rows": _rows(cur)})


@app.post("/api/users")
async def create_user(request: Request) -> ORJSONResponse:
  user = _require_login(request)
  _require_role(user, "admin")
  body = await request.json()
  username = str((body or {}).get("username") or "").strip()
  password = str((body or {}).get("password") or "")
  display_name = str((body or {}).get("displayName") or "").strip() or None
  role = str((body or {}).get("role") or "cajero").strip().lower()
  is_active = bool((body or {}).get("isActive", True))

  if not username or len(username) > 120:
    raise HTTPException(status_code=400, detail="Usuario inválido")
  if not password or len(password) > 256:
    raise HTTPException(status_code=400, detail="Contraseña inválida")
  if role not in {"admin", "cajero", "dir_ingresos"}:
    raise HTTPException(status_code=400, detail="Rol inválido")

  _ensure_auth_schema_microservicios()
  salt = secrets.token_bytes(16)
  iterations = 210000
  pw_hash = _pbkdf2_hash(password, salt, iterations)

  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      INSERT INTO dbo.AppUsers (Username, DisplayName, Role, PasswordSalt, PasswordHash, PasswordIterations, IsActive)
      VALUES (?, ?, ?, ?, ?, ?, ?);
      """,
      (username, display_name, role, salt, pw_hash, iterations, 1 if is_active else 0),
    )
    conn.commit()

  _audit_auth_event("user_create", username, None, request, f"by={user.get('username')}")
  return ORJSONResponse({"ok": True})


@app.get("/api/auth/admin/users")
def auth_admin_list_users(x_admin_key: Optional[str] = Header(default=None)) -> ORJSONResponse:
  _require_admin(x_admin_key)
  _ensure_auth_schema_microservicios()
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      SELECT
        UserId AS id,
        Username AS username,
        COALESCE(DisplayName, Username) AS displayName,
        Role AS role,
        IsActive AS isActive,
        CreatedAt AS createdAt,
        UpdatedAt AS updatedAt
      FROM dbo.AppUsers
      ORDER BY Username ASC;
      """
    )
    return ORJSONResponse({"ok": True, "rows": _rows(cur)})


@app.post("/api/auth/admin/users")
async def auth_admin_create_user(request: Request, x_admin_key: Optional[str] = Header(default=None)) -> ORJSONResponse:
  _require_admin(x_admin_key)
  body = await request.json()
  username = str((body or {}).get("username") or "").strip()
  password = str((body or {}).get("password") or "")
  display_name = str((body or {}).get("displayName") or "").strip() or None
  role = str((body or {}).get("role") or "cajero").strip().lower()
  is_active = bool((body or {}).get("isActive", True))

  if not username or len(username) > 120:
    raise HTTPException(status_code=400, detail="Usuario inválido")
  if not password or len(password) > 256:
    raise HTTPException(status_code=400, detail="Contraseña inválida")
  if role not in {"admin", "cajero", "dir_ingresos"}:
    raise HTTPException(status_code=400, detail="Rol inválido")

  _ensure_auth_schema_microservicios()
  salt = secrets.token_bytes(16)
  iterations = 210000
  pw_hash = _pbkdf2_hash(password, salt, iterations)

  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      INSERT INTO dbo.AppUsers (Username, DisplayName, Role, PasswordSalt, PasswordHash, PasswordIterations, IsActive)
      VALUES (?, ?, ?, ?, ?, ?, ?);
      """,
      (username, display_name, role, salt, pw_hash, iterations, 1 if is_active else 0),
    )
    conn.commit()

  return ORJSONResponse({"ok": True})


@app.post("/api/auth/admin/reset-password")
async def auth_admin_reset_password(request: Request, x_admin_key: Optional[str] = Header(default=None)) -> ORJSONResponse:
  _require_admin(x_admin_key)
  body = await request.json()
  username = str((body or {}).get("username") or "").strip()
  password = str((body or {}).get("password") or "")

  if not username or len(username) > 120:
    raise HTTPException(status_code=400, detail="Usuario inválido")
  if not password or len(password) > 256:
    raise HTTPException(status_code=400, detail="Contraseña inválida")

  _ensure_auth_schema_microservicios()
  salt = secrets.token_bytes(16)
  iterations = 210000
  pw_hash = _pbkdf2_hash(password, salt, iterations)

  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      UPDATE dbo.AppUsers
      SET PasswordSalt = ?,
          PasswordHash = ?,
          PasswordIterations = ?,
          FailedAttempts = 0,
          LockedUntil = NULL,
          UpdatedAt = SYSUTCDATETIME()
      WHERE Username = ?;
      """,
      (salt, pw_hash, iterations, username),
    )
    if cur.rowcount <= 0:
      conn.rollback()
      raise HTTPException(status_code=404, detail="Usuario no encontrado")
    conn.commit()

  return ORJSONResponse({"ok": True})

 
@app.get("/api/health")
def health() -> Dict[str, Any]:
  return {"ok": True}
 
 
@app.get("/api/test-connection")
def test_connection() -> ORJSONResponse:
  try:
    cfg = get_db_target()
    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        """
        SELECT
          DB_NAME() AS databaseName,
          @@SERVERNAME AS serverName,
          SUSER_SNAME() AS loginName,
          GETDATE() AS serverTime
        """
      )
      row = _rows(cur)[0] if cur.description else {}
    return ORJSONResponse(
        {
          "ok": True,
          "message": "Conexion exitosa",
          "target": {
            "server": cfg["server"],
            "port": cfg["port"],
            "database": cfg["database"],
            "encrypt": bool(cfg["encrypt"]),
            "trustServerCertificate": bool(cfg["trustServerCertificate"]),
            "serverName": cfg["serverName"],
          },
          "connection": row,
        }
    )
  except Exception as e:
    return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/api/test-connection/microservicios")
def test_connection_microservicios() -> ORJSONResponse:
  try:
    cfg = get_db_target()
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        SELECT
          DB_NAME() AS databaseName,
          @@SERVERNAME AS serverName,
          SUSER_SNAME() AS loginName,
          GETDATE() AS serverTime
        """
      )
      row = _rows(cur)[0] if cur.description else {}
    return ORJSONResponse(
        {
          "ok": True,
          "message": "Conexion exitosa",
          "target": {
            "server": cfg["server"],
            "port": cfg["port"],
            "database": "MicroServicios",
            "encrypt": bool(cfg["encrypt"]),
            "trustServerCertificate": bool(cfg["trustServerCertificate"]),
            "serverName": cfg["serverName"],
          },
          "connection": row,
        }
    )
  except Exception as e:
    return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})


def _validate_sqlserver_db_name(value: str) -> str:
  name = (value or "").strip()
  if not re.fullmatch(r"[A-Za-z][A-Za-z0-9_]{0,127}", name):
    raise ValueError("Nombre de base de datos inválido")
  return name


def _bootstrap_cri_schema_in_database(database: str) -> Dict[str, Any]:
  ddl: List[str] = []
  ddl.append(
    """
    IF NOT EXISTS (SELECT 1 FROM sys.schemas WHERE name = 'cri')
    BEGIN
      EXEC('CREATE SCHEMA cri AUTHORIZATION dbo;');
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('cri.CRI_Rubro', 'U') IS NULL
    BEGIN
      CREATE TABLE cri.CRI_Rubro (
        id_rubro        INT IDENTITY(1,1) NOT NULL CONSTRAINT PK_CRI_Rubro PRIMARY KEY,
        clave_rubro     CHAR(1) NOT NULL,
        nombre_rubro    NVARCHAR(200) NOT NULL,
        descripcion     NVARCHAR(500) NULL,
        CONSTRAINT UQ_CRI_Rubro_clave UNIQUE (clave_rubro),
        CONSTRAINT CK_CRI_Rubro_clave_digito CHECK (clave_rubro LIKE '[0-9]')
      );
      CREATE INDEX IX_CRI_Rubro_nombre ON cri.CRI_Rubro(nombre_rubro);
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('cri.CRI_Tipo', 'U') IS NULL
    BEGIN
      CREATE TABLE cri.CRI_Tipo (
        id_tipo       INT IDENTITY(1,1) NOT NULL CONSTRAINT PK_CRI_Tipo PRIMARY KEY,
        clave_tipo    CHAR(2) NOT NULL,
        nombre_tipo   NVARCHAR(250) NOT NULL,
        id_rubro      INT NOT NULL,
        CONSTRAINT UQ_CRI_Tipo_clave UNIQUE (clave_tipo),
        CONSTRAINT CK_CRI_Tipo_clave_dos_digitos CHECK (clave_tipo LIKE '[0-9][0-9]'),
        CONSTRAINT FK_CRI_Tipo_Rubro FOREIGN KEY (id_rubro) REFERENCES cri.CRI_Rubro(id_rubro)
      );
      CREATE INDEX IX_CRI_Tipo_id_rubro ON cri.CRI_Tipo(id_rubro);
      CREATE INDEX IX_CRI_Tipo_nombre ON cri.CRI_Tipo(nombre_tipo);
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('cri.CRI_Clase', 'U') IS NULL
    BEGIN
      CREATE TABLE cri.CRI_Clase (
        id_clase      INT IDENTITY(1,1) NOT NULL CONSTRAINT PK_CRI_Clase PRIMARY KEY,
        clave_clase   CHAR(2) NOT NULL,
        nombre_clase  NVARCHAR(250) NOT NULL,
        id_tipo       INT NOT NULL,
        CONSTRAINT UQ_CRI_Clase_por_tipo UNIQUE (id_tipo, clave_clase),
        CONSTRAINT CK_CRI_Clase_clave_dos_digitos CHECK (clave_clase LIKE '[0-9][0-9]'),
        CONSTRAINT FK_CRI_Clase_Tipo FOREIGN KEY (id_tipo) REFERENCES cri.CRI_Tipo(id_tipo)
      );
      CREATE INDEX IX_CRI_Clase_id_tipo ON cri.CRI_Clase(id_tipo);
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('cri.CRI_Concepto', 'U') IS NULL
    BEGIN
      CREATE TABLE cri.CRI_Concepto (
        id_concepto      INT IDENTITY(1,1) NOT NULL CONSTRAINT PK_CRI_Concepto PRIMARY KEY,
        clave_concepto   CHAR(2) NOT NULL,
        nombre_concepto  NVARCHAR(300) NOT NULL,
        id_clase         INT NOT NULL,
        CONSTRAINT UQ_CRI_Concepto_por_clase UNIQUE (id_clase, clave_concepto),
        CONSTRAINT CK_CRI_Concepto_clave_dos_digitos CHECK (clave_concepto LIKE '[0-9][0-9]'),
        CONSTRAINT FK_CRI_Concepto_Clase FOREIGN KEY (id_clase) REFERENCES cri.CRI_Clase(id_clase)
      );
      CREATE INDEX IX_CRI_Concepto_id_clase ON cri.CRI_Concepto(id_clase);
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('cri.Ente_Publico', 'U') IS NULL
    BEGIN
      CREATE TABLE cri.Ente_Publico (
        id_ente          INT IDENTITY(1,1) NOT NULL CONSTRAINT PK_Ente_Publico PRIMARY KEY,
        nombre           NVARCHAR(250) NOT NULL,
        nivel_gobierno   VARCHAR(10) NOT NULL,
        rfc              VARCHAR(13) NULL,
        ejercicio_fiscal SMALLINT NOT NULL,
        CONSTRAINT CK_Ente_nivel CHECK (nivel_gobierno IN ('Federal','Estatal','Municipal')),
        CONSTRAINT CK_Ente_ejercicio CHECK (ejercicio_fiscal BETWEEN 2000 AND 2100),
        CONSTRAINT UQ_Ente_rfc_ejercicio UNIQUE (rfc, ejercicio_fiscal)
      );
      CREATE INDEX IX_Ente_nombre ON cri.Ente_Publico(nombre);
      CREATE INDEX IX_Ente_nivel_ejercicio ON cri.Ente_Publico(nivel_gobierno, ejercicio_fiscal);
    END
    """
  )
  ddl.append(
    """
    IF OBJECT_ID('cri.Ingreso_Presupuestal', 'U') IS NULL
    BEGIN
      CREATE TABLE cri.Ingreso_Presupuestal (
        id_ingreso            BIGINT IDENTITY(1,1) NOT NULL CONSTRAINT PK_Ingreso_Presupuestal PRIMARY KEY,
        id_ente               INT NOT NULL,
        id_concepto           INT NOT NULL,
        ejercicio_fiscal      SMALLINT NOT NULL,
        periodo               TINYINT NOT NULL,
        fecha_registro        DATETIME2(0) NOT NULL CONSTRAINT DF_Ingreso_fecha DEFAULT (SYSUTCDATETIME()),
        usuario_captura       NVARCHAR(100) NOT NULL,
        estimado_ley_ingresos DECIMAL(18,2) NOT NULL CONSTRAINT DF_Ingreso_estimado DEFAULT (0),
        modificaciones        DECIMAL(18,2) NOT NULL CONSTRAINT DF_Ingreso_modif DEFAULT (0),
        devengado             DECIMAL(18,2) NOT NULL CONSTRAINT DF_Ingreso_dev DEFAULT (0),
        recaudado             DECIMAL(18,2) NOT NULL CONSTRAINT DF_Ingreso_rec DEFAULT (0),
        modificado AS (estimado_ley_ingresos + modificaciones) PERSISTED,
        diferencia AS (recaudado - estimado_ley_ingresos) PERSISTED,
        CONSTRAINT CK_Ingreso_ejercicio CHECK (ejercicio_fiscal BETWEEN 2000 AND 2100),
        CONSTRAINT CK_Ingreso_periodo CHECK (periodo BETWEEN 1 AND 12),
        CONSTRAINT FK_Ingreso_Ente FOREIGN KEY (id_ente) REFERENCES cri.Ente_Publico(id_ente),
        CONSTRAINT FK_Ingreso_Concepto FOREIGN KEY (id_concepto) REFERENCES cri.CRI_Concepto(id_concepto)
      );
      CREATE INDEX IX_Ingreso_Filtros ON cri.Ingreso_Presupuestal(ejercicio_fiscal, periodo, id_ente) INCLUDE (id_concepto, estimado_ley_ingresos, modificaciones, devengado, recaudado);
      CREATE INDEX IX_Ingreso_Concepto ON cri.Ingreso_Presupuestal(id_concepto);
    END
    """
  )
  ddl.append(
    """
    EXEC(N'CREATE OR ALTER VIEW cri.vw_Estado_Analitico_Ingresos AS
    SELECT
      e.id_ente,
      e.nombre AS Ente_Publico,
      e.nivel_gobierno AS Nivel_Gobierno,
      i.ejercicio_fiscal,
      i.periodo,
      r.clave_rubro,
      r.nombre_rubro,
      t.clave_tipo,
      t.nombre_tipo,
      CAST(SUM(i.estimado_ley_ingresos) AS DECIMAL(18,2)) AS Estimado,
      CAST(SUM(i.modificado) AS DECIMAL(18,2)) AS Modificado,
      CAST(SUM(i.devengado) AS DECIMAL(18,2)) AS Devengado,
      CAST(SUM(i.recaudado) AS DECIMAL(18,2)) AS Recaudado,
      CAST(
        CASE WHEN SUM(i.estimado_ley_ingresos) = 0
          THEN NULL
          ELSE ((SUM(i.recaudado) - SUM(i.estimado_ley_ingresos)) / NULLIF(SUM(i.estimado_ley_ingresos), 0)) * 100
        END
      AS DECIMAL(18,2)) AS [Diferencia (%)]
    FROM cri.Ingreso_Presupuestal i
    INNER JOIN cri.Ente_Publico e ON e.id_ente = i.id_ente
    INNER JOIN cri.CRI_Concepto co ON co.id_concepto = i.id_concepto
    INNER JOIN cri.CRI_Clase cl ON cl.id_clase = co.id_clase
    INNER JOIN cri.CRI_Tipo t ON t.id_tipo = cl.id_tipo
    INNER JOIN cri.CRI_Rubro r ON r.id_rubro = t.id_rubro
    GROUP BY
      e.id_ente, e.nombre, e.nivel_gobierno, i.ejercicio_fiscal, i.periodo,
      r.clave_rubro, r.nombre_rubro, t.clave_tipo, t.nombre_tipo;');
    """
  )
  ddl.append(
    """
    EXEC(N'CREATE OR ALTER VIEW cri.vw_Resumen_por_Rubro AS
    SELECT
      r.clave_rubro,
      r.nombre_rubro,
      i.ejercicio_fiscal,
      i.periodo,
      CAST(SUM(i.estimado_ley_ingresos) AS DECIMAL(18,2)) AS Estimado,
      CAST(SUM(i.modificado) AS DECIMAL(18,2)) AS Modificado,
      CAST(SUM(i.devengado) AS DECIMAL(18,2)) AS Devengado,
      CAST(SUM(i.recaudado) AS DECIMAL(18,2)) AS Recaudado,
      CAST(SUM(i.diferencia) AS DECIMAL(18,2)) AS Diferencia_Monto
    FROM cri.Ingreso_Presupuestal i
    INNER JOIN cri.CRI_Concepto co ON co.id_concepto = i.id_concepto
    INNER JOIN cri.CRI_Clase cl ON cl.id_clase = co.id_clase
    INNER JOIN cri.CRI_Tipo t ON t.id_tipo = cl.id_tipo
    INNER JOIN cri.CRI_Rubro r ON r.id_rubro = t.id_rubro
    GROUP BY
      r.clave_rubro, r.nombre_rubro, i.ejercicio_fiscal, i.periodo;');
    """
  )
  ddl.append(
    """
    MERGE cri.CRI_Rubro AS tgt
    USING (VALUES
      ('1', N'IMPUESTOS', N'Rubro 1 — Impuestos'),
      ('2', N'CUOTAS Y APORTACIONES DE SEGURIDAD SOCIAL', N'Rubro 2 — Cuotas y Aportaciones de Seguridad Social'),
      ('3', N'CONTRIBUCIONES DE MEJORAS', N'Rubro 3 — Contribuciones de mejoras'),
      ('4', N'DERECHOS', N'Rubro 4 — Derechos'),
      ('5', N'PRODUCTOS', N'Rubro 5 — Productos'),
      ('6', N'APROVECHAMIENTOS', N'Rubro 6 — Aprovechamientos'),
      ('7', N'INGRESOS POR VENTAS DE BIENES Y SERVICIOS', N'Rubro 7 — Ventas de bienes y servicios'),
      ('8', N'PARTICIPACIONES Y APORTACIONES', N'Rubro 8 — Participaciones y aportaciones'),
      ('9', N'TRANSFERENCIAS, ASIGNACIONES, SUBSIDIOS Y OTRAS AYUDAS', N'Rubro 9 — Transferencias y otras ayudas'),
      ('0', N'INGRESOS DERIVADOS DE FINANCIAMIENTOS', N'Rubro 0 — Financiamientos')
    ) AS src (clave_rubro, nombre_rubro, descripcion)
    ON tgt.clave_rubro = src.clave_rubro
    WHEN MATCHED THEN
      UPDATE SET nombre_rubro = src.nombre_rubro, descripcion = src.descripcion
    WHEN NOT MATCHED THEN
      INSERT (clave_rubro, nombre_rubro, descripcion)
      VALUES (src.clave_rubro, src.nombre_rubro, src.descripcion);
    """
  )
  ddl.append(
    """
    MERGE cri.CRI_Tipo AS tgt
    USING (
      SELECT
        v.clave_tipo,
        v.nombre_tipo,
        r.id_rubro
      FROM (VALUES
        ('11', N'Impuestos sobre los ingresos', '1'),
        ('12', N'Impuestos sobre el patrimonio', '1'),
        ('13', N'Impuestos sobre la producción, el consumo y las transacciones', '1'),
        ('14', N'Impuestos al comercio exterior', '1'),
        ('15', N'Impuestos sobre Nóminas y Asimilables', '1'),
        ('16', N'Impuestos Ecológicos', '1'),
        ('17', N'Accesorios', '1'),
        ('18', N'Otros Impuestos', '1'),
        ('19', N'Causados en ejercicios anteriores pendientes de pago', '1'),
        ('21', N'Aportaciones para Fondos de Vivienda (INFONAVIT)', '2'),
        ('22', N'Cuotas para el Seguro Social (IMSS)', '2'),
        ('23', N'Cuotas de Ahorro para el Retiro (SAR/AFORE)', '2'),
        ('24', N'Otras Cuotas y Aportaciones para la seguridad social', '2'),
        ('25', N'Accesorios', '2'),
        ('31', N'Contribución de mejoras por obras públicas', '3'),
        ('39', N'Causadas en ejercicios anteriores pendientes de pago', '3'),
        ('41', N'Derechos por uso, goce, aprovechamiento de bienes de dominio público', '4'),
        ('42', N'Derechos a los hidrocarburos', '4'),
        ('43', N'Derechos por prestación de servicios', '4'),
        ('44', N'Otros Derechos', '4'),
        ('45', N'Accesorios', '4'),
        ('49', N'Causados en ejercicios anteriores pendientes de pago', '4'),
        ('51', N'Productos de tipo corriente', '5'),
        ('52', N'Productos de capital', '5'),
        ('59', N'Causados en ejercicios anteriores pendientes de pago', '5'),
        ('61', N'Aprovechamientos de tipo corriente', '6'),
        ('62', N'Aprovechamientos de capital', '6'),
        ('69', N'Causados en ejercicios anteriores pendientes de pago', '6'),
        ('71', N'Ingresos por ventas de bienes/servicios de organismos descentralizados', '7'),
        ('72', N'Ingresos de operación de entidades paraestatales empresariales', '7'),
        ('73', N'Ingresos por ventas de establecimientos del Gobierno Central', '7'),
        ('81', N'Participaciones', '8'),
        ('82', N'Aportaciones', '8'),
        ('83', N'Convenios', '8'),
        ('91', N'Transferencias Internas y Asignaciones al Sector Público', '9'),
        ('92', N'Transferencias al Resto del Sector Público', '9'),
        ('93', N'Subsidios y Subvenciones', '9'),
        ('94', N'Ayudas sociales', '9'),
        ('95', N'Pensiones y Jubilaciones', '9'),
        ('96', N'Transferencias a Fideicomisos, mandatos y análogos', '9'),
        ('01', N'Endeudamiento interno', '0'),
        ('02', N'Endeudamiento externo', '0')
      ) AS v(clave_tipo, nombre_tipo, clave_rubro)
      INNER JOIN cri.CRI_Rubro r ON r.clave_rubro = v.clave_rubro
    ) AS src (clave_tipo, nombre_tipo, id_rubro)
    ON tgt.clave_tipo = src.clave_tipo
    WHEN MATCHED THEN
      UPDATE SET nombre_tipo = src.nombre_tipo, id_rubro = src.id_rubro
    WHEN NOT MATCHED THEN
      INSERT (clave_tipo, nombre_tipo, id_rubro)
      VALUES (src.clave_tipo, src.nombre_tipo, src.id_rubro);
    """
  )
  ddl.append(
    """
    MERGE cri.CRI_Clase AS tgt
    USING (
      SELECT
        t.id_tipo,
        '01' AS clave_clase,
        CONCAT(N'Clase 01 del tipo ', RTRIM(t.clave_tipo), N' — ', RTRIM(t.nombre_tipo)) AS nombre_clase
      FROM cri.CRI_Tipo t
    ) AS src (id_tipo, clave_clase, nombre_clase)
    ON tgt.id_tipo = src.id_tipo
   AND tgt.clave_clase = src.clave_clase
    WHEN MATCHED THEN
      UPDATE SET nombre_clase = src.nombre_clase
    WHEN NOT MATCHED THEN
      INSERT (clave_clase, nombre_clase, id_tipo)
      VALUES (src.clave_clase, src.nombre_clase, src.id_tipo);
    """
  )
  ddl.append(
    """
    MERGE cri.CRI_Concepto AS tgt
    USING (
      SELECT
        c.id_clase,
        '01' AS clave_concepto,
        CONCAT(
          N'Concepto 01 de clase ',
          RTRIM(c.clave_clase),
          N' (tipo ',
          RTRIM(t.clave_tipo),
          N') — ',
          RTRIM(t.nombre_tipo)
        ) AS nombre_concepto
      FROM cri.CRI_Clase c
      INNER JOIN cri.CRI_Tipo t
        ON t.id_tipo = c.id_tipo
    ) AS src (id_clase, clave_concepto, nombre_concepto)
    ON tgt.id_clase = src.id_clase
   AND tgt.clave_concepto = src.clave_concepto
    WHEN MATCHED THEN
      UPDATE SET nombre_concepto = src.nombre_concepto
    WHEN NOT MATCHED THEN
      INSERT (clave_concepto, nombre_concepto, id_clase)
      VALUES (src.clave_concepto, src.nombre_concepto, src.id_clase);
    """
  )
  ddl.append(
    """
    IF NOT EXISTS (
      SELECT 1
      FROM sys.indexes
      WHERE object_id = OBJECT_ID('cri.Ingreso_Presupuestal')
        AND name = 'UQ_Ingreso_EnteConceptoPeriodo'
    )
    BEGIN
      CREATE UNIQUE INDEX UQ_Ingreso_EnteConceptoPeriodo
      ON cri.Ingreso_Presupuestal(id_ente, id_concepto, ejercicio_fiscal, periodo);
    END
    """
  )
  ddl.append(
    """
    MERGE cri.Ente_Publico AS tgt
    USING (VALUES
      (N'Municipio de Tulum', 'Municipal', 'XAXX010101000', CAST(YEAR(GETDATE()) AS smallint))
    ) AS src (nombre, nivel_gobierno, rfc, ejercicio_fiscal)
    ON tgt.rfc = src.rfc
   AND tgt.ejercicio_fiscal = src.ejercicio_fiscal
    WHEN MATCHED THEN
      UPDATE SET nombre = src.nombre, nivel_gobierno = src.nivel_gobierno
    WHEN NOT MATCHED THEN
      INSERT (nombre, nivel_gobierno, rfc, ejercicio_fiscal)
      VALUES (src.nombre, src.nivel_gobierno, src.rfc, src.ejercicio_fiscal);
    """
  )
  ddl.append(
    """
    ;WITH Ente AS (
      SELECT TOP 1
        e.id_ente,
        e.ejercicio_fiscal
      FROM cri.Ente_Publico e
      WHERE e.rfc = 'XAXX010101000'
        AND e.ejercicio_fiscal = CAST(YEAR(GETDATE()) AS smallint)
      ORDER BY e.id_ente ASC
    ),
    Conceptos AS (
      SELECT
        co.id_concepto,
        CAST(1000 + (co.id_concepto % 250) AS decimal(18,2)) AS estimado,
        CAST(50 + (co.id_concepto % 25) AS decimal(18,2)) AS modificaciones,
        CAST(800 + (co.id_concepto % 200) AS decimal(18,2)) AS devengado,
        CAST(750 + (co.id_concepto % 180) AS decimal(18,2)) AS recaudado
      FROM cri.CRI_Concepto co
    )
    MERGE cri.Ingreso_Presupuestal AS tgt
    USING (
      SELECT
        e.id_ente,
        c.id_concepto,
        e.ejercicio_fiscal AS ejercicio_fiscal,
        CAST(1 AS tinyint) AS periodo,
        CAST(N'bootstrap' AS nvarchar(100)) AS usuario_captura,
        c.estimado,
        c.modificaciones,
        c.devengado,
        c.recaudado
      FROM Ente e
      CROSS JOIN Conceptos c
    ) AS src (id_ente, id_concepto, ejercicio_fiscal, periodo, usuario_captura, estimado_ley_ingresos, modificaciones, devengado, recaudado)
    ON tgt.id_ente = src.id_ente
   AND tgt.id_concepto = src.id_concepto
   AND tgt.ejercicio_fiscal = src.ejercicio_fiscal
   AND tgt.periodo = src.periodo
    WHEN MATCHED THEN
      UPDATE SET
        usuario_captura = src.usuario_captura,
        estimado_ley_ingresos = src.estimado_ley_ingresos,
        modificaciones = src.modificaciones,
        devengado = src.devengado,
        recaudado = src.recaudado
    WHEN NOT MATCHED THEN
      INSERT (id_ente, id_concepto, ejercicio_fiscal, periodo, usuario_captura, estimado_ley_ingresos, modificaciones, devengado, recaudado)
      VALUES (src.id_ente, src.id_concepto, src.ejercicio_fiscal, src.periodo, src.usuario_captura, src.estimado_ley_ingresos, src.modificaciones, src.devengado, src.recaudado);
    """
  )

  conn = get_conn_for_database(database)
  try:
    cur = conn.cursor()
    cur.execute("SET NOCOUNT ON; SET XACT_ABORT ON;")
    for stmt in ddl:
      cur.execute(stmt)
    conn.commit()
  except Exception:
    conn.rollback()
    raise
  finally:
    conn.close()

  with get_conn_for_database(database) as check_conn:
    cur = check_conn.cursor()
    cur.execute(
      """
      SELECT s.name AS schemaName, o.name AS objectName, o.type_desc AS objectType
      FROM sys.objects o
      INNER JOIN sys.schemas s ON s.schema_id = o.schema_id
      WHERE s.name = 'cri'
        AND o.type IN ('U','V')
      ORDER BY o.type_desc, o.name;
      """
    )
    return {"database": database, "objects": _rows(cur)}


@app.post("/api/admin/bootstrap-microservicios")
async def bootstrap_microservicios(x_admin_key: Optional[str] = Header(default=None)) -> ORJSONResponse:
  _require_admin(x_admin_key)
  try:
    db_name = _validate_sqlserver_db_name("MicroServicios")
    conn = get_conn_for_database("master")
    created_db = False
    try:
      conn.autocommit = True
      cur = conn.cursor()
      cur.execute("SELECT DB_ID(?) AS dbid;", (db_name,))
      row = _rows(cur)[0] if cur.description else {}
      exists = row.get("dbid") is not None
      if not exists:
        cur.execute(f"CREATE DATABASE [{db_name}];")
        created_db = True
    finally:
      conn.close()

    applied = _bootstrap_cri_schema_in_database(db_name)
    return ORJSONResponse({"ok": True, "database": db_name, "createdDatabase": created_db, "applied": applied})
  except Exception as e:
    return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/api/cri/catalogo")
def cri_catalogo() -> ORJSONResponse:
  try:
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        SELECT
          id_rubro,
          RTRIM(clave_rubro) AS clave_rubro,
          RTRIM(nombre_rubro) AS nombre_rubro,
          RTRIM(COALESCE(descripcion, '')) AS descripcion
        FROM cri.CRI_Rubro
        ORDER BY clave_rubro ASC;
        """
      )
      rubros = _rows(cur)

      cur.execute(
        """
        SELECT
          t.id_tipo,
          RTRIM(t.clave_tipo) AS clave_tipo,
          RTRIM(t.nombre_tipo) AS nombre_tipo,
          t.id_rubro,
          RTRIM(r.clave_rubro) AS clave_rubro
        FROM cri.CRI_Tipo t
        INNER JOIN cri.CRI_Rubro r
          ON r.id_rubro = t.id_rubro
        ORDER BY r.clave_rubro ASC, t.clave_tipo ASC;
        """
      )
      tipos = _rows(cur)

      cur.execute("SELECT COUNT(*) AS cnt FROM cri.CRI_Clase;")
      clases_cnt = _rows(cur)[0]["cnt"]
      cur.execute("SELECT COUNT(*) AS cnt FROM cri.CRI_Concepto;")
      conceptos_cnt = _rows(cur)[0]["cnt"]

      cur.execute(
        """
        SELECT
          c.id_clase,
          RTRIM(c.clave_clase) AS clave_clase,
          RTRIM(c.nombre_clase) AS nombre_clase,
          c.id_tipo,
          RTRIM(t.clave_tipo) AS clave_tipo,
          t.id_rubro,
          RTRIM(r.clave_rubro) AS clave_rubro
        FROM cri.CRI_Clase c
        INNER JOIN cri.CRI_Tipo t
          ON t.id_tipo = c.id_tipo
        INNER JOIN cri.CRI_Rubro r
          ON r.id_rubro = t.id_rubro
        ORDER BY r.clave_rubro ASC, t.clave_tipo ASC, c.clave_clase ASC;
        """
      )
      clases = _rows(cur)

      cur.execute(
        """
        SELECT
          co.id_concepto,
          RTRIM(co.clave_concepto) AS clave_concepto,
          RTRIM(co.nombre_concepto) AS nombre_concepto,
          co.id_clase,
          RTRIM(c.clave_clase) AS clave_clase,
          c.id_tipo,
          RTRIM(t.clave_tipo) AS clave_tipo,
          t.id_rubro,
          RTRIM(r.clave_rubro) AS clave_rubro
        FROM cri.CRI_Concepto co
        INNER JOIN cri.CRI_Clase c
          ON c.id_clase = co.id_clase
        INNER JOIN cri.CRI_Tipo t
          ON t.id_tipo = c.id_tipo
        INNER JOIN cri.CRI_Rubro r
          ON r.id_rubro = t.id_rubro
        ORDER BY r.clave_rubro ASC, t.clave_tipo ASC, c.clave_clase ASC, co.clave_concepto ASC;
        """
      )
      conceptos = _rows(cur)

    tipos_by_rubro: Dict[int, List[Dict[str, Any]]] = {}
    for t in tipos:
      rid = int(t.get("id_rubro"))
      tipos_by_rubro.setdefault(rid, []).append(
        {
          "id_tipo": t.get("id_tipo"),
          "clave_tipo": t.get("clave_tipo"),
          "nombre_tipo": t.get("nombre_tipo"),
        }
      )

    items: List[Dict[str, Any]] = []
    for r in rubros:
      rid = int(r.get("id_rubro"))
      items.append(
        {
          "id_rubro": rid,
          "clave_rubro": r.get("clave_rubro"),
          "nombre_rubro": r.get("nombre_rubro"),
          "descripcion": r.get("descripcion"),
          "tipos": tipos_by_rubro.get(rid, []),
        }
      )

    return ORJSONResponse(
      {
        "ok": True,
        "database": "MicroServicios",
        "counts": {
          "rubros": len(rubros),
          "tipos": len(tipos),
          "clases": int(clases_cnt),
          "conceptos": int(conceptos_cnt),
        },
        "rubros": items,
        "clases": clases,
        "conceptos": conceptos,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/api/reportes/cri/entes")
def cri_report_entes() -> ORJSONResponse:
  try:
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        SELECT
          e.id_ente,
          RTRIM(e.nombre) AS nombre,
          RTRIM(e.nivel_gobierno) AS nivel_gobierno,
          RTRIM(COALESCE(e.rfc, '')) AS rfc,
          e.ejercicio_fiscal
        FROM cri.Ente_Publico e
        ORDER BY e.ejercicio_fiscal DESC, e.id_ente ASC;
        """
      )
      rows = _rows(cur)
    return ORJSONResponse({"ok": True, "count": len(rows), "items": rows})
  except Exception as e:
    return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})


def _cri_report_int(raw: Any) -> Optional[int]:
  if raw in (None, "", "null"):
    return None
  s = str(raw).strip()
  if not s:
    return None
  if s.isdigit():
    try:
      return int(s)
    except Exception:
      return None
  try:
    return int(float(s))
  except Exception:
    return None


@app.get("/api/reportes/cri/estado-analitico")
def cri_estado_analitico(request: Request) -> ORJSONResponse:
  try:
    id_ente = _cri_report_int(request.query_params.get("idEnte"))
    ejercicio_fiscal = _cri_report_int(request.query_params.get("ejercicioFiscal"))
    periodo = _cri_report_int(request.query_params.get("periodo"))

    limit = _cri_report_int(request.query_params.get("limit")) or 200
    offset = _cri_report_int(request.query_params.get("offset")) or 0
    limit = max(1, min(500, limit))
    offset = max(0, offset)

    where_sql = """
      WHERE (? IS NULL OR id_ente = ?)
        AND (? IS NULL OR ejercicio_fiscal = ?)
        AND (? IS NULL OR periodo = ?)
    """
    where_params = (id_ente, id_ente, ejercicio_fiscal, ejercicio_fiscal, periodo, periodo)

    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        "SELECT COUNT(*) AS cnt FROM cri.vw_Estado_Analitico_Ingresos " + where_sql + ";",
        where_params,
      )
      total = int(_rows(cur)[0]["cnt"])

      cur.execute(
        """
        SELECT
          id_ente,
          RTRIM(COALESCE(Ente_Publico, '')) AS Ente_Publico,
          RTRIM(COALESCE(Nivel_Gobierno, '')) AS Nivel_Gobierno,
          ejercicio_fiscal,
          periodo,
          RTRIM(COALESCE(clave_rubro, '')) AS clave_rubro,
          RTRIM(COALESCE(nombre_rubro, '')) AS nombre_rubro,
          RTRIM(COALESCE(clave_tipo, '')) AS clave_tipo,
          RTRIM(COALESCE(nombre_tipo, '')) AS nombre_tipo,
          Estimado,
          Modificado,
          Devengado,
          Recaudado,
          [Diferencia (%)]
        FROM cri.vw_Estado_Analitico_Ingresos
        """
        + where_sql
        + """
        ORDER BY ejercicio_fiscal DESC, periodo DESC, clave_rubro ASC, clave_tipo ASC
        OFFSET ? ROWS
        FETCH NEXT ? ROWS ONLY;
        """,
        where_params + (offset, limit),
      )
      rows = _rows(cur)

    next_offset = offset + len(rows)
    has_more = next_offset < total
    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "idEnte": id_ente,
          "ejercicioFiscal": ejercicio_fiscal,
          "periodo": periodo,
          "limit": limit,
          "offset": offset,
        },
        "count": len(rows),
        "total": total,
        "hasMore": has_more,
        "nextOffset": next_offset if has_more else None,
        "rows": rows,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/reportes/cri/resumen-por-rubro")
def cri_resumen_por_rubro(request: Request) -> ORJSONResponse:
  try:
    ejercicio_fiscal = _cri_report_int(request.query_params.get("ejercicioFiscal"))
    periodo = _cri_report_int(request.query_params.get("periodo"))

    limit = _cri_report_int(request.query_params.get("limit")) or 200
    offset = _cri_report_int(request.query_params.get("offset")) or 0
    limit = max(1, min(500, limit))
    offset = max(0, offset)

    where_sql = """
      WHERE (? IS NULL OR ejercicio_fiscal = ?)
        AND (? IS NULL OR periodo = ?)
    """
    where_params = (ejercicio_fiscal, ejercicio_fiscal, periodo, periodo)

    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        "SELECT COUNT(*) AS cnt FROM cri.vw_Resumen_por_Rubro " + where_sql + ";",
        where_params,
      )
      total = int(_rows(cur)[0]["cnt"])

      cur.execute(
        """
        SELECT
          RTRIM(COALESCE(clave_rubro, '')) AS clave_rubro,
          RTRIM(COALESCE(nombre_rubro, '')) AS nombre_rubro,
          ejercicio_fiscal,
          periodo,
          Estimado,
          Modificado,
          Devengado,
          Recaudado,
          Diferencia_Monto
        FROM cri.vw_Resumen_por_Rubro
        """
        + where_sql
        + """
        ORDER BY ejercicio_fiscal DESC, periodo DESC, clave_rubro ASC
        OFFSET ? ROWS
        FETCH NEXT ? ROWS ONLY;
        """,
        where_params + (offset, limit),
      )
      rows = _rows(cur)

    next_offset = offset + len(rows)
    has_more = next_offset < total
    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "ejercicioFiscal": ejercicio_fiscal,
          "periodo": periodo,
          "limit": limit,
          "offset": offset,
        },
        "count": len(rows),
        "total": total,
        "hasMore": has_more,
        "nextOffset": next_offset if has_more else None,
        "rows": rows,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/reportes/cri/estado-analitico.csv")
def cri_estado_analitico_csv(request: Request) -> StreamingResponse:
  id_ente = _cri_report_int(request.query_params.get("idEnte"))
  ejercicio_fiscal = _cri_report_int(request.query_params.get("ejercicioFiscal"))
  periodo = _cri_report_int(request.query_params.get("periodo"))

  max_rows = _cri_report_int(request.query_params.get("maxRows")) or 50000
  max_rows = max(1, min(200000, max_rows))

  where_sql = """
    WHERE (? IS NULL OR id_ente = ?)
      AND (? IS NULL OR ejercicio_fiscal = ?)
      AND (? IS NULL OR periodo = ?)
  """
  where_params = (id_ente, id_ente, ejercicio_fiscal, ejercicio_fiscal, periodo, periodo)

  columns = [
    "Ente_Publico",
    "Nivel_Gobierno",
    "ejercicio_fiscal",
    "periodo",
    "clave_rubro",
    "nombre_rubro",
    "clave_tipo",
    "nombre_tipo",
    "Estimado",
    "Modificado",
    "Devengado",
    "Recaudado",
    "Diferencia (%)",
  ]

  def gen() -> Iterable[bytes]:
    conn = get_conn_for_database("MicroServicios")
    try:
      cur = conn.cursor()
      cur.execute("SET NOCOUNT ON; SET XACT_ABORT ON;")
      yield (",".join(columns) + "\n").encode("utf-8")
      cur.execute(
        """
        SELECT TOP (?)
          RTRIM(COALESCE(Ente_Publico, '')) AS Ente_Publico,
          RTRIM(COALESCE(Nivel_Gobierno, '')) AS Nivel_Gobierno,
          ejercicio_fiscal,
          periodo,
          RTRIM(COALESCE(clave_rubro, '')) AS clave_rubro,
          RTRIM(COALESCE(nombre_rubro, '')) AS nombre_rubro,
          RTRIM(COALESCE(clave_tipo, '')) AS clave_tipo,
          RTRIM(COALESCE(nombre_tipo, '')) AS nombre_tipo,
          Estimado,
          Modificado,
          Devengado,
          Recaudado,
          [Diferencia (%)] AS [Diferencia (%)]
        FROM cri.vw_Estado_Analitico_Ingresos
        """
        + where_sql
        + """
        ORDER BY ejercicio_fiscal DESC, periodo DESC, clave_rubro ASC, clave_tipo ASC;
        """,
        (max_rows,) + where_params,
      )
      while cur.description is None:
        if not cur.nextset():
          return
      cols = [c[0] for c in cur.description]
      index = {name: i for i, name in enumerate(cols)}
      while True:
        batch = cur.fetchmany(500)
        if not batch:
          break
        for row in batch:
          line = ",".join(_csv_escape(row[index[c]]) if c in index else "" for c in columns) + "\n"
          yield line.encode("utf-8")
    finally:
      conn.close()

  headers = {"Content-Disposition": 'attachment; filename="cri_estado_analitico.csv"'}
  return StreamingResponse(gen(), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/reportes/cri/resumen-por-rubro.csv")
def cri_resumen_por_rubro_csv(request: Request) -> StreamingResponse:
  ejercicio_fiscal = _cri_report_int(request.query_params.get("ejercicioFiscal"))
  periodo = _cri_report_int(request.query_params.get("periodo"))

  max_rows = _cri_report_int(request.query_params.get("maxRows")) or 50000
  max_rows = max(1, min(200000, max_rows))

  where_sql = """
    WHERE (? IS NULL OR ejercicio_fiscal = ?)
      AND (? IS NULL OR periodo = ?)
  """
  where_params = (ejercicio_fiscal, ejercicio_fiscal, periodo, periodo)

  columns = [
    "clave_rubro",
    "nombre_rubro",
    "ejercicio_fiscal",
    "periodo",
    "Estimado",
    "Modificado",
    "Devengado",
    "Recaudado",
    "Diferencia_Monto",
  ]

  def gen() -> Iterable[bytes]:
    conn = get_conn_for_database("MicroServicios")
    try:
      cur = conn.cursor()
      cur.execute("SET NOCOUNT ON; SET XACT_ABORT ON;")
      yield (",".join(columns) + "\n").encode("utf-8")
      cur.execute(
        """
        SELECT TOP (?)
          RTRIM(COALESCE(clave_rubro, '')) AS clave_rubro,
          RTRIM(COALESCE(nombre_rubro, '')) AS nombre_rubro,
          ejercicio_fiscal,
          periodo,
          Estimado,
          Modificado,
          Devengado,
          Recaudado,
          Diferencia_Monto
        FROM cri.vw_Resumen_por_Rubro
        """
        + where_sql
        + """
        ORDER BY ejercicio_fiscal DESC, periodo DESC, clave_rubro ASC;
        """,
        (max_rows,) + where_params,
      )
      while cur.description is None:
        if not cur.nextset():
          return
      cols = [c[0] for c in cur.description]
      index = {name: i for i, name in enumerate(cols)}
      while True:
        batch = cur.fetchmany(500)
        if not batch:
          break
        for row in batch:
          line = ",".join(_csv_escape(row[index[c]]) if c in index else "" for c in columns) + "\n"
          yield line.encode("utf-8")
    finally:
      conn.close()

  headers = {"Content-Disposition": 'attachment; filename="cri_resumen_por_rubro.csv"'}
  return StreamingResponse(gen(), media_type="text/csv; charset=utf-8", headers=headers)
 
 
@app.get("/api/config/umas")
def get_config_umas() -> ORJSONResponse:
  items = []
  for year in sorted(UMA_MXN_BY_VIGENCIA_YEAR.keys()):
    v = UMA_MXN_BY_VIGENCIA_YEAR.get(year)
    if v is None:
      continue
    items.append({"vigenciaYear": year, "umaMxn": float(v)})
  return ORJSONResponse({"ok": True, "items": items, "count": len(items)})
 
 
@app.post("/api/config/umas")
async def upsert_config_umas(request: Request) -> ORJSONResponse:
  try:
    payload = await request.json()
  except Exception:
    payload = {}
 
  year_raw = payload.get("vigenciaYear")
  uma_raw = payload.get("umaMxn")
 
  try:
    year = int(str(year_raw).strip())
  except Exception:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": "vigenciaYear inválido"})
 
  if uma_raw in (None, "", "null"):
    with _UMA_STORE_LOCK:
      UMA_MXN_BY_VIGENCIA_YEAR.pop(year, None)
      try:
        _save_umas_to_disk()
      except Exception as e:
        return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})
    return ORJSONResponse({"ok": True, "vigenciaYear": year, "deleted": True})
 
  try:
    uma = decimal.Decimal(str(uma_raw).strip())
  except Exception:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": "umaMxn inválido"})
 
  with _UMA_STORE_LOCK:
    UMA_MXN_BY_VIGENCIA_YEAR[year] = uma
    try:
      _save_umas_to_disk()
    except Exception as e:
      return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})
  return ORJSONResponse({"ok": True, "vigenciaYear": year, "umaMxn": float(uma)})


@app.get("/api/config/recargos")
def get_config_recargos() -> ORJSONResponse:
  items = []
  for year in sorted(RECARGOS_MORA_BY_YEAR.keys()):
    v = RECARGOS_MORA_BY_YEAR.get(year)
    if v is None:
      continue
    items.append({"year": year, "tasaMora": float(v)})
  return ORJSONResponse({"ok": True, "tipo": "mora", "items": items, "count": len(items)})


@app.post("/api/config/recargos")
async def upsert_config_recargos(request: Request) -> ORJSONResponse:
  _reject_if_too_large(request, _UPLOAD_JSON_MAX_BYTES)
  try:
    payload = await request.json()
  except Exception:
    payload = {}

  year_raw = payload.get("year")
  tasa_raw = payload.get("tasaMora")

  try:
    year = int(str(year_raw).strip())
  except Exception:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": "year inválido"})

  if tasa_raw in (None, "", "null"):
    with _RECARGOS_STORE_LOCK:
      RECARGOS_MORA_BY_YEAR.pop(year, None)
      try:
        _save_recargos_to_disk()
      except Exception as e:
        return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})
    return ORJSONResponse({"ok": True, "year": year, "deleted": True})

  try:
    tasa = decimal.Decimal(str(tasa_raw).strip())
  except Exception:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": "tasaMora inválida"})
  if tasa <= 0:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": "tasaMora inválida"})

  with _RECARGOS_STORE_LOCK:
    RECARGOS_MORA_BY_YEAR[year] = tasa
    try:
      _save_recargos_to_disk()
    except Exception as e:
      return ORJSONResponse(status_code=500, content={"ok": False, "error": str(e)})

  return ORJSONResponse({"ok": True, "year": year, "tasaMora": float(tasa)})


@app.get("/api/config/inpc")
def get_config_inpc() -> ORJSONResponse:
  items: List[Dict[str, Any]] = []
  for p in _list_inpc_files():
    try:
      stat = p.stat()
      raw = p.read_bytes()
      data = orjson.loads(raw) if raw else {}
      flat = _flatten_inpc_payload(data)
      table = _parse_inpc_table(flat)
      min_key = None
      max_key = None
      if table:
        idx_min = min(table.keys())
        idx_max = max(table.keys())
        y1 = idx_min // 12
        m1 = (idx_min % 12) + 1
        y2 = idx_max // 12
        m2 = (idx_max % 12) + 1
        min_key = _predial_ym_key(y1, m1)
        max_key = _predial_ym_key(y2, m2)
      items.append(
        {
          "name": p.name,
          "sizeBytes": int(stat.st_size),
          "updatedAt": datetime.fromtimestamp(stat.st_mtime).isoformat(),
          "monthCount": int(len(table)),
          "minKey": min_key,
          "maxKey": max_key,
        }
      )
    except Exception:
      continue
  return ORJSONResponse({"ok": True, "items": items, "count": len(items)})


@app.get("/api/config/inpc/{filename}")
def get_config_inpc_file(filename: str) -> ORJSONResponse:
  name = _safe_inpc_filename(filename)
  path = _INPC_STORE_DIR / name
  if not path.exists() or not path.is_file():
    raise HTTPException(status_code=404, detail="Archivo INPC no encontrado")
  raw = path.read_bytes()
  data = orjson.loads(raw) if raw else {}
  return ORJSONResponse({"ok": True, "name": name, "data": data})


@app.put("/api/config/inpc/{filename}")
async def put_config_inpc_file(filename: str, request: Request) -> ORJSONResponse:
  name = _safe_inpc_filename(filename)
  _reject_if_too_large(request, _UPLOAD_JSON_MAX_BYTES)
  try:
    payload = await request.json()
  except Exception:
    raise HTTPException(status_code=400, detail="JSON inválido")
  flat = _flatten_inpc_payload(payload)
  table = _parse_inpc_table(flat)
  if not table:
    raise HTTPException(status_code=400, detail="El archivo INPC no contiene meses válidos (YYYY-MM)")
  path = _INPC_STORE_DIR / name
  _INPC_STORE_DIR.mkdir(parents=True, exist_ok=True)
  data = orjson.dumps(payload, option=orjson.OPT_INDENT_2)
  tmp = path.with_suffix(path.suffix + ".tmp")
  tmp.write_bytes(data)
  os.replace(tmp, path)
  return ORJSONResponse({"ok": True, "name": name, "saved": True})


@app.delete("/api/config/inpc/{filename}")
def delete_config_inpc_file(filename: str) -> ORJSONResponse:
  name = _safe_inpc_filename(filename)
  path = _INPC_STORE_DIR / name
  if not path.exists() or not path.is_file():
    raise HTTPException(status_code=404, detail="Archivo INPC no encontrado")
  try:
    os.remove(path)
  except Exception as e:
    raise HTTPException(status_code=500, detail=str(e))
  return ORJSONResponse({"ok": True, "name": name, "deleted": True})


@app.post("/api/config/inpc/upload")
async def upload_config_inpc(file: UploadFile = File(...)) -> ORJSONResponse:
  name = _safe_inpc_filename(file.filename or "")
  raw = await file.read()
  if not raw:
    raise HTTPException(status_code=400, detail="Archivo vacío")
  if len(raw) > int(_UPLOAD_JSON_MAX_BYTES):
    raise HTTPException(status_code=413, detail="Archivo demasiado grande")
  try:
    payload = orjson.loads(raw)
  except Exception:
    raise HTTPException(status_code=400, detail="JSON inválido")
  flat = _flatten_inpc_payload(payload)
  table = _parse_inpc_table(flat)
  if not table:
    raise HTTPException(status_code=400, detail="El archivo INPC no contiene meses válidos (YYYY-MM)")
  path = _INPC_STORE_DIR / name
  _INPC_STORE_DIR.mkdir(parents=True, exist_ok=True)
  data = orjson.dumps(payload, option=orjson.OPT_INDENT_2)
  tmp = path.with_suffix(path.suffix + ".tmp")
  tmp.write_bytes(data)
  os.replace(tmp, path)
  return ORJSONResponse({"ok": True, "name": name, "uploaded": True})
 
 
@app.get("/api/fuentes")
def fuentes(
  solicitudId: int,
  ano: int,
  grupoTramiteId: int = 42,
  cveFteMT: str = "MTULUM",
) -> ORJSONResponse:
  try:
    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        """
        DECLARE @GrupoTramiteId int = ?;
        DECLARE @SolicitudId int = ?;
        DECLARE @Ano int = ?;
        DECLARE @CveFteMT varchar(32) = ?;
 
        SELECT
          GrupoTramiteId,
          SolicitudId,
          SolicitudDetalleEjericicio,
          SolicitudDetalleFteIngId,
          CveFteMT,
          SolicitudDetalleImporteFijo
        FROM TLSOLICITUDFUENTESINGRESO
        WHERE GrupoTramiteId = @GrupoTramiteId
          AND SolicitudId = @SolicitudId
          AND SolicitudDetalleEjericicio = @Ano
          AND CveFteMT = @CveFteMT
        ORDER BY SolicitudDetalleFteIngId ASC
        OPTION (RECOMPILE);
        """,
        (grupoTramiteId, solicitudId, ano, cveFteMT),
      )
      rows = _rows(cur)
    return ORJSONResponse(
        {
          "ok": True,
          "filtros": {"grupoTramiteId": grupoTramiteId, "solicitudId": solicitudId, "ano": ano, "cveFteMT": cveFteMT},
          "count": len(rows),
          "rows": rows,
        }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})
 
 
@app.post("/api/consolidar")
async def consolidar(request: Request, x_admin_key: Optional[str] = Header(default=None)) -> ORJSONResponse:
  _require_admin(x_admin_key)
  _reject_if_too_large(request, _UPLOAD_JSON_MAX_BYTES)
  try:
    body = await request.json()
  except Exception:
    body = {}

  user = _get_current_user(request) or {}
  cached = _idempotency_lookup(request, user, body)
  if cached is not None:
    return cached

 
  try:
    solicitud_id = int(body.get("solicitudId"))
    ano = int(body.get("ano"))
    grupo_tramite_id = int(body.get("grupoTramiteId") or 42)
    cve_fte_mt = str(body.get("cveFteMT") or "MTULUM")
    nuevo_estado = str(body.get("nuevoEstado") or "PP")
    vencimiento = str(body.get("vencimientoFecha")).strip() if body.get("vencimientoFecha") else None

    lk = _lock_key("consolidar", cve_fte_mt, grupo_tramite_id, solicitud_id, ano)
    _acquire_lock(lk, user, ttl_seconds=120)

    resumen = None
    rows = []
    attempts = max(1, int(_DEADLOCK_RETRIES))
    for attempt in range(attempts):
      conn = get_conn()
      try:
        cur = conn.cursor()
        cur.execute(
          """
          DECLARE @GrupoTramiteId int = ?;
          DECLARE @SolicitudId int = ?;
          DECLARE @Ano int = ?;
          DECLARE @CveFteMT varchar(32) = ?;
          DECLARE @NuevoEstado varchar(8) = ?;
          DECLARE @Vencimiento varchar(32) = ?;
   
          SET NOCOUNT ON;
          SET XACT_ABORT ON;
   
          DECLARE @FtePrincipal int;
          DECLARE @FteSecundario int;
          DECLARE @ImporteTotal decimal(18,2);
          DECLARE @Cnt int;
   
          ;WITH x AS (
            SELECT
              SolicitudDetalleFteIngId,
              CAST(SolicitudDetalleImporteFijo AS decimal(18,2)) AS Importe,
              ROW_NUMBER() OVER (ORDER BY SolicitudDetalleFteIngId ASC) AS rn_asc,
              ROW_NUMBER() OVER (ORDER BY SolicitudDetalleFteIngId DESC) AS rn_desc,
              COUNT(*) OVER () AS cnt
            FROM TLSOLICITUDFUENTESINGRESO WITH (UPDLOCK, HOLDLOCK)
            WHERE GrupoTramiteId = @GrupoTramiteId
              AND SolicitudId = @SolicitudId
              AND SolicitudDetalleEjericicio = @Ano
              AND CveFteMT = @CveFteMT
          )
          SELECT
            @FtePrincipal = MAX(CASE WHEN rn_asc = 1 THEN SolicitudDetalleFteIngId END),
            @FteSecundario = MAX(CASE WHEN rn_desc = 1 THEN SolicitudDetalleFteIngId END),
            @ImporteTotal = SUM(Importe),
            @Cnt = MAX(cnt)
          FROM x;
   
          IF @Cnt <> 2 OR @FtePrincipal IS NULL OR @FteSecundario IS NULL OR @FtePrincipal = @FteSecundario
            THROW 51000, 'Se requieren exactamente 2 registros para consolidar (mismo filtro).', 1;
   
          UPDATE TLSOLICITUDFUENTESINGRESO
          SET SolicitudDetalleImporteFijo = @ImporteTotal
          WHERE GrupoTramiteId = @GrupoTramiteId
            AND SolicitudId = @SolicitudId
            AND SolicitudDetalleEjericicio = @Ano
            AND SolicitudDetalleFteIngId = @FtePrincipal
            AND CveFteMT = @CveFteMT;
   
          DELETE FROM TLSOLICITUDFUENTESINGRESO
          WHERE GrupoTramiteId = @GrupoTramiteId
            AND SolicitudId = @SolicitudId
            AND SolicitudDetalleEjericicio = @Ano
            AND SolicitudDetalleFteIngId = @FteSecundario
            AND CveFteMT = @CveFteMT;
   
          UPDATE TLSOLICITUD
          SET
            SolicitudEstado = @NuevoEstado,
            SolicitudVencimientoFecha =
              CASE
                WHEN @Vencimiento IS NULL THEN SolicitudVencimientoFecha
                ELSE COALESCE(
                  TRY_CONVERT(datetime, @Vencimiento, 126),
                  TRY_CONVERT(datetime, @Vencimiento, 23),
                  TRY_CONVERT(datetime, @Vencimiento, 112),
                  SolicitudVencimientoFecha
                )
              END
          WHERE GrupoTramiteId = @GrupoTramiteId
            AND SolicitudId = @SolicitudId;
   
          SELECT
            @FtePrincipal AS FtePrincipal,
            @FteSecundario AS FteSecundario,
            @ImporteTotal AS ImporteTotal;
   
          SELECT
            GrupoTramiteId,
            SolicitudId,
            SolicitudDetalleEjericicio,
            SolicitudDetalleFteIngId,
            CveFteMT,
            SolicitudDetalleImporteFijo
          FROM TLSOLICITUDFUENTESINGRESO
          WHERE GrupoTramiteId = @GrupoTramiteId
            AND SolicitudId = @SolicitudId
            AND SolicitudDetalleEjericicio = @Ano
            AND CveFteMT = @CveFteMT
          ORDER BY SolicitudDetalleFteIngId ASC
          OPTION (RECOMPILE);
          """,
          (grupo_tramite_id, solicitud_id, ano, cve_fte_mt, nuevo_estado, vencimiento),
        )
   
        resumen_rows = _rows(cur)
        resumen = resumen_rows[0] if resumen_rows else None
        cur.nextset()
        rows = _rows(cur)
        conn.commit()
        break
      except Exception as e:
        try:
          conn.rollback()
        except Exception:
          pass
        if attempt < (attempts - 1) and _is_deadlock_error(e):
          time.sleep(0.15 * (2 ** attempt))
          continue
        raise
      finally:
        try:
          conn.close()
        except Exception:
          pass
 
    resp = ORJSONResponse(
      {
        "ok": True,
        "filtros": {"grupoTramiteId": grupo_tramite_id, "solicitudId": solicitud_id, "ano": ano, "cveFteMT": cve_fte_mt},
        "resumen": resumen,
        "rows": rows,
      }
    )
    _idempotency_store(request, user, body, resp)
    return resp
  except HTTPException:
    raise
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})
  finally:
    try:
      if "lk" in locals():
        _release_lock(lk, user)
    except Exception:
      pass
 
 
@app.post("/api/activaciones")
async def activaciones(request: Request, x_admin_key: Optional[str] = Header(default=None)) -> ORJSONResponse:
  _require_admin(x_admin_key)
  _reject_if_too_large(request, _UPLOAD_JSON_MAX_BYTES)
  user = _get_current_user(request) or {}
  try:
    body = await request.json()
  except Exception:
    body = {}
  cached = _idempotency_lookup(request, user, body)
  if cached is not None:
    return cached
  try:
    lk = _lock_key("activaciones")
    _acquire_lock(lk, user, ttl_seconds=180)

    attempts = max(1, int(_DEADLOCK_RETRIES))
    for attempt in range(attempts):
      conn = get_conn()
      try:
        cur = conn.cursor()
        cur.execute("SET NOCOUNT ON; SET XACT_ABORT ON;")

        cur.execute(
          """
          UPDATE KUMiObligacion
          SET MiObligacionEstatus = N'AP'
          WHERE (MiObligacionEstatus <> N'AP');
          SELECT CAST(@@ROWCOUNT AS int) AS affected;
          """
        )
        obligacion_row = cur.fetchone()
        obligacion_rows = int(obligacion_row[0]) if obligacion_row else 0

        cur.execute(
          """
          UPDATE KUMiDocumento
          SET MiDocumentoEstatus = N'AP'
          WHERE (MiDocumentoEstatus <> N'AP');
          SELECT CAST(@@ROWCOUNT AS int) AS affected;
          """
        )
        documento_row = cur.fetchone()
        documento_rows = int(documento_row[0]) if documento_row else 0

        cur.execute(
          """
          UPDATE COQFORMASPAGOPAQUETE
          SET FormaPagoPaqueteCuentaBancaria = 235
          WHERE (CveCaja = 29)
            AND (CveFecAsi >= DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1));
          SELECT CAST(@@ROWCOUNT AS int) AS affected;
          """
        )
        paquete_row = cur.fetchone()
        paquete_rows = int(paquete_row[0]) if paquete_row else 0

        conn.commit()
        resp = ORJSONResponse(
          {
            "ok": True,
            "updated_rows": {
              "KUMiObligacion": obligacion_rows,
              "KUMiDocumento": documento_rows,
              "COQFORMASPAGOPAQUETE": paquete_rows,
            },
          }
        )
        _idempotency_store(request, user, body, resp)
        return resp
      except Exception as e:
        try:
          conn.rollback()
        except Exception:
          pass
        if attempt < (attempts - 1) and _is_deadlock_error(e):
          time.sleep(0.15 * (2 ** attempt))
          continue
        raise
      finally:
        try:
          conn.close()
        except Exception:
          pass
  except HTTPException:
    raise
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})
  finally:
    try:
      if "lk" in locals():
        _release_lock(lk, user)
    except Exception:
      pass


def _parse_date(value: Optional[str]) -> Optional[datetime]:
  if not value:
    return None
  try:
    return datetime.fromisoformat(value)
  except Exception:
    try:
      return datetime.strptime(value, "%Y-%m-%d")
    except Exception:
      return None


def _ensure_report_jobs_dir() -> None:
  try:
    _REPORT_JOBS_DIR.mkdir(parents=True, exist_ok=True)
  except Exception:
    return


def _report_job_file_path(job_id: str, ext: str) -> Path:
  safe_ext = str(ext or "").strip().lower()
  if safe_ext not in {"csv"}:
    safe_ext = "csv"
  _ensure_report_jobs_dir()
  return _REPORT_JOBS_DIR / f"{job_id}.{safe_ext}"


def _purge_expired_report_jobs() -> None:
  try:
    _ensure_auth_schema_microservicios()
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      cur.execute(
        """
        SELECT TOP 50 JobId
        FROM dbo.AppReportJobs
        WHERE ExpiresAt < SYSUTCDATETIME()
          AND Status NOT IN ('expired')
        ORDER BY ExpiresAt ASC;
        """
      )
      rows = cur.fetchall() or []
      for r in rows:
        try:
          jid = str(r[0])
        except Exception:
          continue
        try:
          p = _report_job_file_path(jid, "csv")
          if p.exists() and p.is_file():
            try:
              os.remove(p)
            except Exception:
              pass
        except Exception:
          pass
        cur.execute("UPDATE dbo.AppReportJobs SET Status='expired' WHERE JobId = ?;", (jid,))
      conn.commit()
  except Exception:
    return


def _create_report_job(kind: str, params: Dict[str, Any], user: Dict[str, Any]) -> Dict[str, Any]:
  _purge_expired_report_jobs()
  _ensure_auth_schema_microservicios()
  job_id = str(uuid.uuid4())
  try:
    params_json = orjson.dumps(params or {}, option=orjson.OPT_SORT_KEYS).decode("utf-8")
  except Exception:
    params_json = "{}"
  user_id = int(user.get("id") or 0) if str(user.get("id") or "").strip() else None
  username = str(user.get("username") or "")[:120] or None
  expires_at = datetime.utcnow() + timedelta(seconds=int(_REPORT_JOBS_TTL_SECONDS))
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      INSERT INTO dbo.AppReportJobs (JobId, Kind, ParamsJson, Status, CreatedByUserId, CreatedByUsername, ExpiresAt, FileName)
      VALUES (?, ?, ?, 'queued', ?, ?, ?, ?);
      """,
      (job_id, str(kind)[:80], params_json, user_id, username, expires_at, f"{kind}.csv"),
    )
    conn.commit()
  return {"jobId": job_id, "kind": kind, "status": "queued"}


def _update_report_job(job_id: str, status: str, started: bool = False, finished: bool = False, row_count: Optional[int] = None, error: Optional[str] = None) -> None:
  try:
    _ensure_auth_schema_microservicios()
    with get_conn_for_database("MicroServicios") as conn:
      cur = conn.cursor()
      started_at = datetime.utcnow() if started else None
      finished_at = datetime.utcnow() if finished else None
      cur.execute(
        """
        UPDATE dbo.AppReportJobs
        SET
          Status = ?,
          StartedAt = COALESCE(StartedAt, ?),
          FinishedAt = CASE WHEN ? IS NULL THEN FinishedAt ELSE ? END,
          RowsCount = COALESCE(?, RowsCount),
          ErrorMessage = COALESCE(?, ErrorMessage)
        WHERE JobId = ?;
        """,
        (str(status)[:20], started_at, finished_at, finished_at, row_count, (str(error)[:400] if error else None), job_id),
      )
      conn.commit()
  except Exception:
    return


def _get_report_job(job_id: str) -> Optional[Dict[str, Any]]:
  _purge_expired_report_jobs()
  _ensure_auth_schema_microservicios()
  with get_conn_for_database("MicroServicios") as conn:
    cur = conn.cursor()
    cur.execute(
      """
      SELECT TOP 1
        JobId AS jobId,
        Kind AS kind,
        ParamsJson AS paramsJson,
        Status AS status,
        CreatedAt AS createdAt,
        StartedAt AS startedAt,
        FinishedAt AS finishedAt,
        ExpiresAt AS expiresAt,
        RowsCount AS rowCount,
        FileName AS fileName,
        ErrorMessage AS errorMessage
      FROM dbo.AppReportJobs
      WHERE JobId = ?;
      """,
      (job_id,),
    )
    rows = _rows(cur)
    return rows[0] if rows else None


def _write_csv_file(sql_text: str, sql_params: Tuple[Any, ...], columns: List[str], output_path: Path) -> int:
  row_count = 0
  conn = get_conn()
  try:
    cur = conn.cursor()
    cur.execute(sql_text, sql_params)
    while cur.description is None:
      if not cur.nextset():
        break
    cols = [c[0] for c in cur.description] if cur.description else []
    index = {name: i for i, name in enumerate(cols)}
    with open(output_path, "wb") as f:
      f.write(b"\xef\xbb\xbf")
      f.write((",".join(columns) + "\n").encode("utf-8"))
      while True:
        batch = cur.fetchmany(500)
        if not batch:
          break
        for row in batch:
          line = ",".join(_csv_escape(row[index[c]]) if c in index else "" for c in columns) + "\n"
          f.write(line.encode("utf-8"))
          row_count += 1
  finally:
    try:
      conn.close()
    except Exception:
      pass
  return row_count


def _run_report_job(job_id: str) -> None:
  job = _get_report_job(job_id)
  if not job:
    return
  if str(job.get("status") or "").lower() not in {"queued"}:
    return

  kind = str(job.get("kind") or "")
  try:
    params = orjson.loads(str(job.get("paramsJson") or "{}").encode("utf-8"))
  except Exception:
    params = {}

  _update_report_job(job_id, "running", started=True)
  out_path = _report_job_file_path(job_id, "csv")
  try:
    if kind == "prediales_sabana_csv":
      p = _prediales_sabana_filters(params)
      columns = p["columns"]
      sql_params = p["sqlParams"]
      row_count = _write_csv_file(_SABANA_PREDIALES_CSV, sql_params, columns, out_path)
    elif kind == "prediales_sabana_pagos_csv":
      cve_fte_mt = str(params.get("cveFteMT") or "MTULUM")
      f = _prediales_pagos_filters(dict(params))
      max_rows_raw = params.get("maxRows")
      max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
      max_rows = max(1, min(200000, max_rows))
      columns = [
        "Clave",
        "Clave Catastral",
        "Propietario",
        "Razon social del contribuyente",
        "Direccion",
        "Colonia",
        "Tipo de Predio",
        "Valor del terreno",
        "Área construida",
        "Valor de construcción",
        "Valor catastral",
        "Calificativo",
        "Estado fisico",
        "Periodo inicial",
        "Periodo final",
        "Recibo",
        "Fecha de Pago",
        "Impuesto Corriente y Anticipado",
        "Rezago años anteriores",
        "Rezago",
        "Adicional",
        "Actualizacion",
        "Recargos",
        "Requerimiento",
        "Embargo",
        "Multa",
        "Descuentos",
        "Total",
      ]
      sql_params = (
        cve_fte_mt,
        f["claveCatastral"] or None,
        f["claveCatastralFrom"] or None,
        f["claveCatastralTo"] or None,
        f["predioId"],
        f["pagoFrom"],
        f["pagoTo"],
        max_rows,
      )
      row_count = _write_csv_file(_SABANA_PAGOS_CSV, sql_params, columns, out_path)
    elif kind == "licencias_funcionamiento_csv":
      cve_fte_mt = str(params.get("cveFteMT") or "MTULUM")
      f = _licencias_func_filters(dict(params))
      max_rows_raw = params.get("maxRows")
      max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
      max_rows = max(1, min(200000, max_rows))
      def parse_int_env(name: str, default: Optional[int] = None) -> Optional[int]:
        raw = os.getenv(name)
        if raw in (None, "", "null"):
          return default
        try:
          return int(str(raw).strip())
        except Exception:
          return default
      licencia_principal = parse_int_env("LIC_FUNC_LICENCIA_FTEING")
      actualizaciones_fteing = parse_int_env("LIC_FUNC_ACTUALIZACION_FTEING", 4319020273)
      recargos_fteing = parse_int_env("LIC_FUNC_RECARGOS_FTEING", 4501010101)
      uma_ref = (
        datetime(int(f["ejercicio"]), 2, 1) if f.get("ejercicio") else (f.get("pagoFrom") or f.get("pagoTo") or datetime.now())
      )
      uma_mxn = get_uma_mxn_for_date(uma_ref) or decimal.Decimal("0")
      columns = [
        "No. Licencia",
        "Serie",
        "Folio",
        "Fecha",
        "Nombre",
        "RFC",
        "Observaciones",
        "Domicilio Licencia",
        "Domicilio Local",
        "Tipo Establecimiento",
        "Giro",
        "Base Licencia Nueva",
        "Base Licencia Renovación",
        "Base Basura Nueva",
        "Tipo",
        "Tarifa Licencia",
        "Tarifa Basura",
        "Licencia",
        "Lic Renovación",
        "Basura",
        "Actualizaciones",
        "Recargos",
        "Otros",
        "Total",
      ]
      sql_params = (
        cve_fte_mt,
        f["pagoFrom"],
        f["pagoTo"],
        f["tipo"],
        f["licenciaId"],
        f["licenciaFrom"],
        f["licenciaTo"],
        licencia_principal,
        actualizaciones_fteing,
        recargos_fteing,
        f["ejercicio"],
        uma_mxn,
        max_rows,
      )
      row_count = _write_csv_file(_LIC_FUNC_CSV, sql_params, columns, out_path)
    elif kind == "saneamiento_ambiental_csv":
      cve_fte_mt = str(params.get("cveFteMT") or "MTULUM")
      f = _saneamiento_ambiental_filters(dict(params))
      max_rows_raw = params.get("maxRows")
      max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
      max_rows = max(1, min(200000, max_rows))
      def parse_int_env(name: str, default: int) -> int:
        raw = os.getenv(name)
        if raw in (None, "", "null"):
          return default
        try:
          return int(str(raw).strip())
        except Exception:
          return default
      derecho_fteing = parse_int_env("SAN_AMB_DERECHO_FTEING", 4326010110)
      actualizaciones_fteing = parse_int_env("SAN_AMB_ACTUALIZACION_FTEING", 4326010111)
      recargos_fteing = parse_int_env("SAN_AMB_RECARGOS_FTEING", 4501012601)
      columns = [
        "Padrón",
        "No. Licencia",
        "Serie",
        "Folio",
        "Fecha",
        "Nombre",
        "RFC",
        "Observaciones",
        "Domicilio Licencia",
        "Domicilio Local",
        "Giro",
        "No. Cuartos",
        "Periodo pagado",
        "Derecho",
        "Actualizaciones",
        "Recargos",
        "Total",
      ]
      sql_params = (
        cve_fte_mt,
        f["pagoFrom"],
        f["pagoTo"],
        f["licenciaId"],
        derecho_fteing,
        actualizaciones_fteing,
        recargos_fteing,
        max_rows,
      )
      row_count = _write_csv_file(_SAN_AMB_EXPORT_SELECT, sql_params, columns, out_path)
    else:
      raise Exception("Tipo de reporte no soportado")

    _update_report_job(job_id, "done", finished=True, row_count=int(row_count))
  except Exception as e:
    try:
      if out_path.exists():
        try:
          os.remove(out_path)
        except Exception:
          pass
    except Exception:
      pass
    _update_report_job(job_id, "failed", finished=True, error=str(e))


def _start_report_job(job_id: str) -> None:
  t = threading.Thread(target=_run_report_job, args=(job_id,), daemon=True)
  t.start()


@app.post("/api/reportes/jobs")
async def create_report_job(request: Request) -> ORJSONResponse:
  _reject_if_too_large(request, 131072)
  user = _require_login(request)
  try:
    payload = await request.json()
  except Exception:
    payload = {}
  kind = str((payload or {}).get("kind") or "").strip()
  params = (payload or {}).get("params")
  if not kind:
    raise HTTPException(status_code=400, detail="kind requerido")
  if params is None:
    params = {}
  if not isinstance(params, dict):
    raise HTTPException(status_code=400, detail="params inválido")
  info = _create_report_job(kind, params, user)
  _start_report_job(str(info.get("jobId")))
  job_id = str(info.get("jobId"))
  return ORJSONResponse(
    {
      "ok": True,
      "job": info,
      "urls": {
        "status": f"/api/reportes/jobs/{job_id}",
        "download": f"/api/reportes/jobs/{job_id}/download",
      },
    }
  )


@app.get("/api/reportes/jobs/{job_id}")
def get_report_job(job_id: str, request: Request) -> ORJSONResponse:
  _require_login(request)
  job = _get_report_job(job_id)
  if not job:
    raise HTTPException(status_code=404, detail="Job no encontrado")
  return ORJSONResponse({"ok": True, "job": job})


@app.get("/api/reportes/jobs/{job_id}/download")
def download_report_job(job_id: str, request: Request) -> StreamingResponse:
  _require_login(request)
  job = _get_report_job(job_id)
  if not job:
    raise HTTPException(status_code=404, detail="Job no encontrado")
  status = str(job.get("status") or "").lower()
  if status != "done":
    raise HTTPException(status_code=409, detail="El job no está listo")
  p = _report_job_file_path(job_id, "csv")
  if not p.exists() or not p.is_file():
    raise HTTPException(status_code=404, detail="Archivo no encontrado")
  filename = str(job.get("fileName") or f"{job_id}.csv")
  headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
  def gen() -> Iterable[bytes]:
    with open(p, "rb") as f:
      while True:
        chunk = f.read(1024 * 128)
        if not chunk:
          break
        yield chunk
  return StreamingResponse(gen(), media_type="text/csv; charset=utf-8", headers=headers)


def _prediales_sabana_filters(params: Dict[str, Any]) -> Dict[str, Any]:
  cve_fte_mt = str(params.get("cveFteMT") or "MTULUM")
  q = (params.get("q") or "").strip()
  clave_catastral = (params.get("claveCatastral") or "").strip()
  clave_mode_raw = (params.get("claveMode") or params.get("claveCatastralMode") or "").strip().lower()
  if clave_mode_raw in {"exacto", "exacta", "exact"}:
    clave_mode = "exacto"
  else:
    clave_mode = "contiene"

  predio_id_raw = params.get("predioId") or params.get("padron")
  propietario = (params.get("propietario") or "").strip()
  apellido_paterno = (params.get("apellidoPaterno") or "").strip()
  apellido_materno = (params.get("apellidoMaterno") or "").strip()
  nombre = (params.get("nombre") or "").strip()
  calle = (params.get("calle") or "").strip()
  numero = (params.get("numero") or "").strip()
  estatus = (params.get("estatus") or "").strip()
  adeudo_raw = (params.get("adeudo") or "").strip().lower()
  adeudo = adeudo_raw if adeudo_raw in {"todos", "con", "sin"} else ""

  from_alta = _parse_date(params.get("fromAlta"))
  to_alta = _parse_date(params.get("toAlta"))

  max_rows_raw = params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
  max_rows = max(1, min(200000, max_rows))

  def to_int(value: Any) -> Optional[int]:
    if value in (None, "", "null"):
      return None
    s = str(value).strip()
    if not s:
      return None
    if s.isdigit():
      try:
        return int(s)
      except Exception:
        return None
    try:
      return int(float(s))
    except Exception:
      return None

  predio_id = to_int(predio_id_raw)
  ejercicio_actual = datetime.now().year

  columns = [
    "Padrón",
    "Clave Catastral",
    "Fecha de Alta del Predio",
    "Calle",
    "Código Postal",
    "Número",
    "Propietario",
    "Estatus",
    "Datos escriturales",
    "Superficie del Terreno",
    "Valor del Terreno",
    "Área Construida",
    "Valor de Construcción",
    "Valor Catastral",
    "Año del valor catastral",
    "Impuesto Actual",
    "Impuesto por bimestre",
    "Ut.Bim.Pagado",
    "Tipo de Predio",
    "Estado Físico",
    "Ejer - Per",
  ]

  sql_params = (
    cve_fte_mt,
    q or None,
    clave_catastral or None,
    clave_mode,
    predio_id,
    propietario or None,
    apellido_paterno or None,
    apellido_materno or None,
    nombre or None,
    calle or None,
    numero or None,
    estatus or None,
    adeudo or None,
    from_alta,
    to_alta,
    ejercicio_actual,
    max_rows,
  )

  return {
    "cveFteMT": cve_fte_mt,
    "q": q,
    "claveCatastral": clave_catastral,
    "claveMode": clave_mode,
    "predioId": predio_id,
    "propietario": propietario,
    "apellidoPaterno": apellido_paterno,
    "apellidoMaterno": apellido_materno,
    "nombre": nombre,
    "calle": calle,
    "numero": numero,
    "estatus": estatus,
    "adeudo": adeudo or "todos",
    "fromAlta": params.get("fromAlta") or None,
    "toAlta": params.get("toAlta") or None,
    "maxRows": max_rows,
    "columns": columns,
    "sqlParams": sql_params,
  }
 
 
_SABANA_PREDIALES_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @Q varchar(200) = ?;
DECLARE @ClaveCatastral varchar(80) = ?;
DECLARE @ClaveMode varchar(16) = ?;
DECLARE @PredioId decimal(18,0) = ?;
DECLARE @Propietario varchar(200) = ?;
DECLARE @ApellidoPaterno varchar(120) = ?;
DECLARE @ApellidoMaterno varchar(120) = ?;
DECLARE @Nombre varchar(120) = ?;
DECLARE @Calle varchar(200) = ?;
DECLARE @Numero varchar(50) = ?;
DECLARE @Estatus varchar(8) = ?;
DECLARE @Adeudo varchar(16) = ?;
DECLARE @FromAlta datetime = ?;
DECLARE @ToAlta datetime = ?;
DECLARE @EjercicioActual int = ?;
DECLARE @Limit int = ?;
DECLARE @Offset int = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    p.CveFteMT,
    p.PredioId,
    RTRIM(COALESCE(p.PredioCveCatastral COLLATE DATABASE_DEFAULT, '')) AS PredioCveCatastral,
    p.PredioAltaFecha,
    RTRIM(COALESCE(p.PredioCalle COLLATE DATABASE_DEFAULT, '')) AS PredioCalle,
    p.PredioCodigoPostal,
    RTRIM(COALESCE(p.PredioNumExt COLLATE DATABASE_DEFAULT, '')) AS PredioNumExt,
    RTRIM(COALESCE(p.PredioNumInt COLLATE DATABASE_DEFAULT, '')) AS PredioNumInt,
    RTRIM(COALESCE(p.PredioStatus COLLATE DATABASE_DEFAULT, '')) AS PredioStatus,
    RTRIM(COALESCE(p.CatastroDatosEscriturales COLLATE DATABASE_DEFAULT, '')) AS CatastroDatosEscriturales,
    CAST(COALESCE(p.PredioTotalTerreno, 0) AS decimal(18,2)) AS PredioTotalTerreno,
    CAST(COALESCE(p.PredioArea, 0) AS decimal(18,2)) AS PredioArea,
    CAST(COALESCE(p.PredioTotalConstruccion, 0) AS decimal(18,2)) AS PredioTotalConstruccion,
    CAST(COALESCE(p.PredioTerrenoImporte, 0) AS decimal(18,2)) AS PredioTerrenoImporte,
    CAST(COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS PredioConstruccionImporte,
    CAST(COALESCE(p.PredioCatastralImporte, 0) AS decimal(18,2)) AS PredioCatastralImporte,
    p.PredioUltimoEjericicioPagado,
    p.PredioUltimoPeriodoPagado,
    RTRIM(COALESCE(p.PredioTipo COLLATE DATABASE_DEFAULT, '')) AS PredioTipo,
    RTRIM(COALESCE(per.NombreCompletoPersona COLLATE DATABASE_DEFAULT, per.RazonSocialPersona COLLATE DATABASE_DEFAULT, '')) AS PropietarioNombre,
    RTRIM(COALESCE(per.ApellidoPaternoPersona COLLATE DATABASE_DEFAULT, '')) AS ApellidoPaternoPersona,
    RTRIM(COALESCE(per.ApellidoMaternoPersona COLLATE DATABASE_DEFAULT, '')) AS ApellidoMaternoPersona,
    RTRIM(COALESCE(per.NombrePersona COLLATE DATABASE_DEFAULT, '')) AS NombrePersona,
    RTRIM(COALESCE(ef.EstadoFisicoNombre COLLATE DATABASE_DEFAULT, '')) AS EstadoFisicoNombre,
    CAST(
      COALESCE(
        CAST(p.PredioActualAdeudoImporte AS decimal(18,2)),
        CAST(p.PredioAdeudoImporte AS decimal(18,2)),
        CAST(p.PredioTotalAdeudoImporte AS decimal(18,2)),
        0
      )
    AS decimal(18,2)) AS AdeudoImporte
  FROM AlPredio p
  LEFT JOIN XiPersonas per
    ON per.CveFteMT = p.CveFteMT
   AND per.CvePersona = p.CvePersona
  LEFT JOIN AlEstadoFisico ef
    ON ef.CveFteMT = p.CveFteMT
   AND ef.EstadoFisicoId = p.EstadoFisicoId
  WHERE p.CveFteMT = @CveFteMT
    AND (@FromAlta IS NULL OR p.PredioAltaFecha >= @FromAlta)
    AND (@ToAlta IS NULL OR p.PredioAltaFecha < DATEADD(DAY, 1, @ToAlta))
    AND (@PredioId IS NULL OR p.PredioId = @PredioId)
    AND (@Estatus IS NULL OR RTRIM(COALESCE(p.PredioStatus COLLATE DATABASE_DEFAULT, '')) = @Estatus)
    AND (
      @ClaveCatastral IS NULL OR
      (@ClaveMode = 'exacto' AND RTRIM(COALESCE(p.PredioCveCatastral COLLATE DATABASE_DEFAULT, '')) = @ClaveCatastral) OR
      (@ClaveMode <> 'exacto' AND p.PredioCveCatastral LIKE '%' + @ClaveCatastral + '%')
    )
    AND (@Propietario IS NULL OR per.NombreCompletoPersona LIKE '%' + @Propietario + '%' OR per.RazonSocialPersona LIKE '%' + @Propietario + '%')
    AND (@ApellidoPaterno IS NULL OR per.ApellidoPaternoPersona LIKE '%' + @ApellidoPaterno + '%')
    AND (@ApellidoMaterno IS NULL OR per.ApellidoMaternoPersona LIKE '%' + @ApellidoMaterno + '%')
    AND (@Nombre IS NULL OR per.NombrePersona LIKE '%' + @Nombre + '%')
    AND (@Calle IS NULL OR p.PredioCalle LIKE '%' + @Calle + '%')
    AND (
      @Numero IS NULL OR
      p.PredioNumExt LIKE '%' + @Numero + '%' OR
      p.PredioNumInt LIKE '%' + @Numero + '%'
    )
    AND (
      @Adeudo IS NULL OR @Adeudo = 'todos' OR
      (@Adeudo = 'con' AND COALESCE(p.PredioActualAdeudoImporte, p.PredioAdeudoImporte, p.PredioTotalAdeudoImporte, 0) > 0) OR
      (@Adeudo = 'sin' AND COALESCE(p.PredioActualAdeudoImporte, p.PredioAdeudoImporte, p.PredioTotalAdeudoImporte, 0) <= 0)
    )
    AND (
      @Q IS NULL OR
      p.PredioCveCatastral LIKE '%' + @Q + '%' OR
      p.PredioClavePredial LIKE '%' + @Q + '%' OR
      per.NombreCompletoPersona LIKE '%' + @Q + '%' OR
      per.RazonSocialPersona LIKE '%' + @Q + '%'
    )
),
page AS (
  SELECT *
  FROM base
  ORDER BY PredioId ASC
  OFFSET @Offset ROWS
  FETCH NEXT @Limit ROWS ONLY
),
terr AS (
  SELECT
    t.CveFteMT,
    t.PredioId,
    CAST(COALESCE(SUM(t.PredioTerrenoArea), 0) AS decimal(18,2)) AS TerrenoAreaSum,
    CAST(COALESCE(SUM(t.PredioTerrenoArea * t.PredioTerrenoUnitarioImporte), 0) AS decimal(18,2)) AS TerrenoImporteSum
  FROM ALPREDIOTERRENOS t
  INNER JOIN page p
    ON p.CveFteMT = t.CveFteMT
   AND p.PredioId = t.PredioId
  GROUP BY t.CveFteMT, t.PredioId
),
cons AS (
  SELECT
    c.CveFteMT,
    c.PredioId,
    CAST(COALESCE(SUM(c.PredioConstruccionArea), 0) AS decimal(18,2)) AS ConstruccionAreaSum
  FROM ALPREDIOCONSTRUCCIONES c
  INNER JOIN page p
    ON p.CveFteMT = c.CveFteMT
   AND p.PredioId = c.PredioId
  GROUP BY c.CveFteMT, c.PredioId
),
val AS (
  SELECT
    v.CveFteMT,
    v.PredioId,
    v.PredioValuoCatastralImporte,
    ROW_NUMBER() OVER (
      PARTITION BY v.CveFteMT, v.PredioId
      ORDER BY v.PredioValuoCatastralFecha DESC, v.PredioValuoCatastralEjercicio DESC
    ) AS rn
  FROM ALPREDIOVALUOCATASTRAL v
  INNER JOIN page p
    ON p.CveFteMT = v.CveFteMT
   AND p.PredioId = v.PredioId
),
recibo AS (
  SELECT
    r.CveFteMT,
    r.PredioId,
    r.PredioEdoCuentaEjercicio AS ReciboEjercicio,
    r.PredioEdoCuentaPeriodo AS ReciboPeriodo,
    ROW_NUMBER() OVER (
      PARTITION BY r.CveFteMT, r.PredioId
      ORDER BY r.PredioEdoCuentaEjercicio DESC, r.PredioEdoCuentaPeriodo DESC, r.PredioEdoCuentaReciboId DESC
    ) AS rn
  FROM ALPREDIOEDOCUENTARECIBO r
  INNER JOIN page p
    ON p.CveFteMT = r.CveFteMT
   AND p.PredioId = r.PredioId
),
aval AS (
  SELECT
    a.CveFteMT,
    a.PredioId,
    a.SolicitudAvaluoEjercicioAvaluoActual,
    ROW_NUMBER() OVER (
      PARTITION BY a.CveFteMT, a.PredioId
      ORDER BY a.SolicitudAvaluoEjercicioAvaluoActual DESC, a.SolicitudAvaluoId DESC
    ) AS rn
  FROM AL23SolicitudAvaluo a
  INNER JOIN page p
    ON p.CveFteMT = a.CveFteMT
   AND p.PredioId = a.PredioId
),
mop AS (
  SELECT
    m.CveFteMT,
    m.PredioId,
    CAST(m.PredioEdoCuentaTasa AS decimal(18,6)) AS PredioEdoCuentaTasa,
    ROW_NUMBER() OVER (
      PARTITION BY m.CveFteMT, m.PredioId
      ORDER BY m.PredioEdoCuentaActualizacionFecha DESC, m.PredioEdoCuentaVencimientoFecha DESC, m.PredioEjercicioAlta DESC
    ) AS rn
  FROM ALMOVEDOPRE m
  INNER JOIN page p
    ON p.CveFteMT = m.CveFteMT
   AND p.PredioId = m.PredioId
  WHERE m.PredioEdoCuentaEjercicio = @EjercicioActual
    AND m.PredioEdoCuentaPeriodo = 1
)
SELECT
  CAST(p.PredioId AS int) AS [Padrón],
  p.PredioCveCatastral AS [Clave Catastral],
  p.PredioAltaFecha AS [Fecha de Alta del Predio],
  p.PredioCalle AS [Calle],
  p.PredioCodigoPostal AS [Código Postal],
  CASE
    WHEN NULLIF(RTRIM(p.PredioNumExt), '') IS NOT NULL THEN
      CASE
        WHEN NULLIF(RTRIM(p.PredioNumInt), '') IS NOT NULL THEN CONCAT(RTRIM(p.PredioNumExt), '-', RTRIM(p.PredioNumInt))
        ELSE RTRIM(p.PredioNumExt)
      END
    ELSE
      CASE
        WHEN NULLIF(RTRIM(p.PredioNumInt), '') IS NOT NULL THEN CONCAT('INT ', RTRIM(p.PredioNumInt))
        ELSE 'S/N'
      END
  END AS [Número],
  p.PropietarioNombre AS [Propietario],
  p.PredioStatus AS [Estatus],
  p.CatastroDatosEscriturales AS [Datos escriturales],
  CAST(COALESCE(terr.TerrenoAreaSum, p.PredioTotalTerreno, p.PredioArea, 0) AS decimal(18,2)) AS [Superficie del Terreno],
  CAST(COALESCE(terr.TerrenoImporteSum, p.PredioTerrenoImporte, 0) AS decimal(18,2)) AS [Valor del Terreno],
  CAST(COALESCE(cons.ConstruccionAreaSum, p.PredioTotalConstruccion, 0) AS decimal(18,2)) AS [Área Construida],
  CAST(COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS [Valor de Construcción],
  CAST(COALESCE(CAST(v.PredioValuoCatastralImporte AS decimal(18,2)), p.PredioCatastralImporte, 0) AS decimal(18,2)) AS [Valor Catastral],
  CAST(av.SolicitudAvaluoEjercicioAvaluoActual AS int) AS [Año del valor catastral],
  CAST(
    COALESCE(CAST(v.PredioValuoCatastralImporte AS decimal(18,2)), p.PredioCatastralImporte, 0) *
    COALESCE(m.PredioEdoCuentaTasa, 0)
  AS decimal(18,2)) AS [Impuesto Actual],
  CAST(
    (
      COALESCE(CAST(v.PredioValuoCatastralImporte AS decimal(18,2)), p.PredioCatastralImporte, 0) *
      COALESCE(m.PredioEdoCuentaTasa, 0)
    ) / CAST(6 AS decimal(18,6))
  AS decimal(18,2)) AS [Impuesto por bimestre],
  CASE
    WHEN p.PredioUltimoEjericicioPagado IS NOT NULL AND p.PredioUltimoPeriodoPagado IS NOT NULL
      THEN CONCAT(CAST(p.PredioUltimoEjericicioPagado AS varchar(4)), ' - ', CAST(p.PredioUltimoPeriodoPagado AS varchar(2)))
    WHEN p.PredioUltimoEjericicioPagado IS NOT NULL THEN CAST(p.PredioUltimoEjericicioPagado AS varchar(4))
    WHEN p.PredioUltimoPeriodoPagado IS NOT NULL THEN CAST(p.PredioUltimoPeriodoPagado AS varchar(2))
    ELSE ''
  END AS [Ut.Bim.Pagado],
  p.PredioTipo AS [Tipo de Predio],
  p.EstadoFisicoNombre AS [Estado Físico],
  CASE
    WHEN r.ReciboEjercicio IS NOT NULL AND r.ReciboPeriodo IS NOT NULL THEN CONCAT(CAST(r.ReciboEjercicio AS varchar(4)), '-', CAST(r.ReciboPeriodo AS varchar(2)))
    WHEN r.ReciboEjercicio IS NOT NULL THEN CAST(r.ReciboEjercicio AS varchar(4))
    WHEN r.ReciboPeriodo IS NOT NULL THEN CAST(r.ReciboPeriodo AS varchar(2))
    ELSE ''
  END AS [Ejer - Per]
FROM page p
LEFT JOIN terr
  ON terr.CveFteMT = p.CveFteMT
 AND terr.PredioId = p.PredioId
LEFT JOIN cons
  ON cons.CveFteMT = p.CveFteMT
 AND cons.PredioId = p.PredioId
LEFT JOIN val v
  ON v.CveFteMT = p.CveFteMT
 AND v.PredioId = p.PredioId
 AND v.rn = 1
LEFT JOIN recibo r
  ON r.CveFteMT = p.CveFteMT
 AND r.PredioId = p.PredioId
 AND r.rn = 1
LEFT JOIN aval av
  ON av.CveFteMT = p.CveFteMT
 AND av.PredioId = p.PredioId
 AND av.rn = 1
LEFT JOIN mop m
  ON m.CveFteMT = p.CveFteMT
 AND m.PredioId = p.PredioId
 AND m.rn = 1
ORDER BY p.PredioId ASC
OPTION (RECOMPILE);
"""


def _licencias_func_filters(params: Dict[str, Any]) -> Dict[str, Any]:
  ejercicio = int(params["ejercicio"]) if params.get("ejercicio") not in (None, "", "null") else None
  pago_from = _parse_date(params.get("pagoFrom"))
  pago_to = _parse_date(params.get("pagoTo"))
  if not pago_from and not pago_to and ejercicio:
    pago_from = datetime(ejercicio, 1, 1)
    pago_to = datetime(ejercicio, 12, 31)

  tipo = str(params.get("tipo") or "ambos").strip().lower()
  if tipo not in ("ambos", "basura", "licencia"):
    tipo = "ambos"

  licencia_id = params.get("licenciaId")
  licencia_from = params.get("licenciaFrom")
  licencia_to = params.get("licenciaTo")

  def to_int(value: Any) -> Optional[int]:
    if value in (None, "", "null"):
      return None
    s = str(value).strip()
    if not s:
      return None
    if s.isdigit():
      try:
        return int(s)
      except Exception:
        return None
    try:
      return int(float(s))
    except Exception:
      return None

  return {
    "tipo": tipo,
    "ejercicio": ejercicio,
    "pagoFrom": pago_from,
    "pagoTo": pago_to,
    "licenciaId": to_int(licencia_id),
    "licenciaFrom": to_int(licencia_from),
    "licenciaTo": to_int(licencia_to),
  }


def _saneamiento_ambiental_filters(params: Dict[str, Any]) -> Dict[str, Any]:
  ejercicio = int(params["ejercicio"]) if params.get("ejercicio") not in (None, "", "null") else None
  pago_from = _parse_date(params.get("pagoFrom"))
  pago_to = _parse_date(params.get("pagoTo"))
  if not pago_from and not pago_to and ejercicio:
    pago_from = datetime(ejercicio, 1, 1)
    pago_to = datetime(ejercicio, 12, 31)
  licencia_id = params.get("licenciaId")

  def to_int(value: Any) -> Optional[int]:
    if value in (None, "", "null"):
      return None
    s = str(value).strip()
    if not s:
      return None
    if s.isdigit():
      try:
        return int(s)
      except Exception:
        return None
    try:
      return int(float(s))
    except Exception:
      return None

  return {"ejercicio": ejercicio, "pagoFrom": pago_from, "pagoTo": pago_to, "licenciaId": to_int(licencia_id)}


_SAN_AMB_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @DerechoFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;
DECLARE @Limit int = ?;
DECLARE @Offset int = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    r.ReciboPredioId,
    r.CveFecAsi,
    r.ContriRec,
    r.RFCRecibo,
    r.ReciboObservaciones
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 6100
    AND r.ReciboTramiteId = 5
    AND r.EdoRec = 'A'
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
    AND (
      @LicenciaId IS NULL OR EXISTS (
        SELECT 1
        FROM TESANEAMIENTOAMBIENTAL sa1
        WHERE sa1.CveFteMT = r.CveFteMT
          AND sa1.SaneamientoAmbientalID = r.ReciboPredioId
          AND sa1.LicenciasFuncionamientoId = @LicenciaId
      )
    )
),
page AS (
  SELECT *
  FROM base
  ORDER BY CveFecAsi DESC, CveSerFol DESC, CveFolio DESC
  OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY
),
det AS (
  SELECT
    d.CveFteMT,
    d.CveSerFol,
    d.CveFolio,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @DerechoFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Derecho,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @ActualizacionesFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Actualizaciones,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @RecargosFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Recargos
  FROM COQRECIBODETALLE d
  INNER JOIN page r
    ON r.CveFteMT = d.CveFteMT
   AND r.CveSerFol = d.CveSerFol
   AND r.CveFolio = d.CveFolio
  GROUP BY d.CveFteMT, d.CveSerFol, d.CveFolio
)
SELECT
  CAST(r.ReciboPredioId AS int) AS [PredioId],
  CAST(sa.LicenciasFuncionamientoId AS int) AS [No. Licencia],
  RTRIM(r.CveSerFol COLLATE DATABASE_DEFAULT) AS [Serie],
  CAST(r.CveFolio AS int) AS [Folio],
  r.CveFecAsi AS [Fecha],
  RTRIM(COALESCE(r.ContriRec COLLATE DATABASE_DEFAULT, '')) AS [Nombre],
  RTRIM(COALESCE(r.RFCRecibo COLLATE DATABASE_DEFAULT, '')) AS [RFC],
  RTRIM(COALESCE(r.ReciboObservaciones COLLATE DATABASE_DEFAULT, '')) AS [Observaciones],
  RTRIM(LTRIM(CONCAT(
    COALESCE(p.PredioCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumExt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' ', RTRIM(CAST(p.PredioNumExt AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumInt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' INT ', RTRIM(CAST(p.PredioNumInt AS varchar(50)))) ELSE '' END,
    CASE WHEN p.RegionId IS NOT NULL THEN CONCAT(' REG ', CAST(p.RegionId AS varchar(20))) ELSE '' END,
    CASE WHEN p.ManzanaId IS NOT NULL THEN CONCAT(' MZA ', CAST(p.ManzanaId AS varchar(20))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioLote AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' LOTE ', RTRIM(CAST(p.PredioLote AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNivel AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' NIV ', RTRIM(CAST(p.PredioNivel AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioDepartamento AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' DEPTO ', RTRIM(CAST(p.PredioDepartamento AS varchar(50)))) ELSE '' END
  ))) AS [Domicilio Licencia],
  RTRIM(LTRIM(CONCAT(
    COALESCE(lf.LicenciasFuncionamientoEstablecimientoCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(lf.LicenciasFuncionamientoEstablecimientoEntreLasCalles COLLATE DATABASE_DEFAULT, '')), '') IS NOT NULL
      THEN CONCAT(' ', RTRIM(lf.LicenciasFuncionamientoEstablecimientoEntreLasCalles COLLATE DATABASE_DEFAULT))
      ELSE ''
    END
  ))) AS [Domicilio Local],
  RTRIM(COALESCE(g.GirosComercialesDescripcion COLLATE DATABASE_DEFAULT, '')) AS [Giro],
  TRY_CAST(lf.LicenciasFuncionamientoNumeroControlBasura AS int) AS [No. Cuartos],
  RTRIM(
    CONCAT(
      COALESCE(pago.PeriodoIni, ''),
      CASE WHEN pago.PeriodoFin IS NULL THEN '' ELSE CONCAT(' - ', pago.PeriodoFin) END
    )
  ) AS [Periodo pagado],
  CAST(COALESCE(det.Derecho, 0) AS decimal(18,2)) AS [Derecho],
  CAST(COALESCE(det.Actualizaciones, 0) AS decimal(18,2)) AS [Actualizaciones],
  CAST(COALESCE(det.Recargos, 0) AS decimal(18,2)) AS [Recargos],
  CAST(COALESCE(det.Derecho, 0) + COALESCE(det.Actualizaciones, 0) + COALESCE(det.Recargos, 0) AS decimal(18,2)) AS [Total]
FROM page r
LEFT JOIN TESANEAMIENTOAMBIENTAL sa
  ON sa.CveFteMT = r.CveFteMT
 AND sa.SaneamientoAmbientalID = r.ReciboPredioId
LEFT JOIN TELICENCIASFUNCIONAMIENTO lf
  ON lf.CveFteMT = r.CveFteMT
 AND lf.LicenciasFuncionamientoId = sa.LicenciasFuncionamientoId
LEFT JOIN AlPredio p
  ON p.CveFteMT = r.CveFteMT
 AND p.PredioId = lf.PredioId
LEFT JOIN TEGIROSCOMERCIALES g
  ON g.CveFteMT = r.CveFteMT
 AND g.GirosComercialesId = lf.LicenciasFuncionamientoGiroComercialPrincipalId
OUTER APPLY (
  SELECT TOP 1
    CONCAT(
      CAST(psa.SPagoSaneamientoAmbientalInicialEjercicio AS varchar(4)),
      '/',
      RIGHT('00' + CAST(psa.SPagoSaneamientoAmbientalInicialPeriodo AS varchar(2)), 2)
    ) AS PeriodoIni,
    CONCAT(
      CAST(psa.SPagoSaneamientoAmbientalFinalEjercicio AS varchar(4)),
      '/',
      RIGHT('00' + CAST(psa.SPagoSaneamientoAmbientalFinalPeriodo AS varchar(2)), 2)
    ) AS PeriodoFin
  FROM TESPAGOSANEAMIENTOAMBIENTAL psa
  WHERE psa.CveFteMT = r.CveFteMT
    AND psa.SaneamientoAmbientalID = r.ReciboPredioId
    AND psa.SPagoSaneamientoAmbientalReciboSerie = RTRIM(r.CveSerFol)
    AND psa.SPagoSaneamientoAmbientalReciboFolio = CAST(r.CveFolio AS decimal(18,0))
  ORDER BY psa.SPagoSaneamientoAmbientalId DESC
) pago
LEFT JOIN det
  ON det.CveFteMT = r.CveFteMT
 AND det.CveSerFol = r.CveSerFol
 AND det.CveFolio = r.CveFolio
ORDER BY r.CveFecAsi DESC, r.CveSerFol DESC, r.CveFolio DESC
OPTION (RECOMPILE);
"""


_SAN_AMB_TOTALS_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @DerechoFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT r.CveFteMT, r.CveSerFol, r.CveFolio, r.ReciboPredioId
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 6100
    AND r.ReciboTramiteId = 5
    AND r.EdoRec = 'A'
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
    AND (
      @LicenciaId IS NULL OR EXISTS (
        SELECT 1
        FROM TESANEAMIENTOAMBIENTAL sa1
        WHERE sa1.CveFteMT = r.CveFteMT
          AND sa1.SaneamientoAmbientalID = r.ReciboPredioId
          AND sa1.LicenciasFuncionamientoId = @LicenciaId
      )
    )
)
SELECT
  CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @DerechoFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS [Derecho],
  CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @ActualizacionesFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS [Actualizaciones],
  CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @RecargosFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS [Recargos],
  CAST(COALESCE(SUM(CASE WHEN d.CveFteIng IN (@DerechoFteIng, @ActualizacionesFteIng, @RecargosFteIng) THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS [Total]
FROM base r
LEFT JOIN COQRECIBODETALLE d
  ON d.CveFteMT = r.CveFteMT
 AND d.CveSerFol = r.CveSerFol
 AND d.CveFolio = r.CveFolio
OPTION (RECOMPILE);
"""

_SAN_AMB_MONTHLY_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @DerechoFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    r.ReciboPredioId,
    r.CveFecAsi
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 6100
    AND r.ReciboTramiteId = 5
    AND r.EdoRec = 'A'
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
    AND (
      @LicenciaId IS NULL OR EXISTS (
        SELECT 1
        FROM TESANEAMIENTOAMBIENTAL sa1
        WHERE sa1.CveFteMT = r.CveFteMT
          AND sa1.SaneamientoAmbientalID = r.ReciboPredioId
          AND sa1.LicenciasFuncionamientoId = @LicenciaId
      )
    )
),
det AS (
  SELECT
    d.CveFteMT,
    d.CveSerFol,
    d.CveFolio,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @DerechoFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Derecho,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @ActualizacionesFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Actualizaciones,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @RecargosFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Recargos
  FROM COQRECIBODETALLE d
  INNER JOIN base r
    ON r.CveFteMT = d.CveFteMT
   AND r.CveSerFol = d.CveSerFol
   AND r.CveFolio = d.CveFolio
  GROUP BY d.CveFteMT, d.CveSerFol, d.CveFolio
)
SELECT
  CAST(YEAR(r.CveFecAsi) AS int) AS Ejercicio,
  CAST(MONTH(r.CveFecAsi) AS int) AS Mes,
  CONCAT(
    CAST(YEAR(r.CveFecAsi) AS varchar(4)),
    '-',
    RIGHT('00' + CAST(MONTH(r.CveFecAsi) AS varchar(2)), 2)
  ) AS Periodo,
  CAST(COUNT(1) AS int) AS Recibos,
  CAST(COALESCE(COUNT(DISTINCT sa.LicenciasFuncionamientoId), 0) AS int) AS Licencias,
  CAST(COALESCE(SUM(COALESCE(det.Derecho, 0)), 0) AS decimal(18,2)) AS Derecho,
  CAST(COALESCE(SUM(COALESCE(det.Actualizaciones, 0)), 0) AS decimal(18,2)) AS Actualizaciones,
  CAST(COALESCE(SUM(COALESCE(det.Recargos, 0)), 0) AS decimal(18,2)) AS Recargos,
  CAST(COALESCE(SUM(COALESCE(det.Derecho, 0) + COALESCE(det.Actualizaciones, 0) + COALESCE(det.Recargos, 0)), 0) AS decimal(18,2)) AS Total
FROM base r
LEFT JOIN TESANEAMIENTOAMBIENTAL sa
  ON sa.CveFteMT = r.CveFteMT
 AND sa.SaneamientoAmbientalID = r.ReciboPredioId
LEFT JOIN det
  ON det.CveFteMT = r.CveFteMT
 AND det.CveSerFol = r.CveSerFol
 AND det.CveFolio = r.CveFolio
GROUP BY YEAR(r.CveFecAsi), MONTH(r.CveFecAsi)
ORDER BY YEAR(r.CveFecAsi) DESC, MONTH(r.CveFecAsi) DESC
OPTION (RECOMPILE);
"""

_SAN_AMB_CANCELADOS_COUNT_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @LicenciaId decimal(18,0) = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

SELECT
  CAST(COALESCE(COUNT(1), 0) AS int) AS Cancelados
FROM COQRECIBOS r
WHERE r.CveFteMT = @CveFteMT
  AND r.ReciboGrupoId = 6100
  AND r.ReciboTramiteId = 5
  AND r.EdoRec = 'C'
  AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
  AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
  AND (
    @LicenciaId IS NULL OR EXISTS (
      SELECT 1
      FROM TESANEAMIENTOAMBIENTAL sa1
      WHERE sa1.CveFteMT = r.CveFteMT
        AND sa1.SaneamientoAmbientalID = r.ReciboPredioId
        AND sa1.LicenciasFuncionamientoId = @LicenciaId
    )
  )
OPTION (RECOMPILE);
"""

_SAN_AMB_CANCELADOS_LIST_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @Limit int = ?;
DECLARE @Offset int = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    r.ReciboPredioId,
    r.CveFecAsi,
    r.FecCanRecibo,
    r.UsuCanRecibo,
    r.ConceptoCanRecibo,
    r.ContriRec,
    r.RFCRecibo,
    r.ReciboObservaciones
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 6100
    AND r.ReciboTramiteId = 5
    AND r.EdoRec = 'C'
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
    AND (
      @LicenciaId IS NULL OR EXISTS (
        SELECT 1
        FROM TESANEAMIENTOAMBIENTAL sa1
        WHERE sa1.CveFteMT = r.CveFteMT
          AND sa1.SaneamientoAmbientalID = r.ReciboPredioId
          AND sa1.LicenciasFuncionamientoId = @LicenciaId
      )
    )
),
page AS (
  SELECT
    r.*
  FROM base r
  ORDER BY COALESCE(r.FecCanRecibo, r.CveFecAsi) DESC, r.CveSerFol DESC, r.CveFolio DESC
  OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY
)
SELECT
  CAST(r.ReciboPredioId AS int) AS [Padrón],
  CAST(sa.LicenciasFuncionamientoId AS int) AS [No. Licencia],
  RTRIM(r.CveSerFol COLLATE DATABASE_DEFAULT) AS [Serie],
  CAST(r.CveFolio AS int) AS [Folio],
  r.CveFecAsi AS [Fecha],
  r.FecCanRecibo AS [Fecha cancelación],
  RTRIM(COALESCE(r.UsuCanRecibo COLLATE DATABASE_DEFAULT, '')) AS [Usuario cancelación],
  RTRIM(COALESCE(r.ConceptoCanRecibo COLLATE DATABASE_DEFAULT, '')) AS [Motivo cancelación],
  RTRIM(COALESCE(r.ContriRec COLLATE DATABASE_DEFAULT, '')) AS [Nombre],
  RTRIM(COALESCE(r.RFCRecibo COLLATE DATABASE_DEFAULT, '')) AS [RFC],
  RTRIM(COALESCE(r.ReciboObservaciones COLLATE DATABASE_DEFAULT, '')) AS [Observaciones]
FROM page r
LEFT JOIN TESANEAMIENTOAMBIENTAL sa
  ON sa.CveFteMT = r.CveFteMT
 AND sa.SaneamientoAmbientalID = r.ReciboPredioId
ORDER BY COALESCE(r.FecCanRecibo, r.CveFecAsi) DESC, r.CveSerFol DESC, r.CveFolio DESC
OPTION (RECOMPILE);
"""

_SAN_AMB_EXPORT_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @DerechoFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;
DECLARE @MaxRows int = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    r.ReciboPredioId,
    r.CveFecAsi,
    r.ContriRec,
    r.RFCRecibo,
    r.ReciboObservaciones
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 6100
    AND r.ReciboTramiteId = 5
    AND r.EdoRec = 'A'
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
    AND (
      @LicenciaId IS NULL OR EXISTS (
        SELECT 1
        FROM TESANEAMIENTOAMBIENTAL sa1
        WHERE sa1.CveFteMT = r.CveFteMT
          AND sa1.SaneamientoAmbientalID = r.ReciboPredioId
          AND sa1.LicenciasFuncionamientoId = @LicenciaId
      )
    )
),
page AS (
  SELECT TOP (@MaxRows) *
  FROM base
  ORDER BY CveFecAsi DESC, CveSerFol DESC, CveFolio DESC
),
det AS (
  SELECT
    d.CveFteMT,
    d.CveSerFol,
    d.CveFolio,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @DerechoFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Derecho,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @ActualizacionesFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Actualizaciones,
    CAST(COALESCE(SUM(CASE WHEN d.CveFteIng = @RecargosFteIng THEN d.ReciboDetImpAntesDev ELSE 0 END), 0) AS decimal(18,2)) AS Recargos
  FROM COQRECIBODETALLE d
  INNER JOIN page r
    ON r.CveFteMT = d.CveFteMT
   AND r.CveSerFol = d.CveSerFol
   AND r.CveFolio = d.CveFolio
  GROUP BY d.CveFteMT, d.CveSerFol, d.CveFolio
)
SELECT
  CAST(r.ReciboPredioId AS int) AS [Padrón],
  CAST(sa.LicenciasFuncionamientoId AS int) AS [No. Licencia],
  RTRIM(r.CveSerFol COLLATE DATABASE_DEFAULT) AS [Serie],
  CAST(r.CveFolio AS int) AS [Folio],
  r.CveFecAsi AS [Fecha],
  RTRIM(COALESCE(r.ContriRec COLLATE DATABASE_DEFAULT, '')) AS [Nombre],
  RTRIM(COALESCE(r.RFCRecibo COLLATE DATABASE_DEFAULT, '')) AS [RFC],
  RTRIM(COALESCE(r.ReciboObservaciones COLLATE DATABASE_DEFAULT, '')) AS [Observaciones],
  RTRIM(LTRIM(CONCAT(
    COALESCE(p.PredioCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumExt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' ', RTRIM(CAST(p.PredioNumExt AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumInt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' INT ', RTRIM(CAST(p.PredioNumInt AS varchar(50)))) ELSE '' END,
    CASE WHEN p.RegionId IS NOT NULL THEN CONCAT(' REG ', CAST(p.RegionId AS varchar(20))) ELSE '' END,
    CASE WHEN p.ManzanaId IS NOT NULL THEN CONCAT(' MZA ', CAST(p.ManzanaId AS varchar(20))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioLote AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' LOTE ', RTRIM(CAST(p.PredioLote AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNivel AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' NIV ', RTRIM(CAST(p.PredioNivel AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioDepartamento AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' DEPTO ', RTRIM(CAST(p.PredioDepartamento AS varchar(50)))) ELSE '' END
  ))) AS [Domicilio Licencia],
  RTRIM(LTRIM(CONCAT(
    COALESCE(lf.LicenciasFuncionamientoEstablecimientoCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(lf.LicenciasFuncionamientoEstablecimientoEntreLasCalles COLLATE DATABASE_DEFAULT, '')), '') IS NOT NULL
      THEN CONCAT(' ', RTRIM(lf.LicenciasFuncionamientoEstablecimientoEntreLasCalles COLLATE DATABASE_DEFAULT))
      ELSE ''
    END
  ))) AS [Domicilio Local],
  RTRIM(COALESCE(g.GirosComercialesDescripcion COLLATE DATABASE_DEFAULT, '')) AS [Giro],
  TRY_CAST(lf.LicenciasFuncionamientoNumeroControlBasura AS int) AS [No. Cuartos],
  RTRIM(
    CONCAT(
      COALESCE(pago.PeriodoIni, ''),
      CASE WHEN pago.PeriodoFin IS NULL THEN '' ELSE CONCAT(' - ', pago.PeriodoFin) END
    )
  ) AS [Periodo pagado],
  CAST(COALESCE(det.Derecho, 0) AS decimal(18,2)) AS [Derecho],
  CAST(COALESCE(det.Actualizaciones, 0) AS decimal(18,2)) AS [Actualizaciones],
  CAST(COALESCE(det.Recargos, 0) AS decimal(18,2)) AS [Recargos],
  CAST(COALESCE(det.Derecho, 0) + COALESCE(det.Actualizaciones, 0) + COALESCE(det.Recargos, 0) AS decimal(18,2)) AS [Total]
FROM page r
LEFT JOIN TESANEAMIENTOAMBIENTAL sa
  ON sa.CveFteMT = r.CveFteMT
 AND sa.SaneamientoAmbientalID = r.ReciboPredioId
LEFT JOIN TELICENCIASFUNCIONAMIENTO lf
  ON lf.CveFteMT = r.CveFteMT
 AND lf.LicenciasFuncionamientoId = sa.LicenciasFuncionamientoId
LEFT JOIN AlPredio p
  ON p.CveFteMT = r.CveFteMT
 AND p.PredioId = lf.PredioId
LEFT JOIN TEGIROSCOMERCIALES g
  ON g.CveFteMT = r.CveFteMT
 AND g.GirosComercialesId = lf.LicenciasFuncionamientoGiroComercialPrincipalId
OUTER APPLY (
  SELECT TOP 1
    CONCAT(
      CAST(psa.SPagoSaneamientoAmbientalInicialEjercicio AS varchar(4)),
      '/',
      RIGHT('00' + CAST(psa.SPagoSaneamientoAmbientalInicialPeriodo AS varchar(2)), 2)
    ) AS PeriodoIni,
    CONCAT(
      CAST(psa.SPagoSaneamientoAmbientalFinalEjercicio AS varchar(4)),
      '/',
      RIGHT('00' + CAST(psa.SPagoSaneamientoAmbientalFinalPeriodo AS varchar(2)), 2)
    ) AS PeriodoFin
  FROM TESPAGOSANEAMIENTOAMBIENTAL psa
  WHERE psa.CveFteMT = r.CveFteMT
    AND psa.SaneamientoAmbientalID = r.ReciboPredioId
    AND psa.SPagoSaneamientoAmbientalReciboSerie = RTRIM(r.CveSerFol)
    AND psa.SPagoSaneamientoAmbientalReciboFolio = CAST(r.CveFolio AS decimal(18,0))
  ORDER BY psa.SPagoSaneamientoAmbientalId DESC
) pago
LEFT JOIN det
  ON det.CveFteMT = r.CveFteMT
 AND det.CveSerFol = r.CveSerFol
 AND det.CveFolio = r.CveFolio
ORDER BY r.CveFecAsi DESC, r.CveSerFol DESC, r.CveFolio DESC
OPTION (RECOMPILE);
"""


_LIC_FUNC_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @Tipo varchar(16) = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @LicenciaFrom decimal(18,0) = ?;
DECLARE @LicenciaTo decimal(18,0) = ?;
DECLARE @LicenciaPrincipalFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;
DECLARE @Ejercicio int = ?;
DECLARE @UMAMxn decimal(18,6) = ?;
DECLARE @Limit int = ?;
DECLARE @Offset int = ?;

DECLARE @IncluirLicencia bit = CASE WHEN @Tipo <> 'basura' THEN 1 ELSE 0 END;
DECLARE @IncluirBasura bit = CASE WHEN @Tipo <> 'licencia' THEN 1 ELSE 0 END;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

SELECT
  CAST(r.ReciboPredioId AS int) AS [No. Licencia],
  RTRIM(r.CveSerFol COLLATE DATABASE_DEFAULT) AS [Serie],
  CAST(r.CveFolio AS int) AS [Folio],
  r.CveFecAsi AS [Fecha],
  RTRIM(COALESCE(r.ContriRec COLLATE DATABASE_DEFAULT, '')) AS [Nombre],
  RTRIM(COALESCE(r.RFCRecibo COLLATE DATABASE_DEFAULT, '')) AS [RFC],
  RTRIM(COALESCE(r.ReciboObservaciones COLLATE DATABASE_DEFAULT, '')) AS [Observaciones],
  RTRIM(LTRIM(CONCAT(
    COALESCE(p.PredioCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumExt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' ', RTRIM(CAST(p.PredioNumExt AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumInt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' INT ', RTRIM(CAST(p.PredioNumInt AS varchar(50)))) ELSE '' END,
    CASE WHEN p.RegionId IS NOT NULL THEN CONCAT(' REG ', CAST(p.RegionId AS varchar(20))) ELSE '' END,
    CASE WHEN p.ManzanaId IS NOT NULL THEN CONCAT(' MZA ', CAST(p.ManzanaId AS varchar(20))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioLote AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' LOTE ', RTRIM(CAST(p.PredioLote AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNivel AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' NIV ', RTRIM(CAST(p.PredioNivel AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioDepartamento AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' DEPTO ', RTRIM(CAST(p.PredioDepartamento AS varchar(50)))) ELSE '' END
  ))) AS [Domicilio Licencia],
  RTRIM(LTRIM(CONCAT(
    COALESCE(lf.LicenciasFuncionamientoEstablecimientoCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(lf.LicenciasFuncionamientoEstablecimientoEntreLasCalles COLLATE DATABASE_DEFAULT, '')), '') IS NOT NULL
      THEN CONCAT(' ', RTRIM(lf.LicenciasFuncionamientoEstablecimientoEntreLasCalles COLLATE DATABASE_DEFAULT))
      ELSE ''
    END
  ))) AS [Domicilio Local],
  'Licencia' AS [Tipo Establecimiento],
  RTRIM(COALESCE(g.GirosComercialesDescripcion COLLATE DATABASE_DEFAULT, '')) AS [Giro],
  CAST(COALESCE(tb.TarifasLicenciasBasuraImporteLicenciaUMA, 0) * COALESCE(@UMAMxn, 0) AS decimal(18,2)) AS [Base Licencia Nueva],
  CAST(COALESCE(tb.TarifasLicenciasBasuraImporteRenovacionUMA, 0) * COALESCE(@UMAMxn, 0) AS decimal(18,2)) AS [Base Licencia Renovación],
  CAST(COALESCE(tb.TarifasLicenciasBasuraImporteBasuraUMA, 0) * COALESCE(@UMAMxn, 0) AS decimal(18,2)) AS [Base Basura Nueva],
  CASE WHEN @Tipo = 'basura' THEN 'Basura' WHEN @Tipo = 'licencia' THEN 'Licencia' ELSE 'Ambos' END AS [Tipo],
  det.Licencia AS [Tarifa Licencia],
  det.BasuraTarifa AS [Tarifa Basura],
  det.Licencia AS [Licencia],
  det.LicRenovacion AS [Lic Renovación],
  det.Basura AS [Basura],
  det.Actualizaciones AS [Actualizaciones],
  det.Recargos AS [Recargos],
  det.Otros AS [Otros],
  CAST(det.Licencia + det.LicRenovacion + det.Basura + det.Actualizaciones + det.Recargos + det.Otros AS decimal(18,2)) AS [Total]
FROM COQRECIBOS r
LEFT JOIN TELICENCIASFUNCIONAMIENTO lf
  ON lf.CveFteMT = r.CveFteMT
 AND lf.LicenciasFuncionamientoId = r.ReciboPredioId
LEFT JOIN AlPredio p
  ON p.CveFteMT = r.CveFteMT
 AND p.PredioId = lf.PredioId
LEFT JOIN TEGIROSCOMERCIALES g
  ON g.CveFteMT = r.CveFteMT
 AND g.GirosComercialesId = lf.LicenciasFuncionamientoGiroComercialPrincipalId
LEFT JOIN TETARIFASLICENCIASBASURA tb
  ON tb.CveFteMT = r.CveFteMT
 AND tb.TarifasLicenciasBasuraGiroId = lf.LicenciasFuncionamientoGiroComercialPrincipalId
 AND tb.TarifasLicenciasBasuraEjercicio = COALESCE(@Ejercicio, YEAR(r.CveFecAsi))
OUTER APPLY (
  SELECT
    CAST(COALESCE(SUM(CASE WHEN @IncluirLicencia = 1 AND x.CveFteIng = cfg.FteIngLicencia AND x.CveFteIng <> cfg.FteIngRenovacion THEN x.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Licencia,
    CAST(COALESCE(SUM(CASE WHEN @IncluirLicencia = 1 AND x.CveFteIng = cfg.FteIngRenovacion THEN x.Importe ELSE 0 END), 0) AS decimal(18,2)) AS LicRenovacion,
    CAST(COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND x.CveFteIng = cfg.FteIngBasura THEN x.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Basura,
    CAST(
      COALESCE(
        SUM(
          CASE
            WHEN @IncluirBasura = 1 AND x.CveFteIng = cfg.FteIngBasura
              THEN CASE WHEN x.Cnt > 1 THEN (x.Importe / NULLIF(CAST(x.Cnt AS decimal(18,6)), 0)) * 12 ELSE x.Importe END
            ELSE 0
          END
        ),
        0
      )
    AS decimal(18,2)) AS BasuraTarifa,
    CAST(COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND x.CveFteIng = @ActualizacionesFteIng THEN x.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Actualizaciones,
    CAST(COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND x.CveFteIng = @RecargosFteIng THEN x.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Recargos,
    CAST(COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND x.CveFteIng NOT IN (cfg.FteIngLicencia, cfg.FteIngRenovacion, cfg.FteIngBasura, @ActualizacionesFteIng, @RecargosFteIng) THEN x.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Otros
  FROM (
    SELECT
      d.CveFteIng,
      CAST(COALESCE(SUM(d.ReciboDetImpAntesDev), 0) AS decimal(18,2)) AS Importe,
      COUNT(1) AS Cnt
    FROM COQRECIBODETALLE d
    WHERE d.CveFteMT = r.CveFteMT
      AND d.CveSerFol = r.CveSerFol
      AND d.CveFolio = r.CveFolio
    GROUP BY d.CveFteIng
  ) x
  CROSS JOIN (
    SELECT
      CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngLicencia, @LicenciaPrincipalFteIng, -1) AS decimal(18,0)) AS FteIngLicencia,
      CAST(COALESCE(tb.CveFteIngRenovacionId, -1) AS decimal(18,0)) AS FteIngRenovacion,
      CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngBasura, -1) AS decimal(18,0)) AS FteIngBasura
  ) cfg
) det
WHERE r.CveFteMT = @CveFteMT
  AND r.ReciboGrupoId = 68
  AND r.ReciboTramiteId = 12
  AND r.EdoRec = 'A'
  AND (@LicenciaId IS NULL OR r.ReciboPredioId = @LicenciaId)
  AND (@LicenciaFrom IS NULL OR r.ReciboPredioId >= @LicenciaFrom)
  AND (@LicenciaTo IS NULL OR r.ReciboPredioId <= @LicenciaTo)
  AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
  AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
ORDER BY r.CveFecAsi DESC, r.CveSerFol DESC, r.CveFolio DESC
OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY
OPTION (RECOMPILE);
"""

_LIC_FUNC_TOTALS_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @Tipo varchar(16) = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @LicenciaFrom decimal(18,0) = ?;
DECLARE @LicenciaTo decimal(18,0) = ?;
DECLARE @LicenciaPrincipalFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;
DECLARE @Ejercicio int = ?;
DECLARE @UMAMxn decimal(18,6) = ?;

DECLARE @IncluirLicencia bit = CASE WHEN @Tipo <> 'basura' THEN 1 ELSE 0 END;
DECLARE @IncluirBasura bit = CASE WHEN @Tipo <> 'licencia' THEN 1 ELSE 0 END;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    r.ReciboPredioId,
    r.CveFecAsi
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 68
    AND r.ReciboTramiteId = 12
    AND r.EdoRec = 'A'
    AND (@LicenciaId IS NULL OR r.ReciboPredioId = @LicenciaId)
    AND (@LicenciaFrom IS NULL OR r.ReciboPredioId >= @LicenciaFrom)
    AND (@LicenciaTo IS NULL OR r.ReciboPredioId <= @LicenciaTo)
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
),
cfg AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngLicencia, @LicenciaPrincipalFteIng, -1) AS decimal(18,0)) AS FteIngLicencia,
    CAST(COALESCE(tb.CveFteIngRenovacionId, -1) AS decimal(18,0)) AS FteIngRenovacion,
    CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngBasura, -1) AS decimal(18,0)) AS FteIngBasura
  FROM base r
  LEFT JOIN TELICENCIASFUNCIONAMIENTO lf
    ON lf.CveFteMT = r.CveFteMT
   AND lf.LicenciasFuncionamientoId = r.ReciboPredioId
  LEFT JOIN TETARIFASLICENCIASBASURA tb
    ON tb.CveFteMT = r.CveFteMT
   AND tb.TarifasLicenciasBasuraGiroId = lf.LicenciasFuncionamientoGiroComercialPrincipalId
   AND tb.TarifasLicenciasBasuraEjercicio = COALESCE(@Ejercicio, YEAR(r.CveFecAsi))
),
det AS (
  SELECT
    d.CveFteMT,
    d.CveSerFol,
    d.CveFolio,
    d.CveFteIng,
    CAST(COALESCE(SUM(d.ReciboDetImpAntesDev), 0) AS decimal(18,2)) AS Importe,
    COUNT(1) AS Cnt
  FROM COQRECIBODETALLE d
  INNER JOIN base r
    ON r.CveFteMT = d.CveFteMT
   AND r.CveSerFol = d.CveSerFol
   AND r.CveFolio = d.CveFolio
  GROUP BY d.CveFteMT, d.CveSerFol, d.CveFolio, d.CveFteIng
)
SELECT
  CAST(
    COALESCE(
      SUM(
        CASE
          WHEN @IncluirLicencia = 1 AND d.CveFteIng = cfg.FteIngLicencia AND d.CveFteIng <> cfg.FteIngRenovacion THEN d.Importe
          ELSE 0
        END
      ),
      0
    )
  AS decimal(18,2)) AS [Licencia],
  CAST(
    COALESCE(
      SUM(CASE WHEN @IncluirLicencia = 1 AND d.CveFteIng = cfg.FteIngRenovacion THEN d.Importe ELSE 0 END),
      0
    )
  AS decimal(18,2)) AS [Lic Renovación],
  CAST(
    COALESCE(
      SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = cfg.FteIngBasura THEN d.Importe ELSE 0 END),
      0
    )
  AS decimal(18,2)) AS [Basura],
  CAST(
    COALESCE(
      SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = @ActualizacionesFteIng THEN d.Importe ELSE 0 END),
      0
    )
  AS decimal(18,2)) AS [Actualizaciones],
  CAST(
    COALESCE(
      SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = @RecargosFteIng THEN d.Importe ELSE 0 END),
      0
    )
  AS decimal(18,2)) AS [Recargos],
  CAST(
    COALESCE(
      SUM(
        CASE
          WHEN @IncluirBasura = 1
           AND d.CveFteIng NOT IN (cfg.FteIngLicencia, cfg.FteIngRenovacion, cfg.FteIngBasura, @ActualizacionesFteIng, @RecargosFteIng)
            THEN d.Importe
          ELSE 0
        END
      ),
      0
    )
  AS decimal(18,2)) AS [Otros],
  CAST(
    COALESCE(SUM(CASE WHEN @IncluirLicencia = 1 AND d.CveFteIng = cfg.FteIngLicencia AND d.CveFteIng <> cfg.FteIngRenovacion THEN d.Importe ELSE 0 END), 0) +
    COALESCE(SUM(CASE WHEN @IncluirLicencia = 1 AND d.CveFteIng = cfg.FteIngRenovacion THEN d.Importe ELSE 0 END), 0) +
    COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = cfg.FteIngBasura THEN d.Importe ELSE 0 END), 0) +
    COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = @ActualizacionesFteIng THEN d.Importe ELSE 0 END), 0) +
    COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = @RecargosFteIng THEN d.Importe ELSE 0 END), 0) +
    COALESCE(
      SUM(
        CASE
          WHEN @IncluirBasura = 1
           AND d.CveFteIng NOT IN (cfg.FteIngLicencia, cfg.FteIngRenovacion, cfg.FteIngBasura, @ActualizacionesFteIng, @RecargosFteIng)
            THEN d.Importe
          ELSE 0
        END
      ),
      0
    )
  AS decimal(18,2)) AS [Total]
FROM base r
LEFT JOIN cfg
  ON cfg.CveFteMT = r.CveFteMT
 AND cfg.CveSerFol = r.CveSerFol
 AND cfg.CveFolio = r.CveFolio
LEFT JOIN det d
  ON d.CveFteMT = r.CveFteMT
 AND d.CveSerFol = r.CveSerFol
 AND d.CveFolio = r.CveFolio
OPTION (RECOMPILE);
"""

_LIC_FUNC_CSV = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @Tipo varchar(16) = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @LicenciaFrom decimal(18,0) = ?;
DECLARE @LicenciaTo decimal(18,0) = ?;
DECLARE @LicenciaPrincipalFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;
DECLARE @Ejercicio int = ?;
DECLARE @UMAMxn decimal(18,6) = ?;
DECLARE @MaxRows int = ?;

DECLARE @IncluirLicencia bit = CASE WHEN @Tipo <> 'basura' THEN 1 ELSE 0 END;
DECLARE @IncluirBasura bit = CASE WHEN @Tipo <> 'licencia' THEN 1 ELSE 0 END;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    r.ReciboPredioId,
    r.CveFecAsi,
    r.ContriRec,
    r.RFCRecibo,
    r.ReciboObservaciones
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 68
    AND r.ReciboTramiteId = 12
    AND r.EdoRec = 'A'
    AND (@LicenciaId IS NULL OR r.ReciboPredioId = @LicenciaId)
    AND (@LicenciaFrom IS NULL OR r.ReciboPredioId >= @LicenciaFrom)
    AND (@LicenciaTo IS NULL OR r.ReciboPredioId <= @LicenciaTo)
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
),
page AS (
  SELECT TOP (@MaxRows) *
  FROM base
  ORDER BY CveFecAsi DESC, CveSerFol DESC, CveFolio DESC
),
info AS (
  SELECT
    r.*,
    lf.PredioId AS LicPredioId,
    lf.LicenciasFuncionamientoGiroComercialPrincipalId AS GiroId,
    RTRIM(COALESCE(lf.LicenciasFuncionamientoEstablecimientoCalle COLLATE DATABASE_DEFAULT, '')) AS EstCalle,
    RTRIM(COALESCE(lf.LicenciasFuncionamientoEstablecimientoEntreLasCalles COLLATE DATABASE_DEFAULT, '')) AS EstEntre,
    CAST(COALESCE(tb.TarifasLicenciasBasuraImporteLicenciaUMA, 0) * COALESCE(@UMAMxn, 0) AS decimal(18,2)) AS BaseLicNueva,
    CAST(COALESCE(tb.TarifasLicenciasBasuraImporteRenovacionUMA, 0) * COALESCE(@UMAMxn, 0) AS decimal(18,2)) AS BaseLicRen,
    CAST(COALESCE(tb.TarifasLicenciasBasuraImporteBasuraUMA, 0) * COALESCE(@UMAMxn, 0) AS decimal(18,2)) AS BaseBasNueva,
    CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngLicencia, @LicenciaPrincipalFteIng, -1) AS decimal(18,0)) AS FteIngLicencia,
    CAST(COALESCE(tb.CveFteIngRenovacionId, -1) AS decimal(18,0)) AS FteIngRenovacion,
    CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngBasura, -1) AS decimal(18,0)) AS FteIngBasura
  FROM page r
  LEFT JOIN TELICENCIASFUNCIONAMIENTO lf
    ON lf.CveFteMT = r.CveFteMT
   AND lf.LicenciasFuncionamientoId = r.ReciboPredioId
  LEFT JOIN TETARIFASLICENCIASBASURA tb
    ON tb.CveFteMT = r.CveFteMT
   AND tb.TarifasLicenciasBasuraGiroId = lf.LicenciasFuncionamientoGiroComercialPrincipalId
   AND tb.TarifasLicenciasBasuraEjercicio = COALESCE(@Ejercicio, YEAR(r.CveFecAsi))
),
det AS (
  SELECT
    d.CveFteMT,
    d.CveSerFol,
    d.CveFolio,
    d.CveFteIng,
    CAST(COALESCE(SUM(d.ReciboDetImpAntesDev), 0) AS decimal(18,2)) AS Importe,
    COUNT(1) AS Cnt
  FROM COQRECIBODETALLE d
  INNER JOIN page r
    ON r.CveFteMT = d.CveFteMT
   AND r.CveSerFol = d.CveSerFol
   AND r.CveFolio = d.CveFolio
  GROUP BY d.CveFteMT, d.CveSerFol, d.CveFolio, d.CveFteIng
),
calc AS (
  SELECT
    i.CveFteMT,
    i.CveSerFol,
    i.CveFolio,
    CAST(COALESCE(SUM(CASE WHEN @IncluirLicencia = 1 AND d.CveFteIng = i.FteIngLicencia AND d.CveFteIng <> i.FteIngRenovacion THEN d.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Licencia,
    CAST(COALESCE(SUM(CASE WHEN @IncluirLicencia = 1 AND d.CveFteIng = i.FteIngRenovacion THEN d.Importe ELSE 0 END), 0) AS decimal(18,2)) AS LicRenovacion,
    CAST(COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = i.FteIngBasura THEN d.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Basura,
    CAST(
      COALESCE(
        SUM(
          CASE
            WHEN @IncluirBasura = 1 AND d.CveFteIng = i.FteIngBasura
              THEN CASE WHEN d.Cnt > 1 THEN (d.Importe / NULLIF(CAST(d.Cnt AS decimal(18,6)), 0)) * 12 ELSE d.Importe END
            ELSE 0
          END
        ),
        0
      )
    AS decimal(18,2)) AS BasuraTarifa,
    CAST(COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = @ActualizacionesFteIng THEN d.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Actualizaciones,
    CAST(COALESCE(SUM(CASE WHEN @IncluirBasura = 1 AND d.CveFteIng = @RecargosFteIng THEN d.Importe ELSE 0 END), 0) AS decimal(18,2)) AS Recargos,
    CAST(
      COALESCE(
        SUM(
          CASE
            WHEN @IncluirBasura = 1
             AND d.CveFteIng NOT IN (i.FteIngLicencia, i.FteIngRenovacion, i.FteIngBasura, @ActualizacionesFteIng, @RecargosFteIng)
              THEN d.Importe
            ELSE 0
          END
        ),
        0
      )
    AS decimal(18,2)) AS Otros
  FROM info i
  LEFT JOIN det d
    ON d.CveFteMT = i.CveFteMT
   AND d.CveSerFol = i.CveSerFol
   AND d.CveFolio = i.CveFolio
  GROUP BY i.CveFteMT, i.CveSerFol, i.CveFolio
)
SELECT
  CAST(i.ReciboPredioId AS int) AS [No. Licencia],
  RTRIM(i.CveSerFol COLLATE DATABASE_DEFAULT) AS [Serie],
  CAST(i.CveFolio AS int) AS [Folio],
  i.CveFecAsi AS [Fecha],
  RTRIM(COALESCE(i.ContriRec COLLATE DATABASE_DEFAULT, '')) AS [Nombre],
  RTRIM(COALESCE(i.RFCRecibo COLLATE DATABASE_DEFAULT, '')) AS [RFC],
  RTRIM(COALESCE(i.ReciboObservaciones COLLATE DATABASE_DEFAULT, '')) AS [Observaciones],
  RTRIM(LTRIM(CONCAT(
    COALESCE(p.PredioCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumExt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' ', RTRIM(CAST(p.PredioNumExt AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNumInt AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' INT ', RTRIM(CAST(p.PredioNumInt AS varchar(50)))) ELSE '' END,
    CASE WHEN p.RegionId IS NOT NULL THEN CONCAT(' REG ', CAST(p.RegionId AS varchar(20))) ELSE '' END,
    CASE WHEN p.ManzanaId IS NOT NULL THEN CONCAT(' MZA ', CAST(p.ManzanaId AS varchar(20))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioLote AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' LOTE ', RTRIM(CAST(p.PredioLote AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioNivel AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' NIV ', RTRIM(CAST(p.PredioNivel AS varchar(50)))) ELSE '' END,
    CASE WHEN NULLIF(RTRIM(COALESCE(CAST(p.PredioDepartamento AS varchar(50)), '')), '') IS NOT NULL THEN CONCAT(' DEPTO ', RTRIM(CAST(p.PredioDepartamento AS varchar(50)))) ELSE '' END
  ))) AS [Domicilio Licencia],
  RTRIM(LTRIM(CONCAT(
    COALESCE(i.EstCalle COLLATE DATABASE_DEFAULT, ''),
    CASE WHEN NULLIF(RTRIM(COALESCE(i.EstEntre COLLATE DATABASE_DEFAULT, '')), '') IS NOT NULL
      THEN CONCAT(' ', RTRIM(i.EstEntre COLLATE DATABASE_DEFAULT))
      ELSE ''
    END
  ))) AS [Domicilio Local],
  'Licencia' AS [Tipo Establecimiento],
  RTRIM(COALESCE(g.GirosComercialesDescripcion COLLATE DATABASE_DEFAULT, '')) AS [Giro],
  i.BaseLicNueva AS [Base Licencia Nueva],
  i.BaseLicRen AS [Base Licencia Renovación],
  i.BaseBasNueva AS [Base Basura Nueva],
  CASE WHEN @Tipo = 'basura' THEN 'Basura' WHEN @Tipo = 'licencia' THEN 'Licencia' ELSE 'Ambos' END AS [Tipo],
  c.Licencia AS [Tarifa Licencia],
  c.BasuraTarifa AS [Tarifa Basura],
  c.Licencia AS [Licencia],
  c.LicRenovacion AS [Lic Renovación],
  c.Basura AS [Basura],
  c.Actualizaciones AS [Actualizaciones],
  c.Recargos AS [Recargos],
  c.Otros AS [Otros],
  CAST(c.Licencia + c.LicRenovacion + c.Basura + c.Actualizaciones + c.Recargos + c.Otros AS decimal(18,2)) AS [Total]
FROM info i
LEFT JOIN calc c
  ON c.CveFteMT = i.CveFteMT
 AND c.CveSerFol = i.CveSerFol
 AND c.CveFolio = i.CveFolio
LEFT JOIN AlPredio p
  ON p.CveFteMT = i.CveFteMT
 AND p.PredioId = i.LicPredioId
LEFT JOIN TEGIROSCOMERCIALES g
  ON g.CveFteMT = i.CveFteMT
 AND g.GirosComercialesId = i.GiroId
ORDER BY i.CveFecAsi DESC, i.CveSerFol DESC, i.CveFolio DESC
OPTION (RECOMPILE);
"""
 
_SABANA_PREDIALES_CSV = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @Q varchar(200) = ?;
DECLARE @ClaveCatastral varchar(80) = ?;
DECLARE @ClaveMode varchar(16) = ?;
DECLARE @PredioId decimal(18,0) = ?;
DECLARE @Propietario varchar(200) = ?;
DECLARE @ApellidoPaterno varchar(120) = ?;
DECLARE @ApellidoMaterno varchar(120) = ?;
DECLARE @Nombre varchar(120) = ?;
DECLARE @Calle varchar(200) = ?;
DECLARE @Numero varchar(50) = ?;
DECLARE @Estatus varchar(8) = ?;
DECLARE @Adeudo varchar(16) = ?;
DECLARE @FromAlta datetime = ?;
DECLARE @ToAlta datetime = ?;
DECLARE @EjercicioActual int = ?;
DECLARE @MaxRows int = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    p.CveFteMT,
    p.PredioId,
    RTRIM(COALESCE(p.PredioCveCatastral COLLATE DATABASE_DEFAULT, '')) AS PredioCveCatastral,
    p.PredioAltaFecha,
    RTRIM(COALESCE(p.PredioCalle COLLATE DATABASE_DEFAULT, '')) AS PredioCalle,
    p.PredioCodigoPostal,
    RTRIM(COALESCE(p.PredioNumExt COLLATE DATABASE_DEFAULT, '')) AS PredioNumExt,
    RTRIM(COALESCE(p.PredioNumInt COLLATE DATABASE_DEFAULT, '')) AS PredioNumInt,
    RTRIM(COALESCE(p.PredioStatus COLLATE DATABASE_DEFAULT, '')) AS PredioStatus,
    RTRIM(COALESCE(p.CatastroDatosEscriturales COLLATE DATABASE_DEFAULT, '')) AS CatastroDatosEscriturales,
    CAST(COALESCE(p.PredioTotalTerreno, 0) AS decimal(18,2)) AS PredioTotalTerreno,
    CAST(COALESCE(p.PredioArea, 0) AS decimal(18,2)) AS PredioArea,
    CAST(COALESCE(p.PredioTotalConstruccion, 0) AS decimal(18,2)) AS PredioTotalConstruccion,
    CAST(COALESCE(p.PredioTerrenoImporte, 0) AS decimal(18,2)) AS PredioTerrenoImporte,
    CAST(COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS PredioConstruccionImporte,
    CAST(COALESCE(p.PredioCatastralImporte, 0) AS decimal(18,2)) AS PredioCatastralImporte,
    p.PredioUltimoEjericicioPagado,
    p.PredioUltimoPeriodoPagado,
    RTRIM(COALESCE(p.PredioTipo COLLATE DATABASE_DEFAULT, '')) AS PredioTipo,
    RTRIM(COALESCE(per.NombreCompletoPersona COLLATE DATABASE_DEFAULT, per.RazonSocialPersona COLLATE DATABASE_DEFAULT, '')) AS PropietarioNombre,
    RTRIM(COALESCE(ef.EstadoFisicoNombre COLLATE DATABASE_DEFAULT, '')) AS EstadoFisicoNombre
  FROM AlPredio p
  LEFT JOIN XiPersonas per
    ON per.CveFteMT = p.CveFteMT
   AND per.CvePersona = p.CvePersona
  LEFT JOIN AlEstadoFisico ef
    ON ef.CveFteMT = p.CveFteMT
   AND ef.EstadoFisicoId = p.EstadoFisicoId
  WHERE p.CveFteMT = @CveFteMT
    AND (@FromAlta IS NULL OR p.PredioAltaFecha >= @FromAlta)
    AND (@ToAlta IS NULL OR p.PredioAltaFecha < DATEADD(DAY, 1, @ToAlta))
    AND (@PredioId IS NULL OR p.PredioId = @PredioId)
    AND (@Estatus IS NULL OR RTRIM(COALESCE(p.PredioStatus COLLATE DATABASE_DEFAULT, '')) = @Estatus)
    AND (
      @ClaveCatastral IS NULL OR
      (@ClaveMode = 'exacto' AND RTRIM(COALESCE(p.PredioCveCatastral COLLATE DATABASE_DEFAULT, '')) = @ClaveCatastral) OR
      (@ClaveMode <> 'exacto' AND p.PredioCveCatastral LIKE '%' + @ClaveCatastral + '%')
    )
    AND (@Propietario IS NULL OR per.NombreCompletoPersona LIKE '%' + @Propietario + '%' OR per.RazonSocialPersona LIKE '%' + @Propietario + '%')
    AND (@ApellidoPaterno IS NULL OR per.ApellidoPaternoPersona LIKE '%' + @ApellidoPaterno + '%')
    AND (@ApellidoMaterno IS NULL OR per.ApellidoMaternoPersona LIKE '%' + @ApellidoMaterno + '%')
    AND (@Nombre IS NULL OR per.NombrePersona LIKE '%' + @Nombre + '%')
    AND (@Calle IS NULL OR p.PredioCalle LIKE '%' + @Calle + '%')
    AND (
      @Numero IS NULL OR
      p.PredioNumExt LIKE '%' + @Numero + '%' OR
      p.PredioNumInt LIKE '%' + @Numero + '%'
    )
    AND (
      @Adeudo IS NULL OR @Adeudo = 'todos' OR
      (@Adeudo = 'con' AND COALESCE(p.PredioActualAdeudoImporte, p.PredioAdeudoImporte, p.PredioTotalAdeudoImporte, 0) > 0) OR
      (@Adeudo = 'sin' AND COALESCE(p.PredioActualAdeudoImporte, p.PredioAdeudoImporte, p.PredioTotalAdeudoImporte, 0) <= 0)
    )
    AND (
      @Q IS NULL OR
      p.PredioCveCatastral LIKE '%' + @Q + '%' OR
      p.PredioClavePredial LIKE '%' + @Q + '%' OR
      per.NombreCompletoPersona LIKE '%' + @Q + '%' OR
      per.RazonSocialPersona LIKE '%' + @Q + '%'
    )
),
toprows AS (
  SELECT TOP (@MaxRows) *
  FROM base
  ORDER BY PredioId ASC
),
terr AS (
  SELECT
    t.CveFteMT,
    t.PredioId,
    CAST(COALESCE(SUM(t.PredioTerrenoArea), 0) AS decimal(18,2)) AS TerrenoAreaSum,
    CAST(COALESCE(SUM(t.PredioTerrenoArea * t.PredioTerrenoUnitarioImporte), 0) AS decimal(18,2)) AS TerrenoImporteSum
  FROM ALPREDIOTERRENOS t
  INNER JOIN toprows p
    ON p.CveFteMT = t.CveFteMT
   AND p.PredioId = t.PredioId
  GROUP BY t.CveFteMT, t.PredioId
),
cons AS (
  SELECT
    c.CveFteMT,
    c.PredioId,
    CAST(COALESCE(SUM(c.PredioConstruccionArea), 0) AS decimal(18,2)) AS ConstruccionAreaSum
  FROM ALPREDIOCONSTRUCCIONES c
  INNER JOIN toprows p
    ON p.CveFteMT = c.CveFteMT
   AND p.PredioId = c.PredioId
  GROUP BY c.CveFteMT, c.PredioId
),
val AS (
  SELECT
    v.CveFteMT,
    v.PredioId,
    v.PredioValuoCatastralImporte,
    ROW_NUMBER() OVER (
      PARTITION BY v.CveFteMT, v.PredioId
      ORDER BY v.PredioValuoCatastralFecha DESC, v.PredioValuoCatastralEjercicio DESC
    ) AS rn
  FROM ALPREDIOVALUOCATASTRAL v
  INNER JOIN toprows p
    ON p.CveFteMT = v.CveFteMT
   AND p.PredioId = v.PredioId
),
recibo AS (
  SELECT
    r.CveFteMT,
    r.PredioId,
    r.PredioEdoCuentaEjercicio AS ReciboEjercicio,
    r.PredioEdoCuentaPeriodo AS ReciboPeriodo,
    ROW_NUMBER() OVER (
      PARTITION BY r.CveFteMT, r.PredioId
      ORDER BY r.PredioEdoCuentaEjercicio DESC, r.PredioEdoCuentaPeriodo DESC, r.PredioEdoCuentaReciboId DESC
    ) AS rn
  FROM ALPREDIOEDOCUENTARECIBO r
  INNER JOIN toprows p
    ON p.CveFteMT = r.CveFteMT
   AND p.PredioId = r.PredioId
),
aval AS (
  SELECT
    a.CveFteMT,
    a.PredioId,
    a.SolicitudAvaluoEjercicioAvaluoActual,
    ROW_NUMBER() OVER (
      PARTITION BY a.CveFteMT, a.PredioId
      ORDER BY a.SolicitudAvaluoEjercicioAvaluoActual DESC, a.SolicitudAvaluoId DESC
    ) AS rn
  FROM AL23SolicitudAvaluo a
  INNER JOIN toprows p
    ON p.CveFteMT = a.CveFteMT
   AND p.PredioId = a.PredioId
),
mop AS (
  SELECT
    m.CveFteMT,
    m.PredioId,
    CAST(m.PredioEdoCuentaTasa AS decimal(18,6)) AS PredioEdoCuentaTasa,
    ROW_NUMBER() OVER (
      PARTITION BY m.CveFteMT, m.PredioId
      ORDER BY m.PredioEdoCuentaActualizacionFecha DESC, m.PredioEdoCuentaVencimientoFecha DESC, m.PredioEjercicioAlta DESC
    ) AS rn
  FROM ALMOVEDOPRE m
  INNER JOIN toprows p
    ON p.CveFteMT = m.CveFteMT
   AND p.PredioId = m.PredioId
  WHERE m.PredioEdoCuentaEjercicio = @EjercicioActual
    AND m.PredioEdoCuentaPeriodo = 1
)
SELECT
  CAST(p.PredioId AS int) AS [Padrón],
  p.PredioCveCatastral AS [Clave Catastral],
  p.PredioAltaFecha AS [Fecha de Alta del Predio],
  p.PredioCalle AS [Calle],
  p.PredioCodigoPostal AS [Código Postal],
  CASE
    WHEN NULLIF(RTRIM(p.PredioNumExt), '') IS NOT NULL THEN
      CASE
        WHEN NULLIF(RTRIM(p.PredioNumInt), '') IS NOT NULL THEN CONCAT(RTRIM(p.PredioNumExt), '-', RTRIM(p.PredioNumInt))
        ELSE RTRIM(p.PredioNumExt)
      END
    ELSE
      CASE
        WHEN NULLIF(RTRIM(p.PredioNumInt), '') IS NOT NULL THEN CONCAT('INT ', RTRIM(p.PredioNumInt))
        ELSE 'S/N'
      END
  END AS [Número],
  p.PropietarioNombre AS [Propietario],
  p.PredioStatus AS [Estatus],
  p.CatastroDatosEscriturales AS [Datos escriturales],
  CAST(COALESCE(terr.TerrenoAreaSum, p.PredioTotalTerreno, p.PredioArea, 0) AS decimal(18,2)) AS [Superficie del Terreno],
  CAST(COALESCE(terr.TerrenoImporteSum, p.PredioTerrenoImporte, 0) AS decimal(18,2)) AS [Valor del Terreno],
  CAST(COALESCE(cons.ConstruccionAreaSum, p.PredioTotalConstruccion, 0) AS decimal(18,2)) AS [Área Construida],
  CAST(COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS [Valor de Construcción],
  CAST(COALESCE(CAST(v.PredioValuoCatastralImporte AS decimal(18,2)), p.PredioCatastralImporte, 0) AS decimal(18,2)) AS [Valor Catastral],
  CAST(av.SolicitudAvaluoEjercicioAvaluoActual AS int) AS [Año del valor catastral],
  CAST(
    COALESCE(CAST(v.PredioValuoCatastralImporte AS decimal(18,2)), p.PredioCatastralImporte, 0) *
    COALESCE(m.PredioEdoCuentaTasa, 0)
  AS decimal(18,2)) AS [Impuesto Actual],
  CAST(
    (
      COALESCE(CAST(v.PredioValuoCatastralImporte AS decimal(18,2)), p.PredioCatastralImporte, 0) *
      COALESCE(m.PredioEdoCuentaTasa, 0)
    ) / CAST(6 AS decimal(18,6))
  AS decimal(18,2)) AS [Impuesto por bimestre],
  CASE
    WHEN p.PredioUltimoEjericicioPagado IS NOT NULL AND p.PredioUltimoPeriodoPagado IS NOT NULL
      THEN CONCAT(CAST(p.PredioUltimoEjericicioPagado AS varchar(4)), ' - ', CAST(p.PredioUltimoPeriodoPagado AS varchar(2)))
    WHEN p.PredioUltimoEjericicioPagado IS NOT NULL THEN CAST(p.PredioUltimoEjericicioPagado AS varchar(4))
    WHEN p.PredioUltimoPeriodoPagado IS NOT NULL THEN CAST(p.PredioUltimoPeriodoPagado AS varchar(2))
    ELSE ''
  END AS [Ut.Bim.Pagado],
  p.PredioTipo AS [Tipo de Predio],
  p.EstadoFisicoNombre AS [Estado Físico],
  CASE
    WHEN r.ReciboEjercicio IS NOT NULL AND r.ReciboPeriodo IS NOT NULL THEN CONCAT(CAST(r.ReciboEjercicio AS varchar(4)), '-', CAST(r.ReciboPeriodo AS varchar(2)))
    WHEN r.ReciboEjercicio IS NOT NULL THEN CAST(r.ReciboEjercicio AS varchar(4))
    WHEN r.ReciboPeriodo IS NOT NULL THEN CAST(r.ReciboPeriodo AS varchar(2))
    ELSE ''
  END AS [Ejer - Per]
FROM toprows p
LEFT JOIN terr
  ON terr.CveFteMT = p.CveFteMT
 AND terr.PredioId = p.PredioId
LEFT JOIN cons
  ON cons.CveFteMT = p.CveFteMT
 AND cons.PredioId = p.PredioId
LEFT JOIN val v
  ON v.CveFteMT = p.CveFteMT
 AND v.PredioId = p.PredioId
 AND v.rn = 1
LEFT JOIN recibo r
  ON r.CveFteMT = p.CveFteMT
 AND r.PredioId = p.PredioId
 AND r.rn = 1
LEFT JOIN aval av
  ON av.CveFteMT = p.CveFteMT
 AND av.PredioId = p.PredioId
 AND av.rn = 1
LEFT JOIN mop m
  ON m.CveFteMT = p.CveFteMT
 AND m.PredioId = p.PredioId
 AND m.rn = 1
ORDER BY p.PredioId ASC
OPTION (RECOMPILE);
"""
 
 
@app.get("/api/reportes/prediales/sabana")
def sabana_prediales(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    q = (request.query_params.get("q") or "").strip()
    clave_catastral = (request.query_params.get("claveCatastral") or "").strip()
    clave_mode_raw = (request.query_params.get("claveMode") or request.query_params.get("claveCatastralMode") or "").strip().lower()
    if clave_mode_raw in {"exacto", "exacta", "exact"}:
      clave_mode = "exacto"
    else:
      clave_mode = "contiene"

    predio_id_raw = request.query_params.get("predioId") or request.query_params.get("padron")
    propietario = (request.query_params.get("propietario") or "").strip()
    apellido_paterno = (request.query_params.get("apellidoPaterno") or "").strip()
    apellido_materno = (request.query_params.get("apellidoMaterno") or "").strip()
    nombre = (request.query_params.get("nombre") or "").strip()
    calle = (request.query_params.get("calle") or "").strip()
    numero = (request.query_params.get("numero") or "").strip()
    estatus = (request.query_params.get("estatus") or "").strip()
    adeudo_raw = (request.query_params.get("adeudo") or "").strip().lower()
    adeudo = adeudo_raw if adeudo_raw in {"todos", "con", "sin"} else ""

    from_alta = _parse_date(request.query_params.get("fromAlta"))
    to_alta = _parse_date(request.query_params.get("toAlta"))
 
    limit_raw = request.query_params.get("limit")
    offset_raw = request.query_params.get("offset")
    limit = int(limit_raw) if limit_raw and str(limit_raw).isdigit() else 200
    offset = int(offset_raw) if offset_raw and str(offset_raw).isdigit() else 0
    limit = max(1, min(500, limit))
    offset = max(0, offset)

    def to_int(value: Any) -> Optional[int]:
      if value in (None, "", "null"):
        return None
      s = str(value).strip()
      if not s:
        return None
      if s.isdigit():
        try:
          return int(s)
        except Exception:
          return None
      try:
        return int(float(s))
      except Exception:
        return None

    predio_id = to_int(predio_id_raw)
    ejercicio_actual = datetime.now().year
 
    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _SABANA_PREDIALES_SELECT,
        (
          cve_fte_mt,
          q or None,
          clave_catastral or None,
          clave_mode,
          predio_id,
          propietario or None,
          apellido_paterno or None,
          apellido_materno or None,
          nombre or None,
          calle or None,
          numero or None,
          estatus or None,
          adeudo or None,
          from_alta,
          to_alta,
          ejercicio_actual,
          limit + 1,
          offset,
        ),
      )
      data = _rows(cur)
 
    has_more = len(data) > limit
    rows = data[:limit] if has_more else data
    next_offset = offset + limit if has_more else None
 
    return ORJSONResponse(
        {
          "ok": True,
          "filtros": {
            "cveFteMT": cve_fte_mt,
            "q": q,
            "claveCatastral": clave_catastral,
            "claveMode": clave_mode,
            "predioId": predio_id,
            "propietario": propietario,
            "apellidoPaterno": apellido_paterno,
            "apellidoMaterno": apellido_materno,
            "nombre": nombre,
            "calle": calle,
            "numero": numero,
            "estatus": estatus,
            "adeudo": adeudo or "todos",
            "ejercicioActual": ejercicio_actual,
            "fromAlta": request.query_params.get("fromAlta") or None,
            "toAlta": request.query_params.get("toAlta") or None,
            "limit": limit,
            "offset": offset,
          },
          "count": len(rows),
          "hasMore": has_more,
          "nextOffset": next_offset,
          "rows": rows,
        }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})
 
 
@app.get("/api/reportes/prediales/sabana.csv")
def sabana_prediales_csv(request: Request) -> Any:
  params = dict(request.query_params)
  p = _prediales_sabana_filters(params)
  if _boolish(request.query_params.get("async"), False) and int(p.get("maxRows") or 0) >= int(_REPORT_ASYNC_THRESHOLD_ROWS):
    user = _require_login(request)
    info = _create_report_job("prediales_sabana_csv", params, user)
    _start_report_job(str(info.get("jobId")))
    job_id = str(info.get("jobId"))
    return ORJSONResponse(
      {
        "ok": True,
        "async": True,
        "job": info,
        "urls": {"status": f"/api/reportes/jobs/{job_id}", "download": f"/api/reportes/jobs/{job_id}/download"},
      }
    )

  cve_fte_mt = str(p.get("cveFteMT") or "MTULUM")
  q = str(p.get("q") or "")
  clave_catastral = str(p.get("claveCatastral") or "")
  clave_mode = str(p.get("claveMode") or "contiene")
  predio_id = p.get("predioId")
  propietario = str(p.get("propietario") or "")
  apellido_paterno = str(p.get("apellidoPaterno") or "")
  apellido_materno = str(p.get("apellidoMaterno") or "")
  nombre = str(p.get("nombre") or "")
  calle = str(p.get("calle") or "")
  numero = str(p.get("numero") or "")
  estatus = str(p.get("estatus") or "")
  adeudo = str(p.get("adeudo") or "").strip().lower()
  if adeudo == "todos":
    adeudo = ""

  from_alta = _parse_date(p.get("fromAlta"))
  to_alta = _parse_date(p.get("toAlta"))
  max_rows = int(p.get("maxRows") or 50000)
  columns = p.get("columns") or []
  ejercicio_actual = datetime.now().year

  def gen() -> Iterable[bytes]:
    yield b"\xef\xbb\xbf" + (",".join(columns) + "\n").encode("utf-8")
    conn = get_conn()
    try:
      cur = conn.cursor()
      cur.execute(
        _SABANA_PREDIALES_CSV,
        (
          cve_fte_mt,
          q or None,
          clave_catastral or None,
          clave_mode,
          predio_id,
          propietario or None,
          apellido_paterno or None,
          apellido_materno or None,
          nombre or None,
          calle or None,
          numero or None,
          estatus or None,
          adeudo or None,
          from_alta,
          to_alta,
          ejercicio_actual,
          max_rows,
        ),
      )
      while cur.description is None:
        if not cur.nextset():
          return
      cols = [c[0] for c in cur.description]
      index = {name: i for i, name in enumerate(cols)}
      while True:
        batch = cur.fetchmany(500)
        if not batch:
          break
        for row in batch:
          line = ",".join(_csv_escape(row[index[c]]) if c in index else "" for c in columns) + "\n"
          yield line.encode("utf-8")
    finally:
      conn.close()
 
  headers = {"Content-Disposition": 'attachment; filename="prediales_sabana.csv"'}
  return StreamingResponse(gen(), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/reportes/prediales/sabana.xlsx")
def sabana_prediales_xlsx(request: Request) -> StreamingResponse:
  cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
  q = (request.query_params.get("q") or "").strip()
  clave_catastral = (request.query_params.get("claveCatastral") or "").strip()
  clave_mode_raw = (request.query_params.get("claveMode") or request.query_params.get("claveCatastralMode") or "").strip().lower()
  if clave_mode_raw in {"exacto", "exacta", "exact"}:
    clave_mode = "exacto"
  else:
    clave_mode = "contiene"

  predio_id_raw = request.query_params.get("predioId") or request.query_params.get("padron")
  propietario = (request.query_params.get("propietario") or "").strip()
  apellido_paterno = (request.query_params.get("apellidoPaterno") or "").strip()
  apellido_materno = (request.query_params.get("apellidoMaterno") or "").strip()
  nombre = (request.query_params.get("nombre") or "").strip()
  calle = (request.query_params.get("calle") or "").strip()
  numero = (request.query_params.get("numero") or "").strip()
  estatus = (request.query_params.get("estatus") or "").strip()
  adeudo_raw = (request.query_params.get("adeudo") or "").strip().lower()
  adeudo = adeudo_raw if adeudo_raw in {"todos", "con", "sin"} else ""

  from_alta = _parse_date(request.query_params.get("fromAlta"))
  to_alta = _parse_date(request.query_params.get("toAlta"))

  max_rows_raw = request.query_params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 200000
  max_rows = max(1, min(200000, max_rows))

  columns = [
    "Padrón",
    "Clave Catastral",
    "Fecha de Alta del Predio",
    "Calle",
    "Código Postal",
    "Número",
    "Propietario",
    "Estatus",
    "Datos escriturales",
    "Superficie del Terreno",
    "Valor del Terreno",
    "Área Construida",
    "Valor de Construcción",
    "Valor Catastral",
    "Año del valor catastral",
    "Impuesto Actual",
    "Impuesto por bimestre",
    "Ut.Bim.Pagado",
    "Tipo de Predio",
    "Estado Físico",
    "Ejer - Per",
  ]

  def to_int(value: Any) -> Optional[int]:
    if value in (None, "", "null"):
      return None
    s = str(value).strip()
    if not s:
      return None
    if s.isdigit():
      try:
        return int(s)
      except Exception:
        return None
    try:
      return int(float(s))
    except Exception:
      return None

  predio_id = to_int(predio_id_raw)
  ejercicio_actual = datetime.now().year

  date_cols = {"Fecha de Alta del Predio"}
  text_cols = {"Clave Catastral", "Número", "Datos escriturales", "Ut.Bim.Pagado", "Ejer - Per"}
  number_cols = {"Superficie del Terreno", "Área Construida"}
  int_cols = {"Año del valor catastral"}
  currency_cols = {
    "Valor del Terreno",
    "Valor de Construcción",
    "Valor Catastral",
    "Impuesto Actual",
    "Impuesto por bimestre",
  }

  wb = Workbook(write_only=True)
  ws = wb.create_sheet("Padrón catastral")
  if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

  title_font = Font(bold=True, size=14)
  header_font = Font(bold=True)
  header_fill = PatternFill("solid", fgColor="E7E7E7")
  align_left = Alignment(vertical="top", wrap_text=True)
  align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin = Side(style="thin", color="A0A0A0")
  header_border = Border(bottom=thin)

  currency_format = '"$"#,##0.00'
  number_format = "#,##0.00"
  int_format = "0"
  date_format = "dd/mm/yy"
  text_format = "@"

  title_cell = WriteOnlyCell(ws, value="Padrón Catastral")
  title_cell.font = title_font
  ws.append([title_cell])
  ws.append([WriteOnlyCell(ws, value="Generado"), WriteOnlyCell(ws, value=datetime.now())])
  ws.append([WriteOnlyCell(ws, value="CveFteMT"), WriteOnlyCell(ws, value=cve_fte_mt)])
  ws.append([WriteOnlyCell(ws, value="Ejercicio"), WriteOnlyCell(ws, value=ejercicio_actual)])
  ws.append([WriteOnlyCell(ws, value="PredioId"), WriteOnlyCell(ws, value=predio_id or "")])
  ws.append([WriteOnlyCell(ws, value="Clave catastral"), WriteOnlyCell(ws, value=clave_catastral or "")])
  ws.append([WriteOnlyCell(ws, value="Modo clave"), WriteOnlyCell(ws, value=clave_mode)])
  ws.append([WriteOnlyCell(ws, value="Propietario"), WriteOnlyCell(ws, value=propietario or "")])
  ws.append([WriteOnlyCell(ws, value="Apellido paterno"), WriteOnlyCell(ws, value=apellido_paterno or "")])
  ws.append([WriteOnlyCell(ws, value="Apellido materno"), WriteOnlyCell(ws, value=apellido_materno or "")])
  ws.append([WriteOnlyCell(ws, value="Nombre"), WriteOnlyCell(ws, value=nombre or "")])
  ws.append([WriteOnlyCell(ws, value="Calle"), WriteOnlyCell(ws, value=calle or "")])
  ws.append([WriteOnlyCell(ws, value="Número"), WriteOnlyCell(ws, value=numero or "")])
  ws.append([WriteOnlyCell(ws, value="Estatus"), WriteOnlyCell(ws, value=estatus or "")])
  ws.append([WriteOnlyCell(ws, value="Adeudo"), WriteOnlyCell(ws, value=adeudo or "todos")])
  ws.append([WriteOnlyCell(ws, value="Alta desde"), WriteOnlyCell(ws, value=from_alta or "")])
  ws.append([WriteOnlyCell(ws, value="Alta hasta"), WriteOnlyCell(ws, value=to_alta or "")])
  ws.append([WriteOnlyCell(ws, value="Buscar"), WriteOnlyCell(ws, value=q or "")])
  ws.append([])

  header_row = []
  for name in columns:
    cell = WriteOnlyCell(ws, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = header_border
    header_row.append(cell)
  ws.append(header_row)
  ws.freeze_panes = "A21"

  for idx in range(1, len(columns) + 1):
    col_letter = get_column_letter(idx)
    ws.column_dimensions[col_letter].width = 18
  ws.column_dimensions["B"].width = 22
  ws.column_dimensions["D"].width = 40
  ws.column_dimensions["G"].width = 28
  ws.column_dimensions["I"].width = 44

  tmp = SpooledTemporaryFile(max_size=32 * 1024 * 1024)

  def file_iter() -> Iterable[bytes]:
    try:
      with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
          _SABANA_PREDIALES_CSV,
          (
            cve_fte_mt,
            q or None,
            clave_catastral or None,
            clave_mode,
            predio_id,
            propietario or None,
            apellido_paterno or None,
            apellido_materno or None,
            nombre or None,
            calle or None,
            numero or None,
            estatus or None,
            adeudo or None,
            from_alta,
            to_alta,
            ejercicio_actual,
            max_rows,
          ),
        )
        while cur.description is None:
          if not cur.nextset():
            break
        cols = [c[0] for c in cur.description] if cur.description else []
        index = {name: i for i, name in enumerate(cols)}

        while True:
          batch = cur.fetchmany(500)
          if not batch:
            break
          for row in batch:
            out_row = []
            for name in columns:
              raw = row[index[name]] if name in index else None
              value = raw
              fmt = None

              if value is None:
                pass
              elif name in date_cols and isinstance(value, datetime):
                fmt = date_format
              elif name in text_cols and isinstance(value, str):
                if value.startswith("'"):
                  value = value[1:]
                fmt = text_format
              elif name in currency_cols:
                if isinstance(value, decimal.Decimal):
                  d = value
                elif isinstance(value, (int, float)):
                  d = decimal.Decimal(str(value))
                else:
                  d = None
                if d is not None:
                  value = float(d)
                  fmt = currency_format
              elif name in int_cols:
                if isinstance(value, decimal.Decimal):
                  value = int(value)
                fmt = int_format
              elif name in number_cols:
                if isinstance(value, decimal.Decimal):
                  value = float(value)
                fmt = number_format

              cell = WriteOnlyCell(ws, value=value)
              cell.alignment = align_left
              if fmt:
                cell.number_format = fmt
              out_row.append(cell)
            ws.append(out_row)

      wb.save(tmp)
      tmp.seek(0)
      while True:
        chunk = tmp.read(1024 * 1024)
        if not chunk:
          break
        yield chunk
    finally:
      tmp.close()

  headers = {"Content-Disposition": 'attachment; filename="prediales_sabana.xlsx"'}
  return StreamingResponse(
    file_iter(),
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers=headers,
  )
 
 
def _prediales_pagos_filters(params: Dict[str, Any]) -> Dict[str, Any]:
  ejercicio = int(params["ejercicio"]) if params.get("ejercicio") not in (None, "", "null") else None
  pago_from = _parse_date(params.get("pagoFrom"))
  pago_to = _parse_date(params.get("pagoTo"))
  if ejercicio:
    pago_from = datetime(ejercicio, 1, 1)
    pago_to = datetime(ejercicio, 12, 31)
 
  todos = _normalize_bool(params.get("todos"), False)
  clave_catastral = (params.get("claveCatastral") or "").strip()
  clave_from = (params.get("claveCatastralFrom") or "").strip()
  clave_to = (params.get("claveCatastralTo") or "").strip()
  predio_id_raw = params.get("predioId")
  predio_id = int(predio_id_raw) if predio_id_raw not in (None, "", "null") else None
 
  if todos:
    clave_catastral = ""
    clave_from = ""
    clave_to = ""
    predio_id = None
 
  return {
    "todos": todos,
    "claveCatastral": clave_catastral,
    "claveCatastralFrom": clave_from,
    "claveCatastralTo": clave_to,
    "predioId": predio_id,
    "ejercicio": ejercicio,
    "pagoFrom": pago_from,
    "pagoTo": pago_to,
  }
 
 
_SABANA_PAGOS_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @ClaveCatastral varchar(64) = ?;
DECLARE @ClaveCatastralFrom varchar(64) = ?;
DECLARE @ClaveCatastralTo varchar(64) = ?;
DECLARE @PredioId decimal(18,0) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @Limit int = ?;
DECLARE @Offset int = ?;
 
SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    sp.CveFteMT,
    sp.PredioId,
    sp.SPagoPredialId,
    sp.SPagoPredialSerie,
    sp.SPagoPredialFolio,
    sp.SPagoPredialPagoFecha,
    sp.SPagoPredialInicialEjercicio,
    sp.SPagoPredialInicialPeriodo,
    sp.SPagoPredialFinalEjercicio,
    sp.SPagoPredialFinalPeriodo,
    COALESCE(
      NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
      RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
    ) AS ClaveCatNorm,
    RTRIM(COALESCE(per.NombreCompletoPersona COLLATE DATABASE_DEFAULT, '')) AS Propietario,
    RTRIM(COALESCE(per.RazonSocialPersona COLLATE DATABASE_DEFAULT, p.PredioContribuyenteNombre COLLATE DATABASE_DEFAULT, '')) AS RazonSocial,
    RTRIM(COALESCE(p.PredioQroDireccion COLLATE DATABASE_DEFAULT, '')) AS Direccion,
    RTRIM(COALESCE(p.PredioNombreColonia COLLATE DATABASE_DEFAULT, '')) AS Colonia,
    RTRIM(p.PredioTipo) COLLATE DATABASE_DEFAULT AS TipoPredio,
    CAST(COALESCE(p.PredioTerrenoImporte, 0) AS decimal(18,2)) AS ValorTerreno,
    CAST(COALESCE(p.PredioTotalConstruccion, 0) AS decimal(18,2)) AS AreaConstruida,
    CAST(COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS ValorConstruccion,
    CAST(COALESCE(p.PredioTerrenoImporte, 0) + COALESCE(p.PredioTotalConstruccion, 0) + COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS ValorCatastral,
    RTRIM(COALESCE(cal.CalificativoPropietarioNombre COLLATE DATABASE_DEFAULT, '')) AS Calificativo,
    RTRIM(COALESCE(ef.EstadoFisicoNombre COLLATE DATABASE_DEFAULT, '')) AS EstadoFisico,
    p.EstadoFisicoId,
    UPPER(RTRIM(CONVERT(varchar(50), p.PredioTipo))) AS PredioTipoNorm
  FROM ALSPAGOPREDIAL sp
  LEFT JOIN AlPredio p
    ON p.CveFteMT = sp.CveFteMT
   AND p.PredioId = sp.PredioId
  LEFT JOIN XiPersonas per
    ON per.CveFteMT = p.CveFteMT
   AND per.CvePersona = p.CvePersona
  LEFT JOIN ALCALIFICATIVOPROPIETARIO cal
    ON cal.CveFteMT = p.CveFteMT
   AND cal.CalificativoPropietarioId = p.CalificativoPropietarioId
  LEFT JOIN AlEstadoFisico ef
    ON ef.CveFteMT = p.CveFteMT
   AND ef.EstadoFisicoId = p.EstadoFisicoId
  WHERE sp.CveFteMT = @CveFteMT
    AND sp.SPagoPredialEstatus = 'PAG'
    AND (@PredioId IS NULL OR sp.PredioId = @PredioId)
    AND (@PagoFrom IS NULL OR sp.SPagoPredialPagoFecha >= @PagoFrom)
    AND (@PagoTo IS NULL OR sp.SPagoPredialPagoFecha < DATEADD(DAY, 1, @PagoTo))
    AND (
      @ClaveCatastral IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) = (@ClaveCatastral COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralFrom IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) >= (@ClaveCatastralFrom COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralTo IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) <= (@ClaveCatastralTo COLLATE DATABASE_DEFAULT)
    )
),
page AS (
  SELECT *
  FROM base
  ORDER BY SPagoPredialPagoFecha DESC, SPagoPredialId DESC
  OFFSET @Offset ROWS FETCH NEXT @Limit ROWS ONLY
),
det AS (
  SELECT
    rd.CveFteMT,
    rd.CveSerFol,
    rd.CveFolio,
    SUM(COALESCE(rd.ImpIniRec, 0)) AS SumAll,
    SUM(CASE WHEN rd.CveFteIng = 1201010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010101,
    SUM(CASE WHEN rd.CveFteIng = 1201010102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010102,
    SUM(CASE WHEN rd.CveFteIng = 1201010103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010103,
    SUM(CASE WHEN rd.CveFteIng = 1201010105 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010105,
    SUM(CASE WHEN rd.CveFteIng = 1201020101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020101,
    SUM(CASE WHEN rd.CveFteIng = 1201020102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020102,
    SUM(CASE WHEN rd.CveFteIng = 1201020104 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020104,
    SUM(CASE WHEN rd.CveFteIng = 1201030101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030101,
    SUM(CASE WHEN rd.CveFteIng = 1201030102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030102,
    SUM(CASE WHEN rd.CveFteIng = 1201030103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030103,
    SUM(CASE WHEN rd.CveFteIng = 1701010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1701010101,
    SUM(CASE WHEN rd.CveFteIng IN (1201040110, 1201040108, 1201040104, 1201040101, 1201040106, 1201040107, 1201040105, 1701010102) THEN rd.ImpIniRec ELSE 0 END) AS SumDescuentos
  FROM COQRECIBODETALLE rd
  INNER JOIN page sp
    ON sp.CveFteMT = rd.CveFteMT
   AND sp.SPagoPredialSerie = rd.CveSerFol
   AND sp.SPagoPredialFolio = rd.CveFolio
  GROUP BY rd.CveFteMT, rd.CveSerFol, rd.CveFolio
)
SELECT
  CAST(sp.PredioId AS int) AS [Clave],
  '''' + COALESCE(sp.ClaveCatNorm, '') AS [Clave Catastral],
  sp.Propietario AS [Propietario],
  sp.RazonSocial AS [Razon social del contribuyente],
  sp.Direccion AS [Direccion],
  sp.Colonia AS [Colonia],
  sp.TipoPredio AS [Tipo de Predio],
  sp.ValorTerreno AS [Valor del terreno],
  sp.AreaConstruida AS [Área construida],
  sp.ValorConstruccion AS [Valor de construcción],
  sp.ValorCatastral AS [Valor catastral],
  sp.Calificativo AS [Calificativo],
  sp.EstadoFisico AS [Estado fisico],
  CONCAT(sp.SPagoPredialInicialEjercicio, '- ', sp.SPagoPredialInicialPeriodo) AS [Periodo inicial],
  CONCAT(sp.SPagoPredialFinalEjercicio, '- ', sp.SPagoPredialFinalPeriodo) AS [Periodo final],
  RTRIM(sp.SPagoPredialSerie COLLATE DATABASE_DEFAULT) + ' ' + CAST(sp.SPagoPredialFolio AS varchar(20)) AS [Recibo],
  sp.SPagoPredialPagoFecha AS [Fecha de Pago],
  CAST(
    CASE
      WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201010103, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201010102, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201010101, 0)
      ELSE 0
    END
  AS decimal(18,2)) AS [Impuesto Corriente y Anticipado],
  CAST(
    CASE
      WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201030101, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201030102, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201030103, 0)
      ELSE 0
    END
  AS decimal(18,2)) AS [Rezago años anteriores],
  CAST(
    CASE
      WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201020101, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201020102, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201020104, 0)
      ELSE 0
    END
  AS decimal(18,2)) AS [Rezago],
  CAST(
    COALESCE(det.SumAll, 0) -
    (
      (
        CASE
          WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201010103, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201010102, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201010101, 0)
          ELSE 0
        END
      ) +
      (
        CASE
          WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201030101, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201030102, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201030103, 0)
          ELSE 0
        END
      ) +
      (
        CASE
          WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201020101, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201020102, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201020104, 0)
          ELSE 0
        END
      ) +
      COALESCE(det.Sum1201010105, 0) +
      COALESCE(det.Sum1701010101, 0) -
      ABS(COALESCE(det.SumDescuentos, 0))
    )
  AS decimal(18,2)) AS [Adicional],
  CAST(COALESCE(det.Sum1201010105, 0) AS decimal(18,2)) AS [Actualizacion],
  CAST(COALESCE(det.Sum1701010101, 0) AS decimal(18,2)) AS [Recargos],
  CAST(0 AS decimal(18,2)) AS [Requerimiento],
  CAST(0 AS decimal(18,2)) AS [Embargo],
  CAST(0 AS decimal(18,2)) AS [Multa],
  CAST(ABS(COALESCE(det.SumDescuentos, 0)) AS decimal(18,2)) AS [Descuentos],
  CAST(COALESCE(det.SumAll, 0) AS decimal(18,2)) AS [Total]
FROM page sp
LEFT JOIN det
  ON det.CveFteMT = sp.CveFteMT
 AND det.CveSerFol = sp.SPagoPredialSerie
 AND det.CveFolio = sp.SPagoPredialFolio
ORDER BY sp.SPagoPredialPagoFecha DESC, sp.SPagoPredialId DESC
OPTION (RECOMPILE);
"""

_SABANA_PAGOS_CSV = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @ClaveCatastral varchar(64) = ?;
DECLARE @ClaveCatastralFrom varchar(64) = ?;
DECLARE @ClaveCatastralTo varchar(64) = ?;
DECLARE @PredioId decimal(18,0) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @MaxRows int = ?;
 
SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    sp.CveFteMT,
    sp.PredioId,
    sp.SPagoPredialId,
    sp.SPagoPredialSerie,
    sp.SPagoPredialFolio,
    sp.SPagoPredialPagoFecha,
    sp.SPagoPredialInicialEjercicio,
    sp.SPagoPredialInicialPeriodo,
    sp.SPagoPredialFinalEjercicio,
    sp.SPagoPredialFinalPeriodo,
    COALESCE(
      NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
      RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
    ) AS ClaveCatNorm,
    RTRIM(COALESCE(per.NombreCompletoPersona COLLATE DATABASE_DEFAULT, '')) AS Propietario,
    RTRIM(COALESCE(per.RazonSocialPersona COLLATE DATABASE_DEFAULT, p.PredioContribuyenteNombre COLLATE DATABASE_DEFAULT, '')) AS RazonSocial,
    RTRIM(COALESCE(p.PredioQroDireccion COLLATE DATABASE_DEFAULT, '')) AS Direccion,
    RTRIM(COALESCE(p.PredioNombreColonia COLLATE DATABASE_DEFAULT, '')) AS Colonia,
    RTRIM(p.PredioTipo) COLLATE DATABASE_DEFAULT AS TipoPredio,
    CAST(COALESCE(p.PredioTerrenoImporte, 0) AS decimal(18,2)) AS ValorTerreno,
    CAST(COALESCE(p.PredioTotalConstruccion, 0) AS decimal(18,2)) AS AreaConstruida,
    CAST(COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS ValorConstruccion,
    CAST(COALESCE(p.PredioTerrenoImporte, 0) + COALESCE(p.PredioTotalConstruccion, 0) + COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS ValorCatastral,
    RTRIM(COALESCE(cal.CalificativoPropietarioNombre COLLATE DATABASE_DEFAULT, '')) AS Calificativo,
    RTRIM(COALESCE(ef.EstadoFisicoNombre COLLATE DATABASE_DEFAULT, '')) AS EstadoFisico,
    p.EstadoFisicoId,
    UPPER(RTRIM(CONVERT(varchar(50), p.PredioTipo))) AS PredioTipoNorm
  FROM ALSPAGOPREDIAL sp
  LEFT JOIN AlPredio p
    ON p.CveFteMT = sp.CveFteMT
   AND p.PredioId = sp.PredioId
  LEFT JOIN XiPersonas per
    ON per.CveFteMT = p.CveFteMT
   AND per.CvePersona = p.CvePersona
  LEFT JOIN ALCALIFICATIVOPROPIETARIO cal
    ON cal.CveFteMT = p.CveFteMT
   AND cal.CalificativoPropietarioId = p.CalificativoPropietarioId
  LEFT JOIN AlEstadoFisico ef
    ON ef.CveFteMT = p.CveFteMT
   AND ef.EstadoFisicoId = p.EstadoFisicoId
  WHERE sp.CveFteMT = @CveFteMT
    AND sp.SPagoPredialEstatus = 'PAG'
    AND (@PredioId IS NULL OR sp.PredioId = @PredioId)
    AND (@PagoFrom IS NULL OR sp.SPagoPredialPagoFecha >= @PagoFrom)
    AND (@PagoTo IS NULL OR sp.SPagoPredialPagoFecha < DATEADD(DAY, 1, @PagoTo))
    AND (
      @ClaveCatastral IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) = (@ClaveCatastral COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralFrom IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) >= (@ClaveCatastralFrom COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralTo IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) <= (@ClaveCatastralTo COLLATE DATABASE_DEFAULT)
    )
),
page AS (
  SELECT TOP (@MaxRows) *
  FROM base
  ORDER BY SPagoPredialPagoFecha DESC, SPagoPredialId DESC
),
det AS (
  SELECT
    rd.CveFteMT,
    rd.CveSerFol,
    rd.CveFolio,
    SUM(COALESCE(rd.ImpIniRec, 0)) AS SumAll,
    SUM(CASE WHEN rd.CveFteIng = 1201010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010101,
    SUM(CASE WHEN rd.CveFteIng = 1201010102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010102,
    SUM(CASE WHEN rd.CveFteIng = 1201010103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010103,
    SUM(CASE WHEN rd.CveFteIng = 1201010105 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010105,
    SUM(CASE WHEN rd.CveFteIng = 1201020101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020101,
    SUM(CASE WHEN rd.CveFteIng = 1201020102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020102,
    SUM(CASE WHEN rd.CveFteIng = 1201020104 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020104,
    SUM(CASE WHEN rd.CveFteIng = 1201030101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030101,
    SUM(CASE WHEN rd.CveFteIng = 1201030102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030102,
    SUM(CASE WHEN rd.CveFteIng = 1201030103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030103,
    SUM(CASE WHEN rd.CveFteIng = 1701010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1701010101,
    SUM(CASE WHEN rd.CveFteIng IN (1201040110, 1201040108, 1201040104, 1201040101, 1201040106, 1201040107, 1201040105, 1701010102) THEN rd.ImpIniRec ELSE 0 END) AS SumDescuentos
  FROM COQRECIBODETALLE rd
  INNER JOIN page sp
    ON sp.CveFteMT = rd.CveFteMT
   AND sp.SPagoPredialSerie = rd.CveSerFol
   AND sp.SPagoPredialFolio = rd.CveFolio
  GROUP BY rd.CveFteMT, rd.CveSerFol, rd.CveFolio
)
SELECT
  CAST(sp.PredioId AS int) AS [Clave],
  '''' + COALESCE(sp.ClaveCatNorm, '') AS [Clave Catastral],
  sp.Propietario AS [Propietario],
  sp.RazonSocial AS [Razon social del contribuyente],
  sp.Direccion AS [Direccion],
  sp.Colonia AS [Colonia],
  sp.TipoPredio AS [Tipo de Predio],
  sp.ValorTerreno AS [Valor del terreno],
  sp.AreaConstruida AS [Área construida],
  sp.ValorConstruccion AS [Valor de construcción],
  sp.ValorCatastral AS [Valor catastral],
  sp.Calificativo AS [Calificativo],
  sp.EstadoFisico AS [Estado fisico],
  CONCAT(sp.SPagoPredialInicialEjercicio, '- ', sp.SPagoPredialInicialPeriodo) AS [Periodo inicial],
  CONCAT(sp.SPagoPredialFinalEjercicio, '- ', sp.SPagoPredialFinalPeriodo) AS [Periodo final],
  RTRIM(sp.SPagoPredialSerie COLLATE DATABASE_DEFAULT) + ' ' + CAST(sp.SPagoPredialFolio AS varchar(20)) AS [Recibo],
  sp.SPagoPredialPagoFecha AS [Fecha de Pago],
  CAST(
    CASE
      WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201010103, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201010102, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201010101, 0)
      ELSE 0
    END
  AS decimal(18,2)) AS [Impuesto Corriente y Anticipado],
  CAST(
    CASE
      WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201030101, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201030102, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201030103, 0)
      ELSE 0
    END
  AS decimal(18,2)) AS [Rezago años anteriores],
  CAST(
    CASE
      WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201020101, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201020102, 0)
      WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201020104, 0)
      ELSE 0
    END
  AS decimal(18,2)) AS [Rezago],
  CAST(
    COALESCE(det.SumAll, 0) -
    (
      (
        CASE
          WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201010103, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201010102, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201010101, 0)
          ELSE 0
        END
      ) +
      (
        CASE
          WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201030101, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201030102, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201030103, 0)
          ELSE 0
        END
      ) +
      (
        CASE
          WHEN sp.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(det.Sum1201020101, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 1 THEN COALESCE(det.Sum1201020102, 0)
          WHEN sp.PredioTipoNorm IN ('URBANO', 'U') AND sp.EstadoFisicoId = 2 THEN COALESCE(det.Sum1201020104, 0)
          ELSE 0
        END
      ) +
      COALESCE(det.Sum1201010105, 0) +
      COALESCE(det.Sum1701010101, 0) -
      ABS(COALESCE(det.SumDescuentos, 0))
    )
  AS decimal(18,2)) AS [Adicional],
  CAST(COALESCE(det.Sum1201010105, 0) AS decimal(18,2)) AS [Actualizacion],
  CAST(COALESCE(det.Sum1701010101, 0) AS decimal(18,2)) AS [Recargos],
  CAST(0 AS decimal(18,2)) AS [Requerimiento],
  CAST(0 AS decimal(18,2)) AS [Embargo],
  CAST(0 AS decimal(18,2)) AS [Multa],
  CAST(ABS(COALESCE(det.SumDescuentos, 0)) AS decimal(18,2)) AS [Descuentos],
  CAST(COALESCE(det.SumAll, 0) AS decimal(18,2)) AS [Total]
FROM page sp
LEFT JOIN det
  ON det.CveFteMT = sp.CveFteMT
 AND det.CveSerFol = sp.SPagoPredialSerie
 AND det.CveFolio = sp.SPagoPredialFolio
ORDER BY sp.SPagoPredialPagoFecha DESC, sp.SPagoPredialId DESC
OPTION (RECOMPILE);
"""

_SABANA_PAGOS_TOTALS_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @ClaveCatastral varchar(64) = ?;
DECLARE @ClaveCatastralFrom varchar(64) = ?;
DECLARE @ClaveCatastralTo varchar(64) = ?;
DECLARE @PredioId decimal(18,0) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    sp.CveFteMT,
    sp.PredioId,
    sp.SPagoPredialSerie,
    sp.SPagoPredialFolio,
    p.EstadoFisicoId,
    UPPER(RTRIM(CONVERT(varchar(50), p.PredioTipo))) AS PredioTipoNorm
  FROM ALSPAGOPREDIAL sp
  LEFT JOIN AlPredio p
    ON p.CveFteMT = sp.CveFteMT
   AND p.PredioId = sp.PredioId
  WHERE sp.CveFteMT = @CveFteMT
    AND sp.SPagoPredialEstatus = 'PAG'
    AND (@PredioId IS NULL OR sp.PredioId = @PredioId)
    AND (@PagoFrom IS NULL OR sp.SPagoPredialPagoFecha >= @PagoFrom)
    AND (@PagoTo IS NULL OR sp.SPagoPredialPagoFecha < DATEADD(DAY, 1, @PagoTo))
    AND (
      @ClaveCatastral IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) = (@ClaveCatastral COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralFrom IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) >= (@ClaveCatastralFrom COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralTo IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) <= (@ClaveCatastralTo COLLATE DATABASE_DEFAULT)
    )
),
det AS (
  SELECT
    rd.CveFteMT,
    rd.CveSerFol,
    rd.CveFolio,
    SUM(COALESCE(rd.ImpIniRec, 0)) AS SumAll,
    SUM(CASE WHEN rd.CveFteIng = 1201010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010101,
    SUM(CASE WHEN rd.CveFteIng = 1201010102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010102,
    SUM(CASE WHEN rd.CveFteIng = 1201010103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010103,
    SUM(CASE WHEN rd.CveFteIng = 1201010105 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010105,
    SUM(CASE WHEN rd.CveFteIng = 1201020101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020101,
    SUM(CASE WHEN rd.CveFteIng = 1201020102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020102,
    SUM(CASE WHEN rd.CveFteIng = 1201020104 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020104,
    SUM(CASE WHEN rd.CveFteIng = 1201030101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030101,
    SUM(CASE WHEN rd.CveFteIng = 1201030102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030102,
    SUM(CASE WHEN rd.CveFteIng = 1201030103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030103,
    SUM(CASE WHEN rd.CveFteIng = 1701010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1701010101,
    SUM(CASE WHEN rd.CveFteIng IN (1201040110, 1201040108, 1201040104, 1201040101, 1201040106, 1201040107, 1201040105, 1701010102) THEN rd.ImpIniRec ELSE 0 END) AS SumDescuentos
  FROM COQRECIBODETALLE rd
  INNER JOIN base sp
    ON sp.CveFteMT = rd.CveFteMT
   AND sp.SPagoPredialSerie = rd.CveSerFol
   AND sp.SPagoPredialFolio = rd.CveFolio
  GROUP BY rd.CveFteMT, rd.CveSerFol, rd.CveFolio
),
calc AS (
  SELECT
    b.CveFteMT,
    b.SPagoPredialSerie,
    b.SPagoPredialFolio,
    CAST(
      CASE
        WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201010103, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201010102, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201010101, 0)
        ELSE 0
      END
    AS decimal(18,2)) AS Impuesto,
    CAST(
      CASE
        WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201030101, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201030102, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201030103, 0)
        ELSE 0
      END
    AS decimal(18,2)) AS RezagoAnt,
    CAST(
      CASE
        WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201020101, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201020102, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201020104, 0)
        ELSE 0
      END
    AS decimal(18,2)) AS Rezago,
    CAST(COALESCE(d.Sum1201010105, 0) AS decimal(18,2)) AS Actualizacion,
    CAST(COALESCE(d.Sum1701010101, 0) AS decimal(18,2)) AS Recargos,
    CAST(ABS(COALESCE(d.SumDescuentos, 0)) AS decimal(18,2)) AS Descuentos,
    CAST(COALESCE(d.SumAll, 0) AS decimal(18,2)) AS TotalDetalle,
    CAST(
      COALESCE(d.SumAll, 0) -
      (
        (
          CASE
            WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201010103, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201010102, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201010101, 0)
            ELSE 0
          END
        ) +
        (
          CASE
            WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201030101, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201030102, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201030103, 0)
            ELSE 0
          END
        ) +
        (
          CASE
            WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201020101, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201020102, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201020104, 0)
            ELSE 0
          END
        ) +
        COALESCE(d.Sum1201010105, 0) +
        COALESCE(d.Sum1701010101, 0) -
        ABS(COALESCE(d.SumDescuentos, 0))
      )
    AS decimal(18,2)) AS Adicional
  FROM base b
  LEFT JOIN det d
    ON d.CveFteMT = b.CveFteMT
   AND d.CveSerFol = b.SPagoPredialSerie
   AND d.CveFolio = b.SPagoPredialFolio
)
SELECT
  CAST(COALESCE(SUM(calc.Impuesto), 0) AS decimal(18,2)) AS [Impuesto Corriente y Anticipado],
  CAST(COALESCE(SUM(calc.RezagoAnt), 0) AS decimal(18,2)) AS [Rezago años anteriores],
  CAST(COALESCE(SUM(calc.Rezago), 0) AS decimal(18,2)) AS [Rezago],
  CAST(COALESCE(SUM(calc.Adicional), 0) AS decimal(18,2)) AS [Adicional],
  CAST(COALESCE(SUM(calc.Actualizacion), 0) AS decimal(18,2)) AS [Actualizacion],
  CAST(COALESCE(SUM(calc.Recargos), 0) AS decimal(18,2)) AS [Recargos],
  CAST(0 AS decimal(18,2)) AS [Requerimiento],
  CAST(0 AS decimal(18,2)) AS [Embargo],
  CAST(0 AS decimal(18,2)) AS [Multa],
  CAST(COALESCE(SUM(calc.Descuentos), 0) AS decimal(18,2)) AS [Descuentos],
  CAST(COALESCE(SUM(calc.TotalDetalle), 0) AS decimal(18,2)) AS [Total]
FROM calc
OPTION (RECOMPILE);
"""


_PREDIALES_PAGOS_MONTHLY_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @ClaveCatastral varchar(64) = ?;
DECLARE @ClaveCatastralFrom varchar(64) = ?;
DECLARE @ClaveCatastralTo varchar(64) = ?;
DECLARE @PredioId decimal(18,0) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    sp.CveFteMT,
    sp.PredioId,
    sp.SPagoPredialId,
    sp.SPagoPredialSerie,
    sp.SPagoPredialFolio,
    sp.SPagoPredialPagoFecha,
    p.EstadoFisicoId,
    UPPER(RTRIM(CONVERT(varchar(50), p.PredioTipo))) AS PredioTipoNorm
  FROM ALSPAGOPREDIAL sp
  LEFT JOIN AlPredio p
    ON p.CveFteMT = sp.CveFteMT
   AND p.PredioId = sp.PredioId
  WHERE sp.CveFteMT = @CveFteMT
    AND sp.SPagoPredialEstatus = 'PAG'
    AND (@PredioId IS NULL OR sp.PredioId = @PredioId)
    AND (@PagoFrom IS NULL OR sp.SPagoPredialPagoFecha >= @PagoFrom)
    AND (@PagoTo IS NULL OR sp.SPagoPredialPagoFecha < DATEADD(DAY, 1, @PagoTo))
    AND (
      @ClaveCatastral IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) = (@ClaveCatastral COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralFrom IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) >= (@ClaveCatastralFrom COLLATE DATABASE_DEFAULT)
    )
    AND (
      @ClaveCatastralTo IS NULL OR
      COALESCE(
        NULLIF(RTRIM(sp.SPagoPredialCveCatastral COLLATE DATABASE_DEFAULT), '' COLLATE DATABASE_DEFAULT),
        RTRIM(p.PredioCveCatastral COLLATE DATABASE_DEFAULT)
      ) <= (@ClaveCatastralTo COLLATE DATABASE_DEFAULT)
    )
),
det AS (
  SELECT
    rd.CveFteMT,
    rd.CveSerFol,
    rd.CveFolio,
    SUM(COALESCE(rd.ImpIniRec, 0)) AS SumAll,
    SUM(CASE WHEN rd.CveFteIng = 1201010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010101,
    SUM(CASE WHEN rd.CveFteIng = 1201010102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010102,
    SUM(CASE WHEN rd.CveFteIng = 1201010103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010103,
    SUM(CASE WHEN rd.CveFteIng = 1201010105 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201010105,
    SUM(CASE WHEN rd.CveFteIng = 1201020101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020101,
    SUM(CASE WHEN rd.CveFteIng = 1201020102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020102,
    SUM(CASE WHEN rd.CveFteIng = 1201020104 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201020104,
    SUM(CASE WHEN rd.CveFteIng = 1201030101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030101,
    SUM(CASE WHEN rd.CveFteIng = 1201030102 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030102,
    SUM(CASE WHEN rd.CveFteIng = 1201030103 THEN rd.ImpIniRec ELSE 0 END) AS Sum1201030103,
    SUM(CASE WHEN rd.CveFteIng = 1701010101 THEN rd.ImpIniRec ELSE 0 END) AS Sum1701010101,
    SUM(CASE WHEN rd.CveFteIng IN (1201040110, 1201040108, 1201040104, 1201040101, 1201040106, 1201040107, 1201040105, 1701010102) THEN rd.ImpIniRec ELSE 0 END) AS SumDescuentos
  FROM COQRECIBODETALLE rd
  INNER JOIN base sp
    ON sp.CveFteMT = rd.CveFteMT
   AND sp.SPagoPredialSerie = rd.CveSerFol
   AND sp.SPagoPredialFolio = rd.CveFolio
  GROUP BY rd.CveFteMT, rd.CveSerFol, rd.CveFolio
),
calc AS (
  SELECT
    b.CveFteMT,
    b.SPagoPredialId,
    b.SPagoPredialSerie,
    b.SPagoPredialFolio,
    b.SPagoPredialPagoFecha AS PagoFecha,
    CAST(
      CASE
        WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201010103, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201010102, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201010101, 0)
        ELSE 0
      END
    AS decimal(18,2)) AS Impuesto,
    CAST(
      CASE
        WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201030101, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201030102, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201030103, 0)
        ELSE 0
      END
    AS decimal(18,2)) AS RezagoAnt,
    CAST(
      CASE
        WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201020101, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201020102, 0)
        WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201020104, 0)
        ELSE 0
      END
    AS decimal(18,2)) AS Rezago,
    CAST(COALESCE(d.Sum1201010105, 0) AS decimal(18,2)) AS Actualizacion,
    CAST(COALESCE(d.Sum1701010101, 0) AS decimal(18,2)) AS Recargos,
    CAST(ABS(COALESCE(d.SumDescuentos, 0)) AS decimal(18,2)) AS Descuentos,
    CAST(COALESCE(d.SumAll, 0) AS decimal(18,2)) AS TotalDetalle,
    CAST(
      COALESCE(d.SumAll, 0) -
      (
        (
          CASE
            WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201010103, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201010102, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201010101, 0)
            ELSE 0
          END
        ) +
        (
          CASE
            WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201030101, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201030102, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201030103, 0)
            ELSE 0
          END
        ) +
        (
          CASE
            WHEN b.PredioTipoNorm IN ('RURAL', 'R', 'RUSTICO', 'RÚSTICO') THEN COALESCE(d.Sum1201020101, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 1 THEN COALESCE(d.Sum1201020102, 0)
            WHEN b.PredioTipoNorm IN ('URBANO', 'U') AND b.EstadoFisicoId = 2 THEN COALESCE(d.Sum1201020104, 0)
            ELSE 0
          END
        ) +
        COALESCE(d.Sum1201010105, 0) +
        COALESCE(d.Sum1701010101, 0) -
        ABS(COALESCE(d.SumDescuentos, 0))
      )
    AS decimal(18,2)) AS Adicional
  FROM base b
  LEFT JOIN det d
    ON d.CveFteMT = b.CveFteMT
   AND d.CveSerFol = b.SPagoPredialSerie
   AND d.CveFolio = b.SPagoPredialFolio
)
SELECT
  CONCAT(CAST(YEAR(calc.PagoFecha) AS varchar(4)), '-', RIGHT(CONCAT('0', CAST(MONTH(calc.PagoFecha) AS varchar(2))), 2)) AS Periodo,
  CAST(YEAR(calc.PagoFecha) AS int) AS Ejercicio,
  CAST(MONTH(calc.PagoFecha) AS int) AS Mes,
  CAST(COALESCE(COUNT(DISTINCT calc.SPagoPredialId), 0) AS int) AS Recibos,
  CAST(COALESCE(SUM(calc.TotalDetalle), 0) AS decimal(18,2)) AS Total
FROM calc
GROUP BY YEAR(calc.PagoFecha), MONTH(calc.PagoFecha)
ORDER BY YEAR(calc.PagoFecha) DESC, MONTH(calc.PagoFecha) DESC
OPTION (RECOMPILE);
"""
 
 
@app.get("/api/reportes/prediales/sabana-pagos")
def sabana_pagos(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _prediales_pagos_filters(dict(request.query_params))
 
    limit_raw = request.query_params.get("limit")
    offset_raw = request.query_params.get("offset")
    limit = int(limit_raw) if limit_raw and str(limit_raw).isdigit() else 200
    offset = int(offset_raw) if offset_raw and str(offset_raw).isdigit() else 0
    limit = max(1, min(1000, limit))
    offset = max(0, offset)
 
    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _SABANA_PAGOS_SELECT,
        (
          cve_fte_mt,
          f["claveCatastral"] or None,
          f["claveCatastralFrom"] or None,
          f["claveCatastralTo"] or None,
          f["predioId"],
          f["pagoFrom"],
          f["pagoTo"],
          limit + 1,
          offset,
        ),
      )
      data = _rows(cur)
      cur.execute(
        _SABANA_PAGOS_TOTALS_SELECT,
        (
          cve_fte_mt,
          f["claveCatastral"] or None,
          f["claveCatastralFrom"] or None,
          f["claveCatastralTo"] or None,
          f["predioId"],
          f["pagoFrom"],
          f["pagoTo"],
        ),
      )
      totals_rows = _rows(cur)
      totals = totals_rows[0] if totals_rows else {}
 
    has_more = len(data) > limit
    rows = data[:limit] if has_more else data
    next_offset = offset + limit if has_more else None
 
    return ORJSONResponse(
        {
          "ok": True,
          "filtros": {
            "cveFteMT": cve_fte_mt,
            "todos": f["todos"],
            "claveCatastral": f["claveCatastral"],
            "claveCatastralFrom": f["claveCatastralFrom"],
            "claveCatastralTo": f["claveCatastralTo"],
            "predioId": f["predioId"],
            "ejercicio": f["ejercicio"],
            "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
            "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
            "limit": limit,
            "offset": offset,
          },
          "count": len(rows),
          "hasMore": has_more,
          "nextOffset": next_offset,
          "totals": totals,
          "rows": rows,
        }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})
 
 
@app.get("/api/reportes/prediales/sabana-pagos.csv")
def sabana_pagos_csv(request: Request) -> Any:
  if _boolish(request.query_params.get("async"), False):
    max_rows_raw = request.query_params.get("maxRows")
    max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
    if max_rows >= int(_REPORT_ASYNC_THRESHOLD_ROWS):
      user = _require_login(request)
      params = dict(request.query_params)
      info = _create_report_job("prediales_sabana_pagos_csv", params, user)
      _start_report_job(str(info.get("jobId")))
      job_id = str(info.get("jobId"))
      return ORJSONResponse(
        {
          "ok": True,
          "async": True,
          "job": info,
          "urls": {"status": f"/api/reportes/jobs/{job_id}", "download": f"/api/reportes/jobs/{job_id}/download"},
        }
      )
  cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
  f = _prediales_pagos_filters(dict(request.query_params))
 
  max_rows_raw = request.query_params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
  max_rows = max(1, min(200000, max_rows))
 
  columns = [
    "Clave",
    "Clave Catastral",
    "Propietario",
    "Razon social del contribuyente",
    "Direccion",
    "Colonia",
    "Tipo de Predio",
    "Valor del terreno",
    "Área construida",
    "Valor de construcción",
    "Valor catastral",
    "Calificativo",
    "Estado fisico",
    "Periodo inicial",
    "Periodo final",
    "Recibo",
    "Fecha de Pago",
    "Impuesto Corriente y Anticipado",
    "Rezago años anteriores",
    "Rezago",
    "Adicional",
    "Actualizacion",
    "Recargos",
    "Requerimiento",
    "Embargo",
    "Multa",
    "Descuentos",
    "Total",
  ]
 
  def gen() -> Iterable[bytes]:
    yield b"\xef\xbb\xbf" + (",".join(columns) + "\n").encode("utf-8")
    conn = get_conn()
    try:
      cur = conn.cursor()
      cur.execute(
        _SABANA_PAGOS_CSV,
        (
          cve_fte_mt,
          f["claveCatastral"] or None,
          f["claveCatastralFrom"] or None,
          f["claveCatastralTo"] or None,
          f["predioId"],
          f["pagoFrom"],
          f["pagoTo"],
          max_rows,
        ),
      )
      while cur.description is None:
        if not cur.nextset():
          return
      cols = [c[0] for c in cur.description]
      index = {name: i for i, name in enumerate(cols)}
      while True:
        batch = cur.fetchmany(500)
        if not batch:
          break
        for row in batch:
          line = ",".join(_csv_escape(row[index[c]]) if c in index else "" for c in columns) + "\n"
          yield line.encode("utf-8")
    finally:
      conn.close()
 
  headers = {"Content-Disposition": 'attachment; filename="prediales_sabana_pagos.csv"'}
  return StreamingResponse(gen(), media_type="text/csv; charset=utf-8", headers=headers)
 

@app.get("/api/reportes/prediales/sabana-pagos.xlsx")
def sabana_pagos_xlsx(request: Request) -> StreamingResponse:
  cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
  f = _prediales_pagos_filters(dict(request.query_params))

  max_rows_raw = request.query_params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
  max_rows = max(1, min(200000, max_rows))

  columns = [
    "Clave",
    "Clave Catastral",
    "Propietario",
    "Razon social del contribuyente",
    "Direccion",
    "Colonia",
    "Tipo de Predio",
    "Valor del terreno",
    "Área construida",
    "Valor de construcción",
    "Valor catastral",
    "Calificativo",
    "Estado fisico",
    "Periodo inicial",
    "Periodo final",
    "Recibo",
    "Fecha de Pago",
    "Impuesto Corriente y Anticipado",
    "Rezago años anteriores",
    "Rezago",
    "Adicional",
    "Actualizacion",
    "Recargos",
    "Requerimiento",
    "Embargo",
    "Multa",
    "Descuentos",
    "Total",
  ]

  date_cols = {"Fecha de Pago"}
  text_cols = {"Clave Catastral", "Periodo inicial", "Periodo final", "Recibo"}
  number_cols = {"Valor del terreno", "Área construida", "Valor de construcción"}
  currency_cols = {
    "Valor catastral",
    "Impuesto Corriente y Anticipado",
    "Rezago años anteriores",
    "Rezago",
    "Adicional",
    "Actualizacion",
    "Recargos",
    "Requerimiento",
    "Embargo",
    "Multa",
    "Descuentos",
    "Total",
  }

  totals_start_idx = columns.index("Impuesto Corriente y Anticipado")
  totals_cols = columns[totals_start_idx:]
  totals: Dict[str, decimal.Decimal] = {c: decimal.Decimal("0") for c in totals_cols}

  wb = Workbook(write_only=True)
  ws = wb.create_sheet("Sábana pagos")
  if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

  title_font = Font(bold=True, size=14)
  header_font = Font(bold=True)
  header_fill = PatternFill("solid", fgColor="E7E7E7")
  align_left = Alignment(vertical="top", wrap_text=True)
  align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin = Side(style="thin", color="A0A0A0")
  header_border = Border(bottom=thin)
  totals_border = Border(top=Side(style="thin", color="000000"))

  currency_format = '"$"#,##0.00'
  number_format = "#,##0.00"
  date_format = "dd/mm/yy"
  text_format = "@"

  current_row = 0

  title_cell = WriteOnlyCell(ws, value="Sábana de pagos prediales")
  title_cell.font = title_font
  ws.append([title_cell])
  current_row += 1

  ws.append([WriteOnlyCell(ws, value="Generado"), WriteOnlyCell(ws, value=datetime.now())])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="CveFteMT"), WriteOnlyCell(ws, value=cve_fte_mt)])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="Todos"), WriteOnlyCell(ws, value="Sí" if f["todos"] else "No")])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="Ejercicio"), WriteOnlyCell(ws, value=f["ejercicio"] or "")])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="PredioId"), WriteOnlyCell(ws, value=f["predioId"] or "")])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="Clave catastral"), WriteOnlyCell(ws, value=f["claveCatastral"] or "")])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="Clave catastral desde"), WriteOnlyCell(ws, value=f["claveCatastralFrom"] or "")])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="Clave catastral hasta"), WriteOnlyCell(ws, value=f["claveCatastralTo"] or "")])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="Pago desde"), WriteOnlyCell(ws, value=f["pagoFrom"] or "")])
  current_row += 1
  ws.append([WriteOnlyCell(ws, value="Pago hasta"), WriteOnlyCell(ws, value=f["pagoTo"] or "")])
  current_row += 1
  ws.append([])
  current_row += 1

  header_row = []
  for name in columns:
    cell = WriteOnlyCell(ws, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = header_border
    header_row.append(cell)
  ws.append(header_row)
  current_row += 1
  ws.freeze_panes = f"A{current_row + 1}"

  width_map = {
    "A": 10,
    "B": 20,
    "C": 28,
    "D": 30,
    "E": 40,
    "F": 18,
    "G": 16,
    "H": 16,
    "I": 14,
    "J": 18,
    "K": 16,
    "L": 18,
    "M": 14,
    "N": 16,
    "O": 16,
    "P": 14,
    "Q": 14,
  }
  for idx in range(1, len(columns) + 1):
    col_letter = get_column_letter(idx)
    ws.column_dimensions[col_letter].width = width_map.get(col_letter, 16)

  tmp = SpooledTemporaryFile(max_size=32 * 1024 * 1024)

  def file_iter() -> Iterable[bytes]:
    try:
      with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
          _SABANA_PAGOS_CSV,
          (
            cve_fte_mt,
            f["claveCatastral"] or None,
            f["claveCatastralFrom"] or None,
            f["claveCatastralTo"] or None,
            f["predioId"],
            f["pagoFrom"],
            f["pagoTo"],
            max_rows,
          ),
        )
        while cur.description is None:
          if not cur.nextset():
            break
        cols = [c[0] for c in cur.description] if cur.description else []
        index = {name: i for i, name in enumerate(cols)}

        while True:
          batch = cur.fetchmany(500)
          if not batch:
            break
          for row in batch:
            out_row = []
            for name in columns:
              raw = row[index[name]] if name in index else None
              value = raw
              fmt = None

              if value is None:
                pass
              elif name in date_cols and isinstance(value, datetime):
                fmt = date_format
              elif name in text_cols and isinstance(value, str):
                if value.startswith("'"):
                  value = value[1:]
                fmt = text_format
              elif name in currency_cols:
                if isinstance(value, decimal.Decimal):
                  d = value
                elif isinstance(value, (int, float)):
                  d = decimal.Decimal(str(value))
                else:
                  d = None
                if d is not None:
                  value = float(d)
                  fmt = currency_format
                  if name in totals:
                    totals[name] += d
              elif name in number_cols:
                if isinstance(value, decimal.Decimal):
                  value = float(value)
                fmt = number_format

              cell = WriteOnlyCell(ws, value=value)
              cell.alignment = align_left
              if fmt:
                cell.number_format = fmt
              out_row.append(cell)
            ws.append(out_row)

      label_idx = max(0, totals_start_idx - 1)
      totals_row = []
      for i, name in enumerate(columns):
        if i == label_idx:
          c = WriteOnlyCell(ws, value="TOTAL")
          c.font = header_font
          c.border = totals_border
          totals_row.append(c)
          continue
        if i >= totals_start_idx:
          d = totals.get(name, decimal.Decimal("0"))
          c = WriteOnlyCell(ws, value=float(d))
          c.font = header_font
          c.number_format = currency_format
          c.border = totals_border
          totals_row.append(c)
          continue
        c = WriteOnlyCell(ws, value=None)
        c.border = totals_border
        totals_row.append(c)
      ws.append(totals_row)

      wb.save(tmp)
      tmp.seek(0)
      while True:
        chunk = tmp.read(1024 * 1024)
        if not chunk:
          break
        yield chunk
    finally:
      tmp.close()

  headers = {"Content-Disposition": 'attachment; filename="prediales_sabana_pagos.xlsx"'}
  return StreamingResponse(
    file_iter(),
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers=headers,
  )


@app.get("/api/reportes/licencias/funcionamiento")
def licencias_funcionamiento(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _licencias_func_filters(dict(request.query_params))

    limit_raw = request.query_params.get("limit")
    offset_raw = request.query_params.get("offset")
    limit = int(limit_raw) if limit_raw and str(limit_raw).isdigit() else 200
    offset = int(offset_raw) if offset_raw and str(offset_raw).isdigit() else 0
    limit = max(1, min(1000, limit))
    offset = max(0, offset)

    def parse_int_env(name: str, default: Optional[int] = None) -> Optional[int]:
      raw = os.getenv(name)
      if raw in (None, "", "null"):
        return default
      try:
        return int(str(raw).strip())
      except Exception:
        return default

    licencia_principal = parse_int_env("LIC_FUNC_LICENCIA_FTEING")
    actualizaciones_fteing = parse_int_env("LIC_FUNC_ACTUALIZACION_FTEING", 4319020273)
    recargos_fteing = parse_int_env("LIC_FUNC_RECARGOS_FTEING", 4501010101)
    uma_ref = (
      datetime(int(f["ejercicio"]), 2, 1)
      if f.get("ejercicio")
      else (f.get("pagoFrom") or f.get("pagoTo") or datetime.now())
    )
    uma_mxn = get_uma_mxn_for_date(uma_ref) or decimal.Decimal("0")

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _LIC_FUNC_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["tipo"],
          f["licenciaId"],
          f["licenciaFrom"],
          f["licenciaTo"],
          licencia_principal,
          actualizaciones_fteing,
          recargos_fteing,
          f["ejercicio"],
          uma_mxn,
          limit + 1,
          offset,
        ),
      )
      data = _rows(cur)
      cur.execute(
        _LIC_FUNC_TOTALS_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["tipo"],
          f["licenciaId"],
          f["licenciaFrom"],
          f["licenciaTo"],
          licencia_principal,
          actualizaciones_fteing,
          recargos_fteing,
          f["ejercicio"],
          uma_mxn,
        ),
      )
      totals_rows = _rows(cur)
      totals = totals_rows[0] if totals_rows else {}

    has_more = len(data) > limit
    rows = data[:limit] if has_more else data
    next_offset = offset + limit if has_more else None

    return ORJSONResponse(
        {
          "ok": True,
          "filtros": {
            "cveFteMT": cve_fte_mt,
            "tipo": f["tipo"],
            "ejercicio": f["ejercicio"],
            "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
            "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
            "licenciaId": f["licenciaId"],
            "licenciaFrom": f["licenciaFrom"],
            "licenciaTo": f["licenciaTo"],
            "limit": limit,
            "offset": offset,
          },
          "count": len(rows),
          "hasMore": has_more,
          "nextOffset": next_offset,
          "totals": totals,
          "rows": rows,
        }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/reportes/saneamiento/ambiental")
def saneamiento_ambiental(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _saneamiento_ambiental_filters(dict(request.query_params))

    limit_raw = request.query_params.get("limit")
    offset_raw = request.query_params.get("offset")
    limit = int(limit_raw) if limit_raw and str(limit_raw).isdigit() else 200
    offset = int(offset_raw) if offset_raw and str(offset_raw).isdigit() else 0
    limit = max(1, min(1000, limit))
    offset = max(0, offset)

    def parse_int_env(name: str, default: int) -> int:
      raw = os.getenv(name)
      if raw in (None, "", "null"):
        return default
      try:
        return int(str(raw).strip())
      except Exception:
        return default

    derecho_fteing = parse_int_env("SAN_AMB_DERECHO_FTEING", 4326010110)
    actualizaciones_fteing = parse_int_env("SAN_AMB_ACTUALIZACION_FTEING", 4326010111)
    recargos_fteing = parse_int_env("SAN_AMB_RECARGOS_FTEING", 4501012601)

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _SAN_AMB_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["licenciaId"],
          derecho_fteing,
          actualizaciones_fteing,
          recargos_fteing,
          limit + 1,
          offset,
        ),
      )
      data = _rows(cur)
      cur.execute(
        _SAN_AMB_TOTALS_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["licenciaId"],
          derecho_fteing,
          actualizaciones_fteing,
          recargos_fteing,
        ),
      )
      totals_rows = _rows(cur)
      totals = totals_rows[0] if totals_rows else {}

    has_more = len(data) > limit
    rows = data[:limit] if has_more else data
    next_offset = offset + limit if has_more else None

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "ejercicio": f["ejercicio"],
          "licenciaId": f["licenciaId"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
          "limit": limit,
          "offset": offset,
        },
        "count": len(rows),
        "hasMore": has_more,
        "nextOffset": next_offset,
        "totals": totals,
        "rows": rows,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/analitica/saneamiento/ambiental")
def analitica_saneamiento_ambiental(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _saneamiento_ambiental_filters(dict(request.query_params))

    def parse_int_env(name: str, default: int) -> int:
      raw = os.getenv(name)
      if raw in (None, "", "null"):
        return default
      try:
        return int(str(raw).strip())
      except Exception:
        return default

    derecho_fteing = parse_int_env("SAN_AMB_DERECHO_FTEING", 4326010110)
    actualizaciones_fteing = parse_int_env("SAN_AMB_ACTUALIZACION_FTEING", 4326010111)
    recargos_fteing = parse_int_env("SAN_AMB_RECARGOS_FTEING", 4501012601)

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _SAN_AMB_MONTHLY_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["licenciaId"],
          derecho_fteing,
          actualizaciones_fteing,
          recargos_fteing,
        ),
      )
      series = _rows(cur)

      cur.execute(
        _SAN_AMB_TOTALS_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["licenciaId"],
          derecho_fteing,
          actualizaciones_fteing,
          recargos_fteing,
        ),
      )
      totals_rows = _rows(cur)
      totals = totals_rows[0] if totals_rows else {}

      cur.execute(
        _SAN_AMB_CANCELADOS_COUNT_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["licenciaId"],
        ),
      )
      cancelados_rows = _rows(cur)
      cancelados_count = int(cancelados_rows[0].get("Cancelados") or 0) if cancelados_rows else 0

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "ejercicio": f["ejercicio"],
          "licenciaId": f["licenciaId"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
        },
        "totals": totals,
        "series": series,
        "canceladosCount": cancelados_count,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/analitica/prediales/pagos")
def analitica_prediales_pagos(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _prediales_pagos_filters(dict(request.query_params))

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _PREDIALES_PAGOS_MONTHLY_SELECT,
        (
          cve_fte_mt,
          f["claveCatastral"] or None,
          f["claveCatastralFrom"] or None,
          f["claveCatastralTo"] or None,
          f["predioId"],
          f["pagoFrom"],
          f["pagoTo"],
        ),
      )
      series = _rows(cur)

      cur.execute(
        _SABANA_PAGOS_TOTALS_SELECT,
        (
          cve_fte_mt,
          f["claveCatastral"] or None,
          f["claveCatastralFrom"] or None,
          f["claveCatastralTo"] or None,
          f["predioId"],
          f["pagoFrom"],
          f["pagoTo"],
        ),
      )
      totals_rows = _rows(cur)
      totals = totals_rows[0] if totals_rows else {}

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "todos": f["todos"],
          "claveCatastral": f["claveCatastral"],
          "claveCatastralFrom": f["claveCatastralFrom"],
          "claveCatastralTo": f["claveCatastralTo"],
          "predioId": f["predioId"],
          "ejercicio": f["ejercicio"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
        },
        "totals": totals,
        "series": series,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/analitica/prediales/pagos/pronostico")
def analitica_prediales_pagos_pronostico(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _prediales_pagos_filters(dict(request.query_params))

    backtest_raw = request.query_params.get("backtestMonths")
    backtest_months = int(backtest_raw) if backtest_raw and str(backtest_raw).isdigit() else 6
    horizon_months = 6
    backtest_months = max(1, min(24, backtest_months))

    now = datetime.now()
    month_end = _month_end_dt(int(now.year), int(now.month))
    last_complete_month = (int(now.year), int(now.month)) if now.date() >= month_end.date() else _add_months(int(now.year), int(now.month), -1)
    last_complete_month = (int(last_complete_month[0]), int(last_complete_month[1]))
    start_dt = f["pagoFrom"] or datetime(int(now.year), 1, 1)
    end_dt = f["pagoTo"] or now
    start_month = (int(start_dt.year), int(start_dt.month))
    end_month = (int(end_dt.year), int(end_dt.month))
    horizon_end_month = _add_months(end_month[0], end_month[1], horizon_months)
    horizon_end_month = (int(horizon_end_month[0]), int(horizon_end_month[1]))
    include_real_through = last_complete_month if _month_leq(last_complete_month, horizon_end_month) else horizon_end_month

    training_end = end_month if _month_leq(end_month, last_complete_month) else last_complete_month
    query_to = _month_end_dt(include_real_through[0], include_real_through[1])

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _PREDIALES_PAGOS_MONTHLY_SELECT,
        (
          cve_fte_mt,
          f["claveCatastral"] or None,
          f["claveCatastralFrom"] or None,
          f["claveCatastralTo"] or None,
          f["predioId"],
          start_dt,
          query_to,
        ),
      )
      series_desc = _rows(cur)

    series_asc = sorted(series_desc, key=lambda r: str(r.get("Periodo") or ""))
    total_by_period: Dict[str, float] = {}
    for r in series_asc:
      period = str(r.get("Periodo") or "")
      total_by_period[period] = float(r.get("Total") or 0.0)

    all_months = _month_range(start_month, horizon_end_month)
    training_months = _month_range(start_month, training_end)

    y_train = [float(total_by_period.get(_period_str(yy, mm), 0.0)) for yy, mm in training_months]
    n = len(y_train)
    stats = _compute_stats(y_train)

    forecast_rows: List[Dict[str, Any]] = []
    forecast_start = _add_months(end_month[0], end_month[1], 1)
    forecast_start = (int(forecast_start[0]), int(forecast_start[1]))
    forecast_months = _month_range(forecast_start, horizon_end_month)
    y_ext: List[float] = list(y_train)
    for yy, mm in forecast_months:
      period = _period_str(yy, mm)
      meta = _predict_prediales_estacional_yoy(y_ext)
      pred = float(meta.get("pred") or 0.0)
      forecast_rows.append({"Periodo": period, "Pronostico": float(pred)})

      has_real_for_month = _month_leq((yy, mm), include_real_through)
      if has_real_for_month and period in total_by_period:
        y_ext.append(float(total_by_period.get(period, 0.0)))
      else:
        y_ext.append(float(pred))

    forecast_by_period = {str(r["Periodo"]): float(r["Pronostico"] or 0.0) for r in forecast_rows}

    combined: List[Dict[str, Any]] = []
    for yy, mm in all_months:
      period = _period_str(yy, mm)
      real_val = float(total_by_period.get(period, 0.0)) if _month_leq((yy, mm), include_real_through) else None
      pron = float(forecast_by_period[period]) if period in forecast_by_period else None
      combined.append({"Periodo": period, "Real": real_val, "Pronostico": pron})

    window_rows: List[Dict[str, Any]] = []
    for yy, mm in forecast_months:
      period = _period_str(yy, mm)
      has_real = _month_leq((yy, mm), include_real_through)
      real_raw = float(total_by_period.get(period, 0.0)) if has_real else None
      real_val = float(real_raw) if real_raw is not None else 0.0
      pred_val = float(forecast_by_period.get(period, 0.0))
      err_abs: Optional[float] = None
      err_pct: Optional[float] = None
      if real_raw is not None:
        err_abs = float(abs(real_raw - pred_val))
        err_pct = float((err_abs / abs(real_raw) * 100.0) if real_raw != 0 else 0.0)
      window_rows.append(
        {
          "Periodo": period,
          "Real": real_val,
          "Pronostico": pred_val,
          "ErrorAbs": err_abs,
          "ErrorPct": err_pct,
        }
      )

    bt_abs: List[float] = []
    bt_pct: List[float] = []
    backtest_rows: List[Dict[str, Any]] = []
    for r in window_rows:
      if r.get("ErrorAbs") is None or r.get("ErrorPct") is None:
        continue
      bt_abs.append(float(r["ErrorAbs"]))
      bt_pct.append(float(r["ErrorPct"]))
      backtest_rows.append(r)
    mae = float(sum(bt_abs) / len(bt_abs)) if bt_abs else 0.0
    mape = float(sum(bt_pct) / len(bt_pct)) if bt_pct else 0.0

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "todos": f["todos"],
          "claveCatastral": f["claveCatastral"],
          "claveCatastralFrom": f["claveCatastralFrom"],
          "claveCatastralTo": f["claveCatastralTo"],
          "predioId": f["predioId"],
          "ejercicio": f["ejercicio"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
          "horizonMonths": horizon_months,
          "backtestMonths": backtest_months,
        },
        "modelo": {
          "tipo": "estacional_yoy",
          "n": n,
          "mean": float(stats.get("mean") or 0.0),
          "avgDelta": float(stats.get("avg_delta") or 0.0),
          "avgPct": float(stats.get("avg_pct") or 0.0),
          "maeBacktest": mae,
          "mapeBacktest": mape,
        },
        "historico": series_asc,
        "pronosticos": forecast_rows,
        "pronosticosVsReal": window_rows,
        "backtest": backtest_rows,
        "serieCombinada": combined,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


_LIC_FUNC_MONTHLY_SELECT = """
DECLARE @CveFteMT varchar(32) = ?;
DECLARE @PagoFrom datetime = ?;
DECLARE @PagoTo datetime = ?;
DECLARE @Tipo varchar(16) = ?;
DECLARE @LicenciaId decimal(18,0) = ?;
DECLARE @LicenciaFrom decimal(18,0) = ?;
DECLARE @LicenciaTo decimal(18,0) = ?;
DECLARE @LicenciaPrincipalFteIng decimal(18,0) = ?;
DECLARE @ActualizacionesFteIng decimal(18,0) = ?;
DECLARE @RecargosFteIng decimal(18,0) = ?;
DECLARE @Ejercicio int = ?;
DECLARE @UMAMxn decimal(18,6) = ?;

DECLARE @IncluirLicencia bit = CASE WHEN @Tipo <> 'basura' THEN 1 ELSE 0 END;
DECLARE @IncluirBasura bit = CASE WHEN @Tipo <> 'licencia' THEN 1 ELSE 0 END;

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
SET LOCK_TIMEOUT 5000;

WITH base AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    r.ReciboPredioId,
    r.CveFecAsi
  FROM COQRECIBOS r
  WHERE r.CveFteMT = @CveFteMT
    AND r.ReciboGrupoId = 68
    AND r.ReciboTramiteId = 12
    AND r.EdoRec = 'A'
    AND (@LicenciaId IS NULL OR r.ReciboPredioId = @LicenciaId)
    AND (@LicenciaFrom IS NULL OR r.ReciboPredioId >= @LicenciaFrom)
    AND (@LicenciaTo IS NULL OR r.ReciboPredioId <= @LicenciaTo)
    AND (@PagoFrom IS NULL OR r.CveFecAsi >= @PagoFrom)
    AND (@PagoTo IS NULL OR r.CveFecAsi < DATEADD(DAY, 1, @PagoTo))
),
cfg AS (
  SELECT
    r.CveFteMT,
    r.CveSerFol,
    r.CveFolio,
    CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngLicencia, @LicenciaPrincipalFteIng, -1) AS decimal(18,0)) AS FteIngLicencia,
    CAST(COALESCE(tb.CveFteIngRenovacionId, -1) AS decimal(18,0)) AS FteIngRenovacion,
    CAST(COALESCE(tb.TarifasLicenciasBasuraCveFteIngBasura, -1) AS decimal(18,0)) AS FteIngBasura
  FROM base r
  LEFT JOIN TELICENCIASFUNCIONAMIENTO lf
    ON lf.CveFteMT = r.CveFteMT
   AND lf.LicenciasFuncionamientoId = r.ReciboPredioId
  LEFT JOIN TETARIFASLICENCIASBASURA tb
    ON tb.CveFteMT = r.CveFteMT
   AND tb.TarifasLicenciasBasuraGiroId = lf.LicenciasFuncionamientoGiroComercialPrincipalId
   AND tb.TarifasLicenciasBasuraEjercicio = COALESCE(@Ejercicio, YEAR(r.CveFecAsi))
),
det AS (
  SELECT
    d.CveFteMT,
    d.CveSerFol,
    d.CveFolio,
    d.CveFteIng,
    CAST(COALESCE(SUM(d.ReciboDetImpAntesDev), 0) AS decimal(18,2)) AS Importe
  FROM COQRECIBODETALLE d
  INNER JOIN base r
    ON r.CveFteMT = d.CveFteMT
   AND r.CveSerFol = d.CveSerFol
   AND r.CveFolio = d.CveFolio
  GROUP BY d.CveFteMT, d.CveSerFol, d.CveFolio, d.CveFteIng
)
SELECT
  CONCAT(CAST(YEAR(r.CveFecAsi) AS varchar(4)), '-', RIGHT(CONCAT('0', CAST(MONTH(r.CveFecAsi) AS varchar(2))), 2)) AS Periodo,
  CAST(YEAR(r.CveFecAsi) AS int) AS Ejercicio,
  CAST(MONTH(r.CveFecAsi) AS int) AS Mes,
  CAST(COALESCE(COUNT(1), 0) AS int) AS Recibos,
  CAST(
    COALESCE(
      SUM(
        CASE
          WHEN @IncluirLicencia = 1 AND d.CveFteIng = cfg.FteIngLicencia AND d.CveFteIng <> cfg.FteIngRenovacion THEN d.Importe
          WHEN @IncluirLicencia = 1 AND d.CveFteIng = cfg.FteIngRenovacion THEN d.Importe
          WHEN @IncluirBasura = 1 AND d.CveFteIng = cfg.FteIngBasura THEN d.Importe
          WHEN @IncluirBasura = 1 AND d.CveFteIng = @ActualizacionesFteIng THEN d.Importe
          WHEN @IncluirBasura = 1 AND d.CveFteIng = @RecargosFteIng THEN d.Importe
          WHEN @IncluirBasura = 1 AND d.CveFteIng NOT IN (cfg.FteIngLicencia, cfg.FteIngRenovacion, cfg.FteIngBasura, @ActualizacionesFteIng, @RecargosFteIng) THEN d.Importe
          ELSE 0
        END
      ),
      0
    )
  AS decimal(18,2)) AS Total
FROM base r
LEFT JOIN cfg
  ON cfg.CveFteMT = r.CveFteMT
 AND cfg.CveSerFol = r.CveSerFol
 AND cfg.CveFolio = r.CveFolio
LEFT JOIN det d
  ON d.CveFteMT = r.CveFteMT
 AND d.CveSerFol = r.CveSerFol
 AND d.CveFolio = r.CveFolio
GROUP BY YEAR(r.CveFecAsi), MONTH(r.CveFecAsi)
ORDER BY YEAR(r.CveFecAsi) DESC, MONTH(r.CveFecAsi) DESC
OPTION (RECOMPILE);
"""


@app.get("/api/analitica/licencias/funcionamiento")
def analitica_licencias_funcionamiento(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _licencias_func_filters(dict(request.query_params))

    def parse_int_env(name: str, default: Optional[int] = None) -> Optional[int]:
      raw = os.getenv(name)
      if raw in (None, "", "null"):
        return default
      try:
        return int(str(raw).strip())
      except Exception:
        return default

    licencia_principal = parse_int_env("LIC_FUNC_LICENCIA_FTEING")
    actualizaciones_fteing = parse_int_env("LIC_FUNC_ACTUALIZACION_FTEING", 4319020273)
    recargos_fteing = parse_int_env("LIC_FUNC_RECARGOS_FTEING", 4501010101)
    uma_ref = (
      datetime(int(f["ejercicio"]), 2, 1)
      if f.get("ejercicio")
      else (f.get("pagoFrom") or f.get("pagoTo") or datetime.now())
    )
    uma_mxn = get_uma_mxn_for_date(uma_ref) or decimal.Decimal("0")

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _LIC_FUNC_MONTHLY_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["tipo"],
          f["licenciaId"],
          f["licenciaFrom"],
          f["licenciaTo"],
          licencia_principal,
          actualizaciones_fteing,
          recargos_fteing,
          f["ejercicio"],
          uma_mxn,
        ),
      )
      series = _rows(cur)

      cur.execute(
        _LIC_FUNC_TOTALS_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["tipo"],
          f["licenciaId"],
          f["licenciaFrom"],
          f["licenciaTo"],
          licencia_principal,
          actualizaciones_fteing,
          recargos_fteing,
          f["ejercicio"],
          uma_mxn,
        ),
      )
      totals_rows = _rows(cur)
      totals = totals_rows[0] if totals_rows else {}

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "tipo": f["tipo"],
          "ejercicio": f["ejercicio"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
          "licenciaId": f["licenciaId"],
          "licenciaFrom": f["licenciaFrom"],
          "licenciaTo": f["licenciaTo"],
        },
        "totals": totals,
        "series": series,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/analitica/licencias/funcionamiento/pronostico")
def analitica_licencias_funcionamiento_pronostico(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _licencias_func_filters(dict(request.query_params))

    backtest_raw = request.query_params.get("backtestMonths")
    backtest_months = int(backtest_raw) if backtest_raw and str(backtest_raw).isdigit() else 6
    horizon_months = 6
    backtest_months = max(1, min(24, backtest_months))

    def parse_int_env(name: str, default: Optional[int] = None) -> Optional[int]:
      raw = os.getenv(name)
      if raw in (None, "", "null"):
        return default
      try:
        return int(str(raw).strip())
      except Exception:
        return default

    licencia_principal = parse_int_env("LIC_FUNC_LICENCIA_FTEING")
    actualizaciones_fteing = parse_int_env("LIC_FUNC_ACTUALIZACION_FTEING", 4319020273)
    recargos_fteing = parse_int_env("LIC_FUNC_RECARGOS_FTEING", 4501010101)
    uma_ref = (
      datetime(int(f["ejercicio"]), 2, 1)
      if f.get("ejercicio")
      else (f.get("pagoFrom") or f.get("pagoTo") or datetime.now())
    )
    uma_mxn = get_uma_mxn_for_date(uma_ref) or decimal.Decimal("0")

    now = datetime.now()
    month_end = _month_end_dt(int(now.year), int(now.month))
    last_complete_month = (int(now.year), int(now.month)) if now.date() >= month_end.date() else _add_months(int(now.year), int(now.month), -1)
    last_complete_month = (int(last_complete_month[0]), int(last_complete_month[1]))
    start_dt = f["pagoFrom"] or datetime(int(now.year), 1, 1)
    end_dt = f["pagoTo"] or now
    start_month = (int(start_dt.year), int(start_dt.month))
    end_month = (int(end_dt.year), int(end_dt.month))
    horizon_end_month = _add_months(end_month[0], end_month[1], horizon_months)
    horizon_end_month = (int(horizon_end_month[0]), int(horizon_end_month[1]))
    include_real_through = last_complete_month if _month_leq(last_complete_month, horizon_end_month) else horizon_end_month

    training_end = end_month if _month_leq(end_month, last_complete_month) else last_complete_month
    query_to = _month_end_dt(include_real_through[0], include_real_through[1])

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _LIC_FUNC_MONTHLY_SELECT,
        (
          cve_fte_mt,
          start_dt,
          query_to,
          f["tipo"],
          f["licenciaId"],
          f["licenciaFrom"],
          f["licenciaTo"],
          licencia_principal,
          actualizaciones_fteing,
          recargos_fteing,
          f["ejercicio"],
          uma_mxn,
        ),
      )
      series_desc = _rows(cur)

    series_asc = sorted(series_desc, key=lambda r: str(r.get("Periodo") or ""))
    total_by_period: Dict[str, float] = {}
    for r in series_asc:
      period = str(r.get("Periodo") or "")
      total_by_period[period] = float(r.get("Total") or 0.0)

    all_months = _month_range(start_month, horizon_end_month)
    training_months = _month_range(start_month, training_end)

    y_train = [float(total_by_period.get(_period_str(yy, mm), 0.0)) for yy, mm in training_months]
    n = len(y_train)
    stats = _compute_stats(y_train)

    forecast_rows: List[Dict[str, Any]] = []
    forecast_start = _add_months(end_month[0], end_month[1], 1)
    forecast_start = (int(forecast_start[0]), int(forecast_start[1]))
    forecast_months = _month_range(forecast_start, horizon_end_month)
    y_ext: List[float] = list(y_train)
    for yy, mm in forecast_months:
      period = _period_str(yy, mm)
      meta = _predict_prediales_estacional_yoy(y_ext)
      pred = float(meta.get("pred") or 0.0)
      forecast_rows.append({"Periodo": period, "Pronostico": float(pred)})

      has_real_for_month = _month_leq((yy, mm), include_real_through)
      if has_real_for_month and period in total_by_period:
        y_ext.append(float(total_by_period.get(period, 0.0)))
      else:
        y_ext.append(float(pred))

    forecast_by_period = {str(r["Periodo"]): float(r["Pronostico"] or 0.0) for r in forecast_rows}

    combined: List[Dict[str, Any]] = []
    for yy, mm in all_months:
      period = _period_str(yy, mm)
      real_val = float(total_by_period.get(period, 0.0)) if _month_leq((yy, mm), include_real_through) else None
      pron = float(forecast_by_period[period]) if period in forecast_by_period else None
      combined.append({"Periodo": period, "Real": real_val, "Pronostico": pron})

    window_rows: List[Dict[str, Any]] = []
    for yy, mm in forecast_months:
      period = _period_str(yy, mm)
      has_real = _month_leq((yy, mm), include_real_through)
      real_raw = float(total_by_period.get(period, 0.0)) if has_real else None
      real_val = float(real_raw) if real_raw is not None else 0.0
      pred_val = float(forecast_by_period.get(period, 0.0))
      err_abs: Optional[float] = None
      err_pct: Optional[float] = None
      if real_raw is not None:
        err_abs = float(abs(real_raw - pred_val))
        err_pct = float((err_abs / abs(real_raw) * 100.0) if real_raw != 0 else 0.0)
      window_rows.append(
        {
          "Periodo": period,
          "Real": real_val,
          "Pronostico": pred_val,
          "ErrorAbs": err_abs,
          "ErrorPct": err_pct,
        }
      )

    backtest_rows: List[Dict[str, Any]] = []
    bt_abs: List[float] = []
    bt_pct: List[float] = []
    for r in window_rows:
      if r.get("ErrorAbs") is None or r.get("ErrorPct") is None:
        continue
      bt_abs.append(float(r["ErrorAbs"]))
      bt_pct.append(float(r["ErrorPct"]))
      backtest_rows.append(r)

    mae = float(sum(bt_abs) / len(bt_abs)) if bt_abs else 0.0
    mape = float(sum(bt_pct) / len(bt_pct)) if bt_pct else 0.0

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "tipo": f["tipo"],
          "ejercicio": f["ejercicio"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
          "licenciaId": f["licenciaId"],
          "licenciaFrom": f["licenciaFrom"],
          "licenciaTo": f["licenciaTo"],
          "horizonMonths": horizon_months,
          "backtestMonths": backtest_months,
        },
        "modelo": {
          "tipo": "estacional_yoy",
          "n": n,
          "mean": float(stats.get("mean") or 0.0),
          "avgDelta": float(stats.get("avg_delta") or 0.0),
          "avgPct": float(stats.get("avg_pct") or 0.0),
          "maeBacktest": mae,
          "mapeBacktest": mape,
        },
        "historico": series_asc,
        "pronosticos": forecast_rows,
        "pronosticosVsReal": window_rows,
        "backtest": backtest_rows,
        "serieCombinada": combined,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/analitica/saneamiento/ambiental/cancelados")
def analitica_saneamiento_ambiental_cancelados(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _saneamiento_ambiental_filters(dict(request.query_params))

    limit_raw = request.query_params.get("limit")
    offset_raw = request.query_params.get("offset")
    limit = int(limit_raw) if limit_raw and str(limit_raw).isdigit() else 200
    offset = int(offset_raw) if offset_raw and str(offset_raw).isdigit() else 0
    limit = max(1, min(500, limit))
    offset = max(0, offset)

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _SAN_AMB_CANCELADOS_LIST_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["licenciaId"],
          limit + 1,
          offset,
        ),
      )
      data = _rows(cur)
      has_more = len(data) > limit
      rows = data[:limit] if has_more else data
      next_offset = offset + limit if has_more else None

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "ejercicio": f["ejercicio"],
          "licenciaId": f["licenciaId"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
          "limit": limit,
          "offset": offset,
        },
        "count": len(rows),
        "hasMore": has_more,
        "nextOffset": next_offset,
        "rows": rows,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


def _add_months(year: int, month: int, add: int) -> Tuple[int, int]:
  y = int(year)
  m = int(month)
  total = (y * 12 + (m - 1)) + int(add)
  ny = total // 12
  nm = (total % 12) + 1
  return ny, nm


def _period_str(year: int, month: int) -> str:
  return f"{int(year):04d}-{int(month):02d}"


def _month_to_int(year: int, month: int) -> int:
  return int(year) * 12 + (int(month) - 1)


def _month_leq(a: Tuple[int, int], b: Tuple[int, int]) -> bool:
  return _month_to_int(a[0], a[1]) <= _month_to_int(b[0], b[1])


def _month_end_dt(year: int, month: int) -> datetime:
  yy, mm = int(year), int(month)
  ny, nm = _add_months(yy, mm, 1)
  first_next = datetime(ny, nm, 1)
  last_day = first_next - timedelta(days=1)
  return datetime(int(last_day.year), int(last_day.month), int(last_day.day))


def _month_range(start: Tuple[int, int], end: Tuple[int, int]) -> List[Tuple[int, int]]:
  if not _month_leq(start, end):
    return []
  out: List[Tuple[int, int]] = []
  sy, sm = int(start[0]), int(start[1])
  ey, em = int(end[0]), int(end[1])
  cur_y, cur_m = sy, sm
  while _month_leq((cur_y, cur_m), (ey, em)):
    out.append((cur_y, cur_m))
    cur_y, cur_m = _add_months(cur_y, cur_m, 1)
  return out


def _compute_stats(y: List[float]) -> Dict[str, float]:
  if not y:
    return {"mean": 0.0, "avg_delta": 0.0, "avg_pct": 0.0}
  mean_y = float(sum(y) / float(len(y)))
  if len(y) < 2:
    return {"mean": mean_y, "avg_delta": 0.0, "avg_pct": 0.0}

  deltas = [float(y[i] - y[i - 1]) for i in range(1, len(y))]
  avg_delta = float(sum(deltas) / float(len(deltas))) if deltas else 0.0

  pcts: List[float] = []
  for i in range(1, len(y)):
    prev = float(y[i - 1])
    if prev == 0:
      continue
    pcts.append(float((float(y[i]) - prev) / prev))
  avg_pct = float(sum(pcts) / float(len(pcts))) if pcts else 0.0
  if avg_pct > 0.5:
    avg_pct = 0.5
  if avg_pct < -0.5:
    avg_pct = -0.5
  return {"mean": mean_y, "avg_delta": avg_delta, "avg_pct": avg_pct}


def _predict_growth_mean(last_y: float, h: int, stats: Dict[str, float]) -> float:
  yy = float(last_y)
  hh = int(h)
  mean_y = float(stats.get("mean") or 0.0)
  avg_delta = float(stats.get("avg_delta") or 0.0)
  avg_pct = float(stats.get("avg_pct") or 0.0)

  pred_delta = yy + avg_delta * float(hh)
  pred_pct = yy * ((1.0 + avg_pct) ** float(hh))
  base = (pred_delta + pred_pct) / 2.0
  pred = (base + mean_y) / 2.0
  return 0.0 if pred < 0 else float(pred)

def _clamp(x: float, lo: float, hi: float) -> float:
  xx = float(x)
  l = float(lo)
  h = float(hi)
  if xx < l:
    return l
  if xx > h:
    return h
  return xx

def _median(values: List[float]) -> float:
  if not values:
    return 0.0
  xs = sorted(float(v) for v in values)
  n = len(xs)
  mid = n // 2
  if n % 2 == 1:
    return float(xs[mid])
  return float((xs[mid - 1] + xs[mid]) / 2.0)

def _predict_prediales_estacional_yoy(y_ext: List[float]) -> Dict[str, Any]:
  if not y_ext:
    return {"pred": 0.0, "baseline": 0.0, "ratioMedian": None, "ratioCount": 0}

  baseline = float(y_ext[-12]) if len(y_ext) >= 12 else float(y_ext[-1])
  ratios: List[float] = []
  for i in range(12, len(y_ext)):
    prev = float(y_ext[i - 12])
    cur = float(y_ext[i])
    if prev <= 0 or cur <= 0:
      continue
    ratios.append(cur / prev)
  ratios_recent = ratios[-6:] if ratios else []
  ratio_median: Optional[float] = _median(ratios_recent) if ratios_recent else None
  ratio = _clamp(float(ratio_median), 0.7, 1.3) if ratio_median is not None else 1.0

  pred = baseline * ratio
  if pred < 0:
    pred = 0.0
  if baseline == 0.0:
    stats_h = _compute_stats(y_ext)
    pred = _predict_growth_mean(float(y_ext[-1]), 1, stats_h)

  return {"pred": float(pred), "baseline": float(baseline), "ratioMedian": ratio_median, "ratioCount": int(len(ratios_recent))}

def _linear_regression_fit(y: List[float]) -> Dict[str, float]:
  n = len(y)
  if n <= 0:
    return {"intercept": 0.0, "slope": 0.0}
  if n == 1:
    return {"intercept": float(y[0]), "slope": 0.0}
  xs = list(range(n))
  mean_x = (n - 1) / 2.0
  mean_y = sum(y) / float(n)
  sxx = sum((x - mean_x) ** 2 for x in xs)
  if sxx == 0:
    return {"intercept": mean_y, "slope": 0.0}
  sxy = sum((x - mean_x) * (yy - mean_y) for x, yy in zip(xs, y))
  slope = sxy / sxx
  intercept = mean_y - slope * mean_x
  return {"intercept": float(intercept), "slope": float(slope)}


def _linear_regression_predict(intercept: float, slope: float, x: float) -> float:
  y = float(intercept) + float(slope) * float(x)
  return 0.0 if y < 0 else y


@app.get("/api/analitica/saneamiento/ambiental/pronostico")
def analitica_saneamiento_ambiental_pronostico(request: Request) -> ORJSONResponse:
  try:
    cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
    f = _saneamiento_ambiental_filters(dict(request.query_params))

    backtest_raw = request.query_params.get("backtestMonths")
    backtest_months = int(backtest_raw) if backtest_raw and str(backtest_raw).isdigit() else 6
    horizon_months = 6
    backtest_months = max(1, min(24, backtest_months))

    def parse_int_env(name: str, default: int) -> int:
      raw = os.getenv(name)
      if raw in (None, "", "null"):
        return default
      try:
        return int(str(raw).strip())
      except Exception:
        return default

    derecho_fteing = parse_int_env("SAN_AMB_DERECHO_FTEING", 4326010110)
    actualizaciones_fteing = parse_int_env("SAN_AMB_ACTUALIZACION_FTEING", 4326010111)
    recargos_fteing = parse_int_env("SAN_AMB_RECARGOS_FTEING", 4501012601)

    now = datetime.now()
    current_month = (int(now.year), int(now.month))
    start_dt = f["pagoFrom"] or datetime(int(now.year), 1, 1)
    end_dt = f["pagoTo"] or now
    start_month = (int(start_dt.year), int(start_dt.month))
    end_month = (int(end_dt.year), int(end_dt.month))
    horizon_end_month = _add_months(end_month[0], end_month[1], horizon_months)
    horizon_end_month = (int(horizon_end_month[0]), int(horizon_end_month[1]))
    include_real_through = current_month if _month_leq(current_month, horizon_end_month) else horizon_end_month

    training_end = end_month if _month_leq(end_month, current_month) else current_month
    query_to = _month_end_dt(include_real_through[0], include_real_through[1])

    with get_conn() as conn:
      cur = conn.cursor()
      cur.execute(
        _SAN_AMB_MONTHLY_SELECT,
        (
          cve_fte_mt,
          start_dt,
          query_to,
          f["licenciaId"],
          derecho_fteing,
          actualizaciones_fteing,
          recargos_fteing,
        ),
      )
      series_desc = _rows(cur)

    series_asc = sorted(series_desc, key=lambda r: str(r.get("Periodo") or ""))
    total_by_period: Dict[str, float] = {}
    for r in series_asc:
      period = str(r.get("Periodo") or "")
      total_by_period[period] = float(r.get("Total") or 0.0)

    all_months = _month_range(start_month, horizon_end_month)
    training_months = _month_range(start_month, training_end)

    y_train = [float(total_by_period.get(_period_str(yy, mm), 0.0)) for yy, mm in training_months]
    n = len(y_train)
    stats = _compute_stats(y_train)

    forecast_rows: List[Dict[str, Any]] = []
    forecast_start = _add_months(end_month[0], end_month[1], 1)
    forecast_start = (int(forecast_start[0]), int(forecast_start[1]))
    forecast_months = _month_range(forecast_start, horizon_end_month)
    y_ext: List[float] = list(y_train)
    for yy, mm in forecast_months:
      period = _period_str(yy, mm)
      stats_h = _compute_stats(y_ext)
      prev = float(y_ext[-1]) if y_ext else 0.0
      pred = _predict_growth_mean(prev, 1, stats_h)
      forecast_rows.append({"Periodo": period, "Pronostico": float(pred)})

      has_real_for_month = _month_leq((yy, mm), include_real_through)
      if has_real_for_month and period in total_by_period:
        y_ext.append(float(total_by_period.get(period, 0.0)))
      else:
        y_ext.append(float(pred))

    forecast_by_period = {str(r["Periodo"]): float(r["Pronostico"] or 0.0) for r in forecast_rows}

    combined: List[Dict[str, Any]] = []
    for yy, mm in all_months:
      period = _period_str(yy, mm)
      real_val = float(total_by_period.get(period, 0.0)) if _month_leq((yy, mm), include_real_through) else None
      pron = float(forecast_by_period[period]) if period in forecast_by_period else None
      combined.append({"Periodo": period, "Real": real_val, "Pronostico": pron})

    window_rows: List[Dict[str, Any]] = []
    for yy, mm in forecast_months:
      period = _period_str(yy, mm)
      has_real = _month_leq((yy, mm), include_real_through)
      real_raw = float(total_by_period.get(period, 0.0)) if has_real else None
      real_val = float(real_raw) if real_raw is not None else 0.0
      pred_val = float(forecast_by_period.get(period, 0.0))
      err_abs: Optional[float] = None
      err_pct: Optional[float] = None
      if real_raw is not None:
        err_abs = float(abs(real_raw - pred_val))
        err_pct = float((err_abs / abs(real_raw) * 100.0) if real_raw != 0 else 0.0)
      window_rows.append(
        {
          "Periodo": period,
          "Real": real_val,
          "Pronostico": pred_val,
          "ErrorAbs": err_abs,
          "ErrorPct": err_pct,
        }
      )

    start_index = max(2, n - backtest_months)
    backtest_rows: List[Dict[str, Any]] = []
    abs_errors: List[float] = []
    pct_errors: List[float] = []
    for i in range(start_index, n):
      train_slice = y_train[:i]
      stats_i = _compute_stats(train_slice)
      last_i = float(train_slice[-1]) if train_slice else 0.0
      pred_i = _predict_growth_mean(last_i, 1, stats_i)
      real_i = float(y_train[i])
      err_abs = abs(real_i - pred_i)
      err_pct = (err_abs / abs(real_i) * 100.0) if real_i != 0 else 0.0
      abs_errors.append(err_abs)
      pct_errors.append(err_pct)
      backtest_rows.append(
        {
          "Periodo": _period_str(training_months[i][0], training_months[i][1]),
          "Real": real_i,
          "Pronostico": float(pred_i),
          "ErrorAbs": float(err_abs),
          "ErrorPct": float(err_pct),
        }
      )

    mae = float(sum(abs_errors) / len(abs_errors)) if abs_errors else 0.0
    mape = float(sum(pct_errors) / len(pct_errors)) if pct_errors else 0.0

    return ORJSONResponse(
      {
        "ok": True,
        "filtros": {
          "cveFteMT": cve_fte_mt,
          "ejercicio": f["ejercicio"],
          "licenciaId": f["licenciaId"],
          "pagoFrom": f["pagoFrom"].isoformat() if f["pagoFrom"] else None,
          "pagoTo": f["pagoTo"].isoformat() if f["pagoTo"] else None,
          "horizonMonths": horizon_months,
          "backtestMonths": backtest_months,
        },
        "modelo": {
          "tipo": "crecimiento_y_media",
          "n": n,
          "mean": float(stats.get("mean") or 0.0),
          "avgDelta": float(stats.get("avg_delta") or 0.0),
          "avgPct": float(stats.get("avg_pct") or 0.0),
          "maeBacktest": mae,
          "mapeBacktest": mape,
        },
        "historico": series_asc,
        "pronosticos": forecast_rows,
        "pronosticosVsReal": window_rows,
        "backtest": backtest_rows,
        "serieCombinada": combined,
      }
    )
  except Exception as e:
    return ORJSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/api/reportes/saneamiento/ambiental.csv")
def saneamiento_ambiental_csv(request: Request) -> Any:
  cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
  f = _saneamiento_ambiental_filters(dict(request.query_params))

  max_rows_raw = request.query_params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
  max_rows = max(1, min(200000, max_rows))
  if _boolish(request.query_params.get("async"), False) and max_rows >= int(_REPORT_ASYNC_THRESHOLD_ROWS):
    user = _require_login(request)
    params = dict(request.query_params)
    info = _create_report_job("saneamiento_ambiental_csv", params, user)
    _start_report_job(str(info.get("jobId")))
    job_id = str(info.get("jobId"))
    return ORJSONResponse(
      {
        "ok": True,
        "async": True,
        "job": info,
        "urls": {"status": f"/api/reportes/jobs/{job_id}", "download": f"/api/reportes/jobs/{job_id}/download"},
      }
    )

  def parse_int_env(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw in (None, "", "null"):
      return default
    try:
      return int(str(raw).strip())
    except Exception:
      return default

  derecho_fteing = parse_int_env("SAN_AMB_DERECHO_FTEING", 4326010110)
  actualizaciones_fteing = parse_int_env("SAN_AMB_ACTUALIZACION_FTEING", 4326010111)
  recargos_fteing = parse_int_env("SAN_AMB_RECARGOS_FTEING", 4501012601)

  columns = [
    "Padrón",
    "No. Licencia",
    "Serie",
    "Folio",
    "Fecha",
    "Nombre",
    "RFC",
    "Observaciones",
    "Domicilio Licencia",
    "Domicilio Local",
    "Giro",
    "No. Cuartos",
    "Periodo pagado",
    "Derecho",
    "Actualizaciones",
    "Recargos",
    "Total",
  ]

  def gen() -> Iterable[bytes]:
    yield b"\xef\xbb\xbf" + (",".join(columns) + "\n").encode("utf-8")
    conn = get_conn()
    try:
      cur = conn.cursor()
      cur.execute(
        _SAN_AMB_EXPORT_SELECT,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["licenciaId"],
          derecho_fteing,
          actualizaciones_fteing,
          recargos_fteing,
          max_rows,
        ),
      )
      while cur.description is None:
        if not cur.nextset():
          return
      cols = [c[0] for c in cur.description]
      index = {name: i for i, name in enumerate(cols)}
      while True:
        batch = cur.fetchmany(500)
        if not batch:
          break
        for row in batch:
          line = ",".join(_csv_escape(row[index[c]]) if c in index else "" for c in columns) + "\n"
          yield line.encode("utf-8")
    finally:
      conn.close()

  headers = {"Content-Disposition": 'attachment; filename="saneamiento_ambiental.csv"'}
  return StreamingResponse(gen(), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/reportes/saneamiento/ambiental.xlsx")
def saneamiento_ambiental_xlsx(request: Request) -> StreamingResponse:
  cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
  f = _saneamiento_ambiental_filters(dict(request.query_params))

  max_rows_raw = request.query_params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
  max_rows = max(1, min(200000, max_rows))

  def parse_int_env(name: str, default: int) -> int:
    raw = os.getenv(name)
    if raw in (None, "", "null"):
      return default
    try:
      return int(str(raw).strip())
    except Exception:
      return default

  derecho_fteing = parse_int_env("SAN_AMB_DERECHO_FTEING", 4326010110)
  actualizaciones_fteing = parse_int_env("SAN_AMB_ACTUALIZACION_FTEING", 4326010111)
  recargos_fteing = parse_int_env("SAN_AMB_RECARGOS_FTEING", 4501012601)

  columns = [
    "Padrón",
    "No. Licencia",
    "Serie",
    "Folio",
    "Fecha",
    "Nombre",
    "RFC",
    "Observaciones",
    "Domicilio Licencia",
    "Domicilio Local",
    "Giro",
    "No. Cuartos",
    "Periodo pagado",
    "Derecho",
    "Actualizaciones",
    "Recargos",
    "Total",
  ]

  date_cols = {"Fecha"}
  text_cols = {
    "Serie",
    "Nombre",
    "RFC",
    "Observaciones",
    "Domicilio Licencia",
    "Domicilio Local",
    "Giro",
    "Periodo pagado",
  }
  currency_cols = {"Derecho", "Actualizaciones", "Recargos", "Total"}

  totals_start_idx = columns.index("Derecho")
  totals_cols = columns[totals_start_idx:]
  totals: Dict[str, decimal.Decimal] = {c: decimal.Decimal("0") for c in totals_cols}

  wb = Workbook(write_only=True)
  ws = wb.create_sheet("Saneamiento")
  if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

  title_font = Font(bold=True, size=14)
  header_font = Font(bold=True)
  header_fill = PatternFill("solid", fgColor="E7E7E7")
  align_left = Alignment(vertical="top", wrap_text=True)
  align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin = Side(style="thin", color="A0A0A0")
  header_border = Border(bottom=thin)
  totals_border = Border(top=Side(style="thin", color="000000"))

  currency_format = '"$"#,##0.00'
  date_format = "dd/mm/yy"
  text_format = "@"

  title_cell = WriteOnlyCell(ws, value="Sábana de Saneamiento Ambiental")
  title_cell.font = title_font
  ws.append([title_cell])
  ws.append([WriteOnlyCell(ws, value="Generado"), WriteOnlyCell(ws, value=datetime.now())])
  ws.append([WriteOnlyCell(ws, value="CveFteMT"), WriteOnlyCell(ws, value=cve_fte_mt)])
  ws.append([WriteOnlyCell(ws, value="Ejercicio"), WriteOnlyCell(ws, value=f["ejercicio"] or "")])
  ws.append([WriteOnlyCell(ws, value="Pago desde"), WriteOnlyCell(ws, value=f["pagoFrom"] or "")])
  ws.append([WriteOnlyCell(ws, value="Pago hasta"), WriteOnlyCell(ws, value=f["pagoTo"] or "")])
  ws.append([WriteOnlyCell(ws, value="No. Licencia"), WriteOnlyCell(ws, value=f["licenciaId"] or "")])
  ws.append([WriteOnlyCell(ws, value="MaxRows"), WriteOnlyCell(ws, value=max_rows)])
  ws.append([])

  header_row = []
  for name in columns:
    cell = WriteOnlyCell(ws, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = header_border
    header_row.append(cell)
  ws.append(header_row)

  width_map = {
    "A": 10,
    "B": 12,
    "C": 8,
    "D": 10,
    "E": 12,
    "F": 34,
    "G": 14,
    "H": 30,
    "I": 38,
    "J": 38,
    "K": 22,
    "L": 10,
    "M": 18,
    "N": 14,
    "O": 14,
    "P": 14,
    "Q": 14,
  }
  for idx in range(1, len(columns) + 1):
    col_letter = get_column_letter(idx)
    ws.column_dimensions[col_letter].width = width_map.get(col_letter, 16)

  tmp = SpooledTemporaryFile(max_size=32 * 1024 * 1024)

  def file_iter() -> Iterable[bytes]:
    try:
      with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
          _SAN_AMB_EXPORT_SELECT,
          (
            cve_fte_mt,
            f["pagoFrom"],
            f["pagoTo"],
            f["licenciaId"],
            derecho_fteing,
            actualizaciones_fteing,
            recargos_fteing,
            max_rows,
          ),
        )
        while cur.description is None:
          if not cur.nextset():
            break
        cols = [c[0] for c in cur.description] if cur.description else []
        index = {name: i for i, name in enumerate(cols)}

        while True:
          batch = cur.fetchmany(500)
          if not batch:
            break
          for row in batch:
            out_row = []
            for name in columns:
              value = row[index[name]] if name in index else None
              fmt = None
              if name in currency_cols:
                if isinstance(value, decimal.Decimal):
                  d = value
                elif value in (None, ""):
                  d = None
                else:
                  try:
                    d = decimal.Decimal(str(value))
                  except Exception:
                    d = None
                if d is not None:
                  totals[name] += d
                  value = float(d)
                  fmt = currency_format
              elif name in date_cols:
                fmt = date_format
              elif name in text_cols:
                fmt = text_format

              cell = WriteOnlyCell(ws, value=value)
              cell.alignment = align_left
              if fmt:
                cell.number_format = fmt
              out_row.append(cell)
            ws.append(out_row)

      label_idx = max(0, totals_start_idx - 1)
      totals_row = []
      for i, name in enumerate(columns):
        if i == label_idx:
          c = WriteOnlyCell(ws, value="TOTAL")
          c.font = header_font
          c.border = totals_border
          totals_row.append(c)
          continue
        if i >= totals_start_idx:
          d = totals.get(name, decimal.Decimal("0"))
          c = WriteOnlyCell(ws, value=float(d))
          c.font = header_font
          c.number_format = currency_format
          c.border = totals_border
          totals_row.append(c)
          continue
        c = WriteOnlyCell(ws, value=None)
        c.border = totals_border
        totals_row.append(c)
      ws.append(totals_row)

      wb.save(tmp)
      tmp.seek(0)
      while True:
        chunk = tmp.read(1024 * 1024)
        if not chunk:
          break
        yield chunk
    finally:
      try:
        tmp.close()
      except Exception:
        pass

  headers = {"Content-Disposition": 'attachment; filename="saneamiento_ambiental.xlsx"'}
  return StreamingResponse(file_iter(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@app.get("/api/reportes/licencias/funcionamiento.csv")
def licencias_funcionamiento_csv(request: Request) -> Any:
  cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
  f = _licencias_func_filters(dict(request.query_params))

  max_rows_raw = request.query_params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
  max_rows = max(1, min(200000, max_rows))
  if _boolish(request.query_params.get("async"), False) and max_rows >= int(_REPORT_ASYNC_THRESHOLD_ROWS):
    user = _require_login(request)
    params = dict(request.query_params)
    info = _create_report_job("licencias_funcionamiento_csv", params, user)
    _start_report_job(str(info.get("jobId")))
    job_id = str(info.get("jobId"))
    return ORJSONResponse(
      {
        "ok": True,
        "async": True,
        "job": info,
        "urls": {"status": f"/api/reportes/jobs/{job_id}", "download": f"/api/reportes/jobs/{job_id}/download"},
      }
    )

  def parse_int_env(name: str, default: Optional[int] = None) -> Optional[int]:
    raw = os.getenv(name)
    if raw in (None, "", "null"):
      return default
    try:
      return int(str(raw).strip())
    except Exception:
      return default

  licencia_principal = parse_int_env("LIC_FUNC_LICENCIA_FTEING")
  actualizaciones_fteing = parse_int_env("LIC_FUNC_ACTUALIZACION_FTEING", 4319020273)
  recargos_fteing = parse_int_env("LIC_FUNC_RECARGOS_FTEING", 4501010101)
  uma_ref = (
    datetime(int(f["ejercicio"]), 2, 1) if f.get("ejercicio") else (f.get("pagoFrom") or f.get("pagoTo") or datetime.now())
  )
  uma_mxn = get_uma_mxn_for_date(uma_ref) or decimal.Decimal("0")

  columns = [
    "No. Licencia",
    "Serie",
    "Folio",
    "Fecha",
    "Nombre",
    "RFC",
    "Observaciones",
    "Domicilio Licencia",
    "Domicilio Local",
    "Tipo Establecimiento",
    "Giro",
    "Base Licencia Nueva",
    "Base Licencia Renovación",
    "Base Basura Nueva",
    "Tipo",
    "Tarifa Licencia",
    "Tarifa Basura",
    "Licencia",
    "Lic Renovación",
    "Basura",
    "Actualizaciones",
    "Recargos",
    "Otros",
    "Total",
  ]

  def gen() -> Iterable[bytes]:
    yield b"\xef\xbb\xbf" + (",".join(columns) + "\n").encode("utf-8")
    conn = get_conn()
    try:
      cur = conn.cursor()
      cur.execute(
        _LIC_FUNC_CSV,
        (
          cve_fte_mt,
          f["pagoFrom"],
          f["pagoTo"],
          f["tipo"],
          f["licenciaId"],
          f["licenciaFrom"],
          f["licenciaTo"],
          licencia_principal,
          actualizaciones_fteing,
          recargos_fteing,
          f["ejercicio"],
          uma_mxn,
          max_rows,
        ),
      )
      while cur.description is None:
        if not cur.nextset():
          return
      cols = [c[0] for c in cur.description]
      index = {name: i for i, name in enumerate(cols)}
      while True:
        batch = cur.fetchmany(500)
        if not batch:
          break
        for row in batch:
          line = ",".join(_csv_escape(row[index[c]]) if c in index else "" for c in columns) + "\n"
          yield line.encode("utf-8")
    finally:
      conn.close()

  headers = {"Content-Disposition": 'attachment; filename="licencias_funcionamiento.csv"'}
  return StreamingResponse(gen(), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/reportes/licencias/funcionamiento.xlsx")
def licencias_funcionamiento_xlsx(request: Request) -> StreamingResponse:
  cve_fte_mt = str(request.query_params.get("cveFteMT") or "MTULUM")
  f = _licencias_func_filters(dict(request.query_params))

  max_rows_raw = request.query_params.get("maxRows")
  max_rows = int(max_rows_raw) if max_rows_raw and str(max_rows_raw).isdigit() else 50000
  max_rows = max(1, min(200000, max_rows))

  def parse_int_env(name: str, default: Optional[int] = None) -> Optional[int]:
    raw = os.getenv(name)
    if raw in (None, "", "null"):
      return default
    try:
      return int(str(raw).strip())
    except Exception:
      return default

  licencia_principal = parse_int_env("LIC_FUNC_LICENCIA_FTEING")
  actualizaciones_fteing = parse_int_env("LIC_FUNC_ACTUALIZACION_FTEING", 4319020273)
  recargos_fteing = parse_int_env("LIC_FUNC_RECARGOS_FTEING", 4501010101)
  uma_ref = (
    datetime(int(f["ejercicio"]), 2, 1)
    if f.get("ejercicio")
    else (f.get("pagoFrom") or f.get("pagoTo") or datetime.now())
  )
  uma_mxn = get_uma_mxn_for_date(uma_ref) or decimal.Decimal("0")

  columns = [
    "No. Licencia",
    "Serie",
    "Folio",
    "Fecha",
    "Nombre",
    "RFC",
    "Observaciones",
    "Domicilio Licencia",
    "Domicilio Local",
    "Tipo Establecimiento",
    "Giro",
    "Base Licencia Nueva",
    "Base Licencia Renovación",
    "Base Basura Nueva",
    "Tipo",
    "Tarifa Licencia",
    "Tarifa Basura",
    "Licencia",
    "Lic Renovación",
    "Basura",
    "Actualizaciones",
    "Recargos",
    "Otros",
    "Total",
  ]

  date_cols = {"Fecha"}
  text_cols = {
    "Serie",
    "Observaciones",
    "Nombre",
    "RFC",
    "Tipo",
    "Domicilio Licencia",
    "Domicilio Local",
    "Tipo Establecimiento",
    "Giro",
  }
  currency_cols = {
    "Base Licencia Nueva",
    "Base Licencia Renovación",
    "Base Basura Nueva",
    "Tarifa Licencia",
    "Tarifa Basura",
    "Licencia",
    "Lic Renovación",
    "Basura",
    "Actualizaciones",
    "Recargos",
    "Otros",
    "Total",
  }

  totals_start_idx = columns.index("Licencia")
  totals_cols = columns[totals_start_idx:]
  totals: Dict[str, decimal.Decimal] = {c: decimal.Decimal("0") for c in totals_cols}

  wb = Workbook(write_only=True)
  ws = wb.create_sheet("Licencias")
  if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

  title_font = Font(bold=True, size=14)
  header_font = Font(bold=True)
  header_fill = PatternFill("solid", fgColor="E7E7E7")
  align_left = Alignment(vertical="top", wrap_text=True)
  align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin = Side(style="thin", color="A0A0A0")
  header_border = Border(bottom=thin)
  totals_border = Border(top=Side(style="thin", color="000000"))

  currency_format = '"$"#,##0.00'
  date_format = "dd/mm/yy"
  text_format = "@"

  title_cell = WriteOnlyCell(ws, value="Sábana de Licencias de Funcionamiento")
  title_cell.font = title_font
  ws.append([title_cell])
  ws.append([WriteOnlyCell(ws, value="Generado"), WriteOnlyCell(ws, value=datetime.now())])
  ws.append([WriteOnlyCell(ws, value="CveFteMT"), WriteOnlyCell(ws, value=cve_fte_mt)])
  ws.append([WriteOnlyCell(ws, value="Tipo"), WriteOnlyCell(ws, value=f["tipo"])])
  ws.append([WriteOnlyCell(ws, value="Ejercicio"), WriteOnlyCell(ws, value=f["ejercicio"] or "")])
  ws.append([WriteOnlyCell(ws, value="Pago desde"), WriteOnlyCell(ws, value=f["pagoFrom"] or "")])
  ws.append([WriteOnlyCell(ws, value="Pago hasta"), WriteOnlyCell(ws, value=f["pagoTo"] or "")])
  ws.append([WriteOnlyCell(ws, value="Licencia"), WriteOnlyCell(ws, value=f["licenciaId"] or "")])
  ws.append([WriteOnlyCell(ws, value="Rango licencia desde"), WriteOnlyCell(ws, value=f["licenciaFrom"] or "")])
  ws.append([WriteOnlyCell(ws, value="Rango licencia hasta"), WriteOnlyCell(ws, value=f["licenciaTo"] or "")])
  ws.append([WriteOnlyCell(ws, value="MaxRows"), WriteOnlyCell(ws, value=max_rows)])
  ws.append([])

  header_row = []
  for name in columns:
    cell = WriteOnlyCell(ws, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = header_border
    header_row.append(cell)
  ws.append(header_row)

  tmp = SpooledTemporaryFile(max_size=32 * 1024 * 1024)

  def file_iter() -> Iterable[bytes]:
    try:
      with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
          _LIC_FUNC_CSV,
          (
            cve_fte_mt,
            f["pagoFrom"],
            f["pagoTo"],
            f["tipo"],
            f["licenciaId"],
            f["licenciaFrom"],
            f["licenciaTo"],
            licencia_principal,
            actualizaciones_fteing,
            recargos_fteing,
            f["ejercicio"],
            uma_mxn,
            max_rows,
          ),
        )
        while cur.description is None:
          if not cur.nextset():
            break
        cols = [c[0] for c in cur.description] if cur.description else []
        index = {name: i for i, name in enumerate(cols)}

        while True:
          batch = cur.fetchmany(500)
          if not batch:
            break
          for row in batch:
            out_row = []
            for name in columns:
              raw = row[index[name]] if name in index else None
              value = raw
              fmt = None

              if value is None:
                pass
              elif name in date_cols and isinstance(value, datetime):
                fmt = date_format
              elif name in text_cols and isinstance(value, str):
                fmt = text_format
              elif name in currency_cols:
                if isinstance(value, decimal.Decimal):
                  d = value
                elif isinstance(value, (int, float)):
                  d = decimal.Decimal(str(value))
                else:
                  d = None
                if d is not None:
                  value = float(d)
                  fmt = currency_format
                  if name in totals:
                    totals[name] += d

              cell = WriteOnlyCell(ws, value=value)
              cell.alignment = align_left
              if fmt:
                cell.number_format = fmt
              out_row.append(cell)
            ws.append(out_row)

      label_idx = max(0, totals_start_idx - 1)
      totals_row = []
      for i, name in enumerate(columns):
        if i == label_idx:
          c = WriteOnlyCell(ws, value="TOTAL")
          c.font = header_font
          c.border = totals_border
          totals_row.append(c)
          continue
        if i >= totals_start_idx:
          d = totals.get(name, decimal.Decimal("0"))
          c = WriteOnlyCell(ws, value=float(d))
          c.font = header_font
          c.number_format = currency_format
          c.border = totals_border
          totals_row.append(c)
          continue
        c = WriteOnlyCell(ws, value=None)
        c.border = totals_border
        totals_row.append(c)
      ws.append(totals_row)

      wb.save(tmp)
      tmp.seek(0)
      while True:
        chunk = tmp.read(1024 * 1024)
        if not chunk:
          break
        yield chunk
    finally:
      tmp.close()

  headers = {"Content-Disposition": 'attachment; filename="licencias_funcionamiento.xlsx"'}
  return StreamingResponse(
    file_iter(),
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers=headers,
  )


@app.post("/api/reportes/factus")
async def factus(request: Request) -> ORJSONResponse:
  try:
    payload = await request.json()
  except Exception:
    payload = {}

  default_cve_fte_mt = os.getenv("FACTUS_CVE_FTE_MT") or "MTULUM"
  cve_fte_mt = str(payload.get("cveFteMT") or default_cve_fte_mt)
  range_obj = _parse_factus_range(payload.get("range")) or _parse_factus_range(payload)
  raw = payload.get("input") or ""
  items = _parse_factus_input(raw) if not range_obj else []

  def pick_key(keys: List[str], contains_any: List[str], contains_all: Optional[List[str]] = None) -> Optional[str]:
    for k in keys:
      lk = k.lower()
      if any(s in lk for s in contains_any) and (not contains_all or all(s in lk for s in contains_all)):
        return k
    return None

  def to_decimal(value: Any) -> Optional[decimal.Decimal]:
    if value is None:
      return None
    if isinstance(value, decimal.Decimal):
      return value
    if isinstance(value, (int, float)):
      return decimal.Decimal(str(value))
    try:
      return decimal.Decimal(str(value))
    except Exception:
      return None

  def header_fields(header_row: Dict[str, Any]) -> Tuple[str, str, str, str]:
    header_keys = list(header_row.keys()) if header_row else []

    fecha_val = None
    if header_row:
      for k in ["ReciboFecha", "FechaRecibo", "ReciboPagoFecha", "ReciboFechaPago"]:
        if k in header_row:
          fecha_val = header_row.get(k)
          break
      if fecha_val is None:
        fecha_key = pick_key(header_keys, ["fecha"])
        fecha_val = header_row.get(fecha_key) if fecha_key else None
    if isinstance(fecha_val, datetime):
      fecha_val = fecha_val.isoformat()

    nombre_val = header_row.get("ContriRec") if header_row else ""
    rfc_val = header_row.get("RFCRecibo") if header_row else ""
    obs_val = header_row.get("ReciboObservaciones") if header_row else ""
    return (str(fecha_val or ""), str(nombre_val or ""), str(rfc_val or ""), str(obs_val or ""))

  def append_rows_for_receipt(serie: str, folio: int, header_row: Dict[str, Any], detail_rows: List[Dict[str, Any]]) -> None:
    if not detail_rows and not header_row:
      rows_out.append(
        {
          "Serie": serie,
          "Folio": folio,
          "Fecha": "",
          "Nombre": "",
          "RFC": "",
          "Observaciones": "",
          "Concepto": "No encontrado",
          "Total": 0,
        }
      )
      return
    if not detail_rows:
      rows_out.append(
        {
          "Serie": serie,
          "Folio": folio,
          "Fecha": "",
          "Nombre": "",
          "RFC": "",
          "Observaciones": "",
          "Concepto": "Sin detalle",
          "Total": 0,
        }
      )
      return

    fecha_val, nombre_val, rfc_val, obs_val = header_fields(header_row)

    grouped: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for r in detail_rows:
      fte = r.get("CveFteIng")
      fte_str = str(fte) if fte is not None else ""
      if fte_str not in grouped:
        grouped[fte_str] = {"Concepto": None, "Total": decimal.Decimal("0")}
        order.append(fte_str)
      if grouped[fte_str]["Concepto"] in (None, "") and fte_str:
        grouped[fte_str]["Concepto"] = fte_str
      amount_val = to_decimal(r.get("Total"))
      if amount_val is not None:
        grouped[fte_str]["Total"] += amount_val

    for idx, fte_str in enumerate(order):
      concept_val = grouped[fte_str]["Concepto"] or (fte_str if fte_str else "")
      if idx == 0:
        rows_out.append(
          {
            "Serie": serie,
            "Folio": folio,
            "Fecha": fecha_val,
            "Nombre": nombre_val,
            "RFC": rfc_val,
            "Observaciones": obs_val,
            "Concepto": concept_val,
            "Total": float(grouped[fte_str]["Total"]),
          }
        )
      else:
        rows_out.append(
          {
            "Serie": "",
            "Folio": "",
            "Fecha": "",
            "Nombre": "",
            "RFC": "",
            "Observaciones": "",
            "Concepto": concept_val,
            "Total": float(grouped[fte_str]["Total"]),
          }
        )

  rows_out: List[Dict[str, Any]] = []
  database = os.getenv("FACTUS_DB_NAME") or "Tulum_Alterno"
  with get_conn_for_database(database) as conn:
    cur = conn.cursor()
    if range_obj:
      serie = str(range_obj.get("serie") or "").strip().upper()
      folio_from = int(range_obj.get("folioFrom"))
      folio_to = int(range_obj.get("folioTo"))
      max_range_raw = os.getenv("FACTUS_MAX_RANGE") or "500"
      max_range = int(max_range_raw) if str(max_range_raw).isdigit() else 500
      if (folio_to - folio_from + 1) > max_range:
        raise HTTPException(status_code=400, detail=f"Rango demasiado grande (máximo {max_range})")

      cur.execute(
        """
        SELECT CveFolio AS FolioKey, *
        FROM COQRECIBOS
        WHERE CveFteMT = ?
          AND CveSerFol = ?
          AND CveFolio BETWEEN ? AND ?
          AND EdoRec = 'A'
        ORDER BY CveFolio ASC
        OPTION (RECOMPILE);
        """,
        (cve_fte_mt, serie, folio_from, folio_to),
      )
      headers = _rows(cur)
      header_by_folio: Dict[int, Dict[str, Any]] = {}
      for h in headers:
        try:
          fk = int(h.get("FolioKey"))
        except Exception:
          continue
        header_by_folio[fk] = h

      cur.execute(
        """
        SELECT
          CveFolio AS FolioKey,
          CveFteIng,
          SUM(COALESCE(ReciboDetImpAntesDev, 0)) AS Total
        FROM COQRECIBODETALLE
        WHERE CveFteMT = ?
          AND CveSerFol = ?
          AND CveFolio BETWEEN ? AND ?
        GROUP BY CveFolio, CveFteIng
        ORDER BY CveFolio ASC, CveFteIng ASC
        OPTION (RECOMPILE);
        """,
        (cve_fte_mt, serie, folio_from, folio_to),
      )
      details = _rows(cur)
      details_by_folio: Dict[int, List[Dict[str, Any]]] = {}
      for d in details:
        try:
          fk = int(d.get("FolioKey"))
        except Exception:
          continue
        if fk not in details_by_folio:
          details_by_folio[fk] = []
        details_by_folio[fk].append(d)

      for folio in range(folio_from, folio_to + 1):
        append_rows_for_receipt(serie, folio, header_by_folio.get(folio, {}), details_by_folio.get(folio, []))
    else:
      for item in items:
        serie = str(item.get("serie") or "").strip()
        folio = item.get("folio")
        if not serie or folio is None:
          continue

        header_row: Dict[str, Any] = {}
        cur.execute(
          """
          SELECT TOP 1 *
          FROM COQRECIBOS
          WHERE CveFteMT = ?
            AND CveSerFol = ?
            AND CveFolio = ?
            AND EdoRec = 'A'
          OPTION (RECOMPILE);
          """,
          (cve_fte_mt, serie, folio),
        )
        header_try = _rows(cur)
        if header_try:
          header_row = header_try[0]

        cur.execute(
          """
          SELECT
            CveFteIng,
            SUM(COALESCE(ReciboDetImpAntesDev, 0)) AS Total
          FROM COQRECIBODETALLE
          WHERE CveFteMT = ?
            AND CveSerFol = ?
            AND CveFolio = ?
          GROUP BY CveFteIng
          ORDER BY CveFteIng ASC
          OPTION (RECOMPILE);
          """,
          (cve_fte_mt, serie, folio),
        )
        data = _rows(cur)
        append_rows_for_receipt(serie, int(folio), header_row, data)

  return ORJSONResponse({"ok": True, "count": len(rows_out), "rows": rows_out})


_MONTH_NAMES_ES = [
  "Enero",
  "Febrero",
  "Marzo",
  "Abril",
  "Mayo",
  "Junio",
  "Julio",
  "Agosto",
  "Septiembre",
  "Octubre",
  "Noviembre",
  "Diciembre",
]


def _dec(value: Any) -> decimal.Decimal:
  if isinstance(value, decimal.Decimal):
    return value
  if value is None:
    return decimal.Decimal("0")
  if isinstance(value, (int, float)):
    return decimal.Decimal(str(value))
  try:
    return decimal.Decimal(str(value).strip())
  except Exception:
    return decimal.Decimal("0")


_MONEY_Q = decimal.Decimal("0.01")


def _money(value: decimal.Decimal) -> decimal.Decimal:
  try:
    return value.quantize(_MONEY_Q, rounding=decimal.ROUND_HALF_UP)
  except Exception:
    return decimal.Decimal("0.00")


def _predial_ym_index(year: int, month: int) -> int:
  return year * 12 + (month - 1)


def _predial_add_months(year: int, month: int, delta: int) -> Tuple[int, int]:
  idx = _predial_ym_index(year, month) + int(delta)
  y = idx // 12
  m = (idx % 12) + 1
  return y, m


def _predial_ym_key(year: int, month: int) -> str:
  return f"{int(year):04d}-{int(month):02d}"


def _predial_last_day_of_month(year: int, month: int) -> int:
  y2, m2 = _predial_add_months(year, month, 1)
  d = datetime(y2, m2, 1) - timedelta(days=1)
  return int(d.day)


def _predial_safe_date(year: int, month: int, day: int) -> datetime:
  dmax = _predial_last_day_of_month(year, month)
  d = max(1, min(int(day), dmax))
  return datetime(int(year), int(month), d)


def _parse_year_month(obj: Any, label: str) -> Tuple[int, int]:
  if not isinstance(obj, dict):
    raise HTTPException(status_code=400, detail=f"{label} inválido")
  mes_raw = obj.get("mes")
  anio_raw = obj.get("anio")
  try:
    mes = int(str(mes_raw).strip())
    anio = int(str(anio_raw).strip())
  except Exception:
    raise HTTPException(status_code=400, detail=f"{label} inválido")
  if mes < 1 or mes > 12 or anio < 1900 or anio > 2200:
    raise HTTPException(status_code=400, detail=f"{label} inválido")
  return anio, mes


def _predial_months_in_range(start_y: int, start_m: int, end_y: int, end_m: int) -> List[Tuple[int, int]]:
  start_idx = _predial_ym_index(start_y, start_m)
  end_idx = _predial_ym_index(end_y, end_m)
  if end_idx < start_idx:
    return []
  out: List[Tuple[int, int]] = []
  for idx in range(start_idx, end_idx + 1):
    y = idx // 12
    m = (idx % 12) + 1
    out.append((y, m))
  return out


def _parse_inpc_table(raw: Any) -> Dict[int, decimal.Decimal]:
  if raw is None:
    return {}
  if not isinstance(raw, dict):
    raise HTTPException(status_code=400, detail="tablaINPC inválida")
  out: Dict[int, decimal.Decimal] = {}
  for k, v in raw.items():
    ks = str(k or "").strip()
    m = re.match(r"^\s*(\d{4})-(\d{2})\s*$", ks)
    if not m:
      continue
    y = int(m.group(1))
    mo = int(m.group(2))
    if mo < 1 or mo > 12:
      continue
    dv = _dec(v)
    if dv <= 0:
      continue
    out[_predial_ym_index(y, mo)] = dv
  return out


def _parse_tasas_recargos(raw: Any) -> Dict[int, decimal.Decimal]:
  if raw is None:
    return {}
  if not isinstance(raw, dict):
    raise HTTPException(status_code=400, detail="tasasRecargos inválidas")
  out: Dict[int, decimal.Decimal] = {}
  for k, v in raw.items():
    try:
      year = int(str(k).strip())
    except Exception:
      continue
    dv = _dec(v)
    if dv <= 0:
      continue
    out[year] = dv
  return out


def _obtener_inpc_mes_anterior(fecha: datetime, tabla: Dict[int, decimal.Decimal]) -> Tuple[Optional[str], Optional[decimal.Decimal]]:
  if not tabla:
    return None, None
  py, pm = _predial_add_months(fecha.year, fecha.month, -1)
  target = _predial_ym_index(py, pm)
  if target in tabla:
    return _predial_ym_key(py, pm), tabla[target]
  candidates = [idx for idx in tabla.keys() if idx <= target]
  if candidates:
    idx = max(candidates)
    y = idx // 12
    m = (idx % 12) + 1
    return _predial_ym_key(y, m), tabla[idx]
  idx = max(tabla.keys())
  y = idx // 12
  m = (idx % 12) + 1
  return _predial_ym_key(y, m), tabla[idx]


def _factor_actualizacion(inpc_final: Optional[decimal.Decimal], inpc_inicial: Optional[decimal.Decimal]) -> decimal.Decimal:
  if not inpc_final or not inpc_inicial or inpc_final <= 0 or inpc_inicial <= 0:
    return decimal.Decimal("1")
  try:
    f = inpc_final / inpc_inicial
  except Exception:
    return decimal.Decimal("1")
  if f < decimal.Decimal("1"):
    return decimal.Decimal("1")
  return f


def _meses_recargos(fecha_vencimiento: datetime, fecha_pago: datetime) -> int:
  start_y, start_m = _predial_add_months(fecha_vencimiento.year, fecha_vencimiento.month, 1)
  end_y, end_m = fecha_pago.year, fecha_pago.month
  if int(fecha_pago.day) < 17:
    end_y, end_m = _predial_add_months(end_y, end_m, -1)
  start_idx = _predial_ym_index(start_y, start_m)
  end_idx = _predial_ym_index(end_y, end_m)
  if end_idx < start_idx:
    return 0
  months = (end_idx - start_idx) + 1
  return int(max(0, months))


def _predial_bimestre(month: int) -> int:
  return ((int(month) - 1) // 2) + 1


def _predial_bimestres_in_range(start_y: int, start_m: int, end_y: int, end_m: int) -> List[Tuple[int, int]]:
  s_bim_end_m = max(1, min(12, int(((int(start_m) - 1) // 2) + 1) * 2))
  e_bim_end_m = max(1, min(12, int(((int(end_m) - 1) // 2) + 1) * 2))
  start_idx = _predial_ym_index(int(start_y), int(s_bim_end_m))
  end_idx = _predial_ym_index(int(end_y), int(e_bim_end_m))
  if end_idx < start_idx:
    return []
  out: List[Tuple[int, int]] = []
  idx = start_idx
  while idx <= end_idx:
    y = idx // 12
    m = (idx % 12) + 1
    b = _predial_bimestre(m)
    out.append((int(y), int(b)))
    idx += 2
  return out


def _fetch_avaluos_al23(cve_fte_mt: str, predio_id: int, ejercicio_from: int, ejercicio_to: int) -> List[Dict[str, Any]]:
  cve = str(cve_fte_mt or "MTULUM").strip() or "MTULUM"
  pid = int(predio_id)
  y1 = int(ejercicio_from)
  y2 = int(ejercicio_to)
  if y2 < y1:
    y1, y2 = y2, y1
  with get_conn() as conn:
    cur = conn.cursor()
    cur.execute(
      """
      DECLARE @CveFteMT varchar(32) = ?;
      DECLARE @PredioId decimal(18,0) = ?;
      DECLARE @FromEj smallint = ?;
      DECLARE @ToEj smallint = ?;
      SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
      SET LOCK_TIMEOUT 5000;
      SELECT
        a.SolicitudAvaluoId,
        CAST(a.PredioId AS int) AS PredioId,
        a.SolicitudAvaluoStatus,
        a.SolicitudAvaluoAltaFecha,
        CAST(a.SolicitudAvaluoEjercicioAvaluoActual AS int) AS Ejercicio,
        CAST(COALESCE(a.SolicitudAvaluoCatastralImporte, 0) AS decimal(18,2)) AS CatastralImporte,
        CAST(COALESCE(a.SolicitudAvaluoTasaImpuestoPredial, 0) AS decimal(18,8)) AS TasaPredial
      FROM AL23SolicitudAvaluo a
      WHERE a.CveFteMT = @CveFteMT
        AND a.PredioId = @PredioId
        AND a.SolicitudAvaluoStatus = 'A'
        AND a.SolicitudAvaluoEjercicioAvaluoActual BETWEEN @FromEj AND @ToEj
      ORDER BY a.SolicitudAvaluoEjercicioAvaluoActual ASC, a.SolicitudAvaluoAltaFecha ASC, a.SolicitudAvaluoId ASC
      OPTION (RECOMPILE);
      """,
      (cve, pid, y1, y2),
    )
    return _rows(cur)


def _avalue_effective_bim(ejercicio: int, alta_fecha: Any) -> int:
  try:
    if isinstance(alta_fecha, datetime):
      dt = alta_fecha
    else:
      dt = datetime.fromisoformat(str(alta_fecha).replace("Z", "+00:00"))
  except Exception:
    return 1
  y = int(ejercicio)
  if dt.year < y:
    return 1
  if dt.year > y:
    return 6
  return max(1, min(6, _predial_bimestre(dt.month)))


def _build_avaluo_bimestre_map(rows: List[Dict[str, Any]], ejercicio_from: int, ejercicio_to: int) -> Dict[Tuple[int, int], Dict[str, Any]]:
  by_year: Dict[int, List[Dict[str, Any]]] = {}
  for r in rows or []:
    try:
      y = int(r.get("Ejercicio"))
    except Exception:
      continue
    by_year.setdefault(y, []).append(r)

  out: Dict[Tuple[int, int], Dict[str, Any]] = {}
  last_selected_row: Optional[Dict[str, Any]] = None
  min_year = min(by_year.keys()) if by_year else int(ejercicio_from)
  for y in range(int(min_year), int(ejercicio_to) + 1):
    items = by_year.get(y) or []
    enriched = []
    for r in items:
      eff_b = _avalue_effective_bim(y, r.get("SolicitudAvaluoAltaFecha"))
      try:
        aid = int(r.get("SolicitudAvaluoId"))
      except Exception:
        aid = 0
      enriched.append((eff_b, r.get("SolicitudAvaluoAltaFecha"), aid, r))
    enriched.sort(key=lambda t: (int(t[0]), str(t[1] or ""), int(t[2])))

    chosen_for_b6: Optional[Dict[str, Any]] = None
    for b in range(1, 7):
      candidates = [t for t in enriched if int(t[0]) <= b]
      chosen: Optional[Dict[str, Any]] = None
      if candidates:
        chosen = candidates[-1][3]
      elif last_selected_row is not None:
        chosen = last_selected_row
      if chosen is None:
        continue
      if y >= int(ejercicio_from):
        out[(y, b)] = {
          "SolicitudAvaluoId": chosen.get("SolicitudAvaluoId"),
          "SolicitudAvaluoAltaFecha": (
            chosen.get("SolicitudAvaluoAltaFecha").isoformat()
            if isinstance(chosen.get("SolicitudAvaluoAltaFecha"), datetime)
            else chosen.get("SolicitudAvaluoAltaFecha")
          ),
          "SolicitudAvaluoCatastralImporte": str(_money(_dec(chosen.get("CatastralImporte")))),
          "SolicitudAvaluoTasaImpuestoPredial": str(_dec(chosen.get("TasaPredial"))),
        }
      if b == 6:
        chosen_for_b6 = chosen
    if chosen_for_b6 is not None:
      last_selected_row = chosen_for_b6
  return out


def _fetch_predio_alpredio(cve_fte_mt: str, predio_id: Optional[int], clave_catastral: Optional[str], clave_mode: str) -> Optional[Dict[str, Any]]:
  cve = str(cve_fte_mt or "MTULUM").strip() or "MTULUM"
  pid = int(predio_id) if predio_id is not None else None
  clave = (str(clave_catastral or "").strip()) or None
  mode = (str(clave_mode or "exacto").strip().lower()) or "exacto"
  if mode not in {"exacto", "contiene"}:
    mode = "exacto"
  if pid is None and not clave:
    return None
  with get_conn() as conn:
    cur = conn.cursor()
    cur.execute(
      """
      DECLARE @CveFteMT varchar(32) = ?;
      DECLARE @PredioId decimal(18,0) = ?;
      DECLARE @ClaveCatastral varchar(80) = ?;
      DECLARE @ClaveMode varchar(16) = ?;
      SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;
      SET LOCK_TIMEOUT 5000;
      SELECT TOP 1
        p.CveFteMT,
        CAST(p.PredioId AS int) AS PredioId,
        RTRIM(COALESCE(CONVERT(varchar(80), p.PredioCveCatastral) COLLATE DATABASE_DEFAULT, '')) AS PredioCveCatastral,
        RTRIM(COALESCE(CONVERT(varchar(80), p.PredioClavePredial) COLLATE DATABASE_DEFAULT, '')) AS PredioClavePredial,
        p.PredioAltaFecha,
        RTRIM(COALESCE(CONVERT(varchar(32), p.PredioStatus) COLLATE DATABASE_DEFAULT, '')) AS PredioStatus,
        RTRIM(COALESCE(CONVERT(varchar(32), p.PredioTipo) COLLATE DATABASE_DEFAULT, '')) AS PredioTipo,
        RTRIM(COALESCE(CONVERT(varchar(160), p.PredioCalle) COLLATE DATABASE_DEFAULT, '')) AS PredioCalle,
        RTRIM(COALESCE(CONVERT(varchar(32), p.PredioNumExt) COLLATE DATABASE_DEFAULT, '')) AS PredioNumExt,
        RTRIM(COALESCE(CONVERT(varchar(32), p.PredioNumInt) COLLATE DATABASE_DEFAULT, '')) AS PredioNumInt,
        RTRIM(COALESCE(CONVERT(varchar(12), p.PredioCodigoPostal) COLLATE DATABASE_DEFAULT, '')) AS PredioCodigoPostal,
        RTRIM(COALESCE(CONVERT(varchar(250), p.CatastroDatosEscriturales) COLLATE DATABASE_DEFAULT, '')) AS CatastroDatosEscriturales,
        RTRIM(
          COALESCE(
            CONVERT(varchar(250), per.NombreCompletoPersona) COLLATE DATABASE_DEFAULT,
            CONVERT(varchar(250), per.RazonSocialPersona) COLLATE DATABASE_DEFAULT,
            ''
          )
        ) AS PropietarioNombre,
        RTRIM(COALESCE(CONVERT(varchar(32), per.RFCPersona) COLLATE DATABASE_DEFAULT, '')) AS PropietarioRFC,
        CAST(COALESCE(p.PredioTerrenoImporte, 0) AS decimal(18,2)) AS PredioTerrenoImporte,
        CAST(COALESCE(p.PredioConstruccionImporte, 0) AS decimal(18,2)) AS PredioConstruccionImporte,
        CAST(COALESCE(p.PredioCatastralImporte, 0) AS decimal(18,2)) AS PredioCatastralImporte,
        CAST(COALESCE(CAST(val.PredioValuoCatastralImporte AS decimal(18,2)), p.PredioCatastralImporte, 0) AS decimal(18,2)) AS PredioValuoCatastralImporte,
        CAST(COALESCE(p.PredioUltimoEjericicioPagado, 0) AS int) AS PredioUltimoEjericicioPagado,
        CAST(COALESCE(p.PredioUltimoPeriodoPagado, 0) AS int) AS PredioUltimoPeriodoPagado,
        val.PredioValuoCatastralEjercicio,
        val.PredioValuoCatastralFecha
      FROM AlPredio p
      LEFT JOIN XiPersonas per
        ON per.CveFteMT = p.CveFteMT
       AND per.CvePersona = p.CvePersona
      OUTER APPLY (
        SELECT TOP 1
          v.PredioValuoCatastralEjercicio,
          v.PredioValuoCatastralFecha,
          v.PredioValuoCatastralImporte
        FROM ALPREDIOVALUOCATASTRAL v
        WHERE v.CveFteMT = p.CveFteMT
          AND v.PredioId = p.PredioId
        ORDER BY v.PredioValuoCatastralFecha DESC, v.PredioValuoCatastralEjercicio DESC
      ) val
      WHERE p.CveFteMT = @CveFteMT
        AND (@PredioId IS NULL OR p.PredioId = @PredioId)
        AND (
          @ClaveCatastral IS NULL OR
          (
            @ClaveMode = 'exacto'
            AND (
              RTRIM(COALESCE(CONVERT(varchar(80), p.PredioCveCatastral) COLLATE DATABASE_DEFAULT, '')) = @ClaveCatastral
              OR (
                RIGHT(@ClaveCatastral, 1) = '-'
                AND RTRIM(COALESCE(CONVERT(varchar(80), p.PredioCveCatastral) COLLATE DATABASE_DEFAULT, '')) LIKE @ClaveCatastral + '%'
              )
            )
          )
          OR (
            @ClaveMode <> 'exacto'
            AND COALESCE(CONVERT(varchar(80), p.PredioCveCatastral) COLLATE DATABASE_DEFAULT, '') LIKE '%' + @ClaveCatastral + '%'
          )
        )
      ORDER BY p.PredioId ASC
      OPTION (RECOMPILE);
      """,
      (cve, pid, clave, mode),
    )
    rows = _rows(cur)
    return rows[0] if rows else None


@app.get("/api/cajas/predial/predio")
def cajas_predial_buscar_predio(request: Request) -> ORJSONResponse:
  cve_fte_mt = (request.query_params.get("cveFteMT") or "MTULUM").strip() or "MTULUM"
  clave = (request.query_params.get("claveCatastral") or "").strip()
  clave_mode = (request.query_params.get("claveMode") or "exacto").strip().lower()
  predio_id_raw = (request.query_params.get("predioId") or "").strip()
  predio_id = int(predio_id_raw) if predio_id_raw.isdigit() else None
  if predio_id is None and not clave:
    raise HTTPException(status_code=400, detail="Debes indicar predioId o claveCatastral")
  row = _fetch_predio_alpredio(cve_fte_mt, predio_id, clave, clave_mode)
  if not row:
    return ORJSONResponse(status_code=404, content={"ok": False, "detail": "Predio no encontrado"})
  return ORJSONResponse({"ok": True, "predio": row})


@app.post("/api/cajas/predial/pase/preview")
async def cajas_predial_pase_preview(request: Request) -> ORJSONResponse:
  body = await request.json()
  cve_fte_mt = str((body or {}).get("cveFteMT") or "MTULUM").strip() or "MTULUM"
  predio_id_raw = (body or {}).get("predioId")
  predio_id = None
  try:
    if predio_id_raw not in (None, "", "null"):
      predio_id = int(str(predio_id_raw).strip())
  except Exception:
    predio_id = None
  clave = str((body or {}).get("claveCatastral") or "").strip() or None
  clave_mode = str((body or {}).get("claveMode") or "exacto").strip().lower()

  fecha_pago_raw = (body or {}).get("fechaPago")
  fecha_pago = _parse_date(str(fecha_pago_raw or ""))
  if not fecha_pago:
    raise HTTPException(status_code=400, detail="fechaPago inválida")

  periodo_inicio = (body or {}).get("periodoInicio")
  periodo_fin = (body or {}).get("periodoFin")
  start_y_in, start_m_in = _parse_year_month(periodo_inicio, "periodoInicio")
  end_y, end_m = _parse_year_month(periodo_fin, "periodoFin")
  warnings: List[str] = []
  now_local = datetime.now()
  current_year = int(now_local.year)
  adelanto_activo = _normalize_bool(os.getenv("PREDIAL_ADELANTO_ACTIVO"), fallback=False)
  allowed_end_year = current_year + (1 if adelanto_activo else 0)
  if end_y > allowed_end_year:
    warnings.append(
      f"El periodo final se ajustó al año permitido ({allowed_end_year}) porque el año capturado ({end_y}) no es válido."
    )
    end_y = allowed_end_year
    end_m = 12
  if _predial_ym_index(end_y, end_m) < _predial_ym_index(start_y_in, start_m_in):
    raise HTTPException(status_code=400, detail="periodoInicio debe ser <= periodoFin")

  start_b_in = _predial_bimestre(start_m_in)
  start_y = int(start_y_in)
  start_b = int(start_b_in) + 1
  if start_b > 6:
    start_b = 1
    start_y += 1
  start_m = ((start_b - 1) * 2) + 1
  if _predial_ym_index(end_y, end_m) < _predial_ym_index(start_y, start_m):
    warnings.append("Sin adeudo: el periodo final es menor o igual al último periodo pagado.")

  pay_y, pay_m = fecha_pago.year, fecha_pago.month

  tasas_recargos: Dict[int, decimal.Decimal] = {}
  recargos_source: Optional[str] = None
  tasas_raw = (body or {}).get("tasasRecargos")
  if tasas_raw is not None:
    tasas_recargos = _parse_tasas_recargos(tasas_raw)
  else:
    with _RECARGOS_STORE_LOCK:
      tasas_recargos = dict(RECARGOS_MORA_BY_YEAR)
    recargos_source = _RECARGOS_STORE_PATH.name
  tabla_inpc: Dict[int, decimal.Decimal] = {}
  inpc_sources: List[str] = []
  tabla_inpc_raw = (body or {}).get("tablaINPC")
  if tabla_inpc_raw is not None:
    tabla_inpc = _parse_inpc_table(tabla_inpc_raw)
  else:
    with _INPC_STORE_LOCK:
      tabla_inpc, inpc_sources = _load_inpc_table_from_disk()

  dia_vencimiento_raw = (body or {}).get("diaVencimiento")
  try:
    dia_vencimiento = int(str(dia_vencimiento_raw or "15").strip())
  except Exception:
    dia_vencimiento = 15
  dia_vencimiento = max(1, min(28, dia_vencimiento))

  monto_anual_raw = (body or {}).get("montoAnual")
  monto_anual = _dec(monto_anual_raw) if monto_anual_raw not in (None, "", "null") else None
  if monto_anual is not None and monto_anual <= 0:
    raise HTTPException(status_code=400, detail="montoAnual inválido")

  tasa_al_millar_raw = (body or {}).get("tasaAlMillar")
  tasa_al_millar = _dec(tasa_al_millar_raw)

  predio: Optional[Dict[str, Any]] = None
  if predio_id is not None or clave:
    predio = _fetch_predio_alpredio(cve_fte_mt, predio_id, clave, clave_mode)
    if not predio:
      raise HTTPException(status_code=404, detail="Predio no encontrado")
    if predio_id is None:
      try:
        predio_id = int(predio.get("PredioId")) if predio.get("PredioId") is not None else None
      except Exception:
        predio_id = None

  valor_catastral_raw = (body or {}).get("valorCatastral")
  valor_catastral = _dec(valor_catastral_raw) if valor_catastral_raw not in (None, "", "null") else None
  if valor_catastral is None and predio:
    vc = predio.get("PredioValuoCatastralImporte")
    if vc in (None, "", "null") or float(vc or 0) <= 0:
      vc = predio.get("PredioCatastralImporte")
    valor_catastral = _dec(vc)
  if valor_catastral is not None and valor_catastral < 0:
    raise HTTPException(status_code=400, detail="valorCatastral inválido")

  impuesto_anual = monto_anual
  base_calculo = "montoAnual"
  avaluo_map: Dict[Tuple[int, int], Dict[str, Any]] = {}
  if impuesto_anual is None and predio_id is not None:
    try:
      avaluos = _fetch_avaluos_al23(cve_fte_mt, int(predio_id), 1900, int(end_y))
      avaluo_map = _build_avaluo_bimestre_map(avaluos, int(start_y), int(end_y))
    except Exception:
      avaluo_map = {}
    if avaluo_map:
      base_calculo = "AL23SolicitudAvaluo"
      missing = []
      for yy, bb in _predial_bimestres_in_range(int(start_y), int(start_m), int(end_y), int(end_m)):
        if (int(yy), int(bb)) not in avaluo_map:
          missing.append(f"{int(yy)}-B{int(bb)}")
      if missing:
        raise HTTPException(status_code=400, detail=f"Falta avalúo vigente (AL23SolicitudAvaluo) para: {', '.join(missing[:30])}")
  if impuesto_anual is None and not avaluo_map:
    base_calculo = "valorCatastral"
    if tasa_al_millar <= 0:
      raise HTTPException(status_code=400, detail="tasaAlMillar inválida")
    if not valor_catastral or valor_catastral <= 0:
      raise HTTPException(status_code=400, detail="No se pudo determinar el valor catastral / monto anual")
    impuesto_anual = (valor_catastral * (tasa_al_millar / decimal.Decimal("1000")))

  inpc_pago_key, inpc_pago = (None, None)
  if tabla_inpc:
    inpc_pago_key, inpc_pago = _obtener_inpc_mes_anterior(fecha_pago, tabla_inpc)

  desglose_mensual: List[Dict[str, Any]] = []
  total_original = decimal.Decimal("0")
  total_actualizacion = decimal.Decimal("0")
  total_recargos = decimal.Decimal("0")
  total_pagar = decimal.Decimal("0")

  bimestres = _predial_bimestres_in_range(start_y, start_m, end_y, end_m) if _predial_ym_index(end_y, end_m) >= _predial_ym_index(start_y, start_m) else []
  start_idx = _predial_ym_index(int(start_y), int(start_m))
  end_idx = _predial_ym_index(int(end_y), int(end_m))
  if end_idx >= start_idx:
    idx = start_idx
    while idx <= end_idx:
      y = int(idx // 12)
      m = int((idx % 12) + 1)
      b = int(_predial_bimestre(m))
      fecha_venc = _predial_safe_date(y, m, dia_vencimiento)
      avaluo_info = avaluo_map.get((int(y), int(b))) if avaluo_map else None

      cuota_mensual: Optional[decimal.Decimal] = None
      if base_calculo == "AL23SolicitudAvaluo" and avaluo_info:
        base_imp = _dec(avaluo_info.get("SolicitudAvaluoCatastralImporte"))
        tasa_imp = _dec(avaluo_info.get("SolicitudAvaluoTasaImpuestoPredial"))
        anual = base_imp * tasa_imp
        cuota_mensual = anual / decimal.Decimal("12")
      elif impuesto_anual is not None:
        cuota_mensual = impuesto_anual / decimal.Decimal("12")
      if cuota_mensual is None:
        cuota_mensual = decimal.Decimal("0")
      cuota_orig = _money(cuota_mensual)

      status = "con_adeudo"
      if fecha_pago.date() < fecha_venc.date():
        status = "no_vencido"
      elif fecha_pago.date() == fecha_venc.date():
        status = "en_tiempo"

      inpc_v_key, inpc_v = (None, None)
      if tabla_inpc:
        inpc_v_key, inpc_v = _obtener_inpc_mes_anterior(fecha_venc, tabla_inpc)

      factor = decimal.Decimal("1")
      meses_recargo = 0
      if status == "con_adeudo":
        factor = _factor_actualizacion(inpc_pago, inpc_v) if (inpc_pago and inpc_v) else decimal.Decimal("1")
        meses_recargo = _meses_recargos(fecha_venc, fecha_pago)

      monto_act = _money(cuota_orig * factor)
      imp_act = _money(monto_act - cuota_orig) if status == "con_adeudo" else decimal.Decimal("0.00")

      tasa = _recargo_mora_rate_for_year(int(y), tasas_recargos)

      recargos = decimal.Decimal("0.00")
      if status == "con_adeudo" and meses_recargo > 0:
        recargos = _money(monto_act * tasa * decimal.Decimal(str(meses_recargo)))
      subtotal = _money(monto_act + recargos)

      total_original += cuota_orig
      total_actualizacion += imp_act
      total_recargos += recargos
      total_pagar += subtotal

      label = f"{_MONTH_NAMES_ES[m-1]} {y}"
      desglose_mensual.append(
        {
          "mes_label": label,
          "mes": int(m),
          "anio": int(y),
          "bimestre_numero": int(b),
          "fecha_vencimiento": fecha_venc.date().isoformat(),
          "cuota_original": float(cuota_orig),
          "inpc_vencimiento_key": inpc_v_key,
          "inpc_vencimiento": (float(inpc_v) if inpc_v is not None else None),
          "inpc_pago_key": inpc_pago_key,
          "inpc_pago": (float(inpc_pago) if inpc_pago is not None else None),
          "factor_actualizacion": float(_money(factor)),
          "monto_actualizado": float(monto_act),
          "importe_actualizacion": float(imp_act),
          "meses_recargo": int(meses_recargo),
          "tasa_recargos": float(tasa),
          "importe_recargos": float(recargos),
          "subtotal": float(subtotal),
          "status": status,
          "avaluo": avaluo_info,
        }
      )
      idx += 1

  grouped: Dict[str, Dict[str, Any]] = {}
  for item in desglose_mensual:
    y = int(item["anio"])
    m = int(item["mes"])
    b = _predial_bimestre(m)
    key = f"{y}-{b}"
    if key not in grouped:
      grouped[key] = {
        "bimestre_numero": int(b),
        "anio": int(y),
        "meses": [],
        "subtotal_original": decimal.Decimal("0"),
        "subtotal_actualizacion": decimal.Decimal("0"),
        "subtotal_recargos": decimal.Decimal("0"),
        "subtotal_bimestre": decimal.Decimal("0"),
      }
    grouped[key]["meses"].append(str(item["mes_label"]))
    grouped[key]["subtotal_original"] += _dec(item["cuota_original"])
    grouped[key]["subtotal_actualizacion"] += _dec(item["importe_actualizacion"])
    grouped[key]["subtotal_recargos"] += _dec(item["importe_recargos"])
    grouped[key]["subtotal_bimestre"] += _dec(item["subtotal"])

  desglose_bimestral: List[Dict[str, Any]] = []
  for key in sorted(grouped.keys(), key=lambda k: (_predial_ym_index(int(k.split("-")[0]), 1), int(k.split("-")[1]))):
    g = grouped[key]
    anio = int(g["anio"])
    b = int(g["bimestre_numero"])
    m1 = ((b - 1) * 2) + 1
    m2 = m1 + 1
    label = f"{b}er Bimestre ({_MONTH_NAMES_ES[m1-1]}-{_MONTH_NAMES_ES[m2-1]} {anio})" if b == 1 else f"{b}º Bimestre ({_MONTH_NAMES_ES[m1-1]}-{_MONTH_NAMES_ES[m2-1]} {anio})"
    desglose_bimestral.append(
      {
        "bimestre_numero": b,
        "bimestre_label": label,
        "meses": g["meses"],
        "subtotal_original": float(_money(g["subtotal_original"])),
        "subtotal_actualizacion": float(_money(g["subtotal_actualizacion"])),
        "subtotal_recargos": float(_money(g["subtotal_recargos"])),
        "subtotal_bimestre": float(_money(g["subtotal_bimestre"])),
      }
    )

  response = {
    "ok": True,
    "predio": predio,
    "parametros": {
      "cveFteMT": cve_fte_mt,
      "predioId": int(predio.get("PredioId")) if predio and predio.get("PredioId") is not None else predio_id,
      "claveCatastral": (predio.get("PredioCveCatastral") if predio else (clave or "")) or "",
      "base_calculo": base_calculo,
      "monto_anual": (float(_money(impuesto_anual)) if impuesto_anual is not None else None),
      "periodo_inicio": {"mes": int(start_m), "anio": int(start_y)},
      "periodo_inicio_capturado": {"mes": int(start_m_in), "anio": int(start_y_in)},
      "periodo_fin": {"mes": int(end_m), "anio": int(end_y)},
      "fecha_pago": fecha_pago.date().isoformat(),
      "fecha_vencimiento_dia": int(dia_vencimiento),
      "total_meses_calculados": int(len(desglose_mensual)),
      "inpc_sources": inpc_sources,
      "recargos_source": recargos_source,
      "adelanto_activo": adelanto_activo,
      "warnings": warnings,
    },
    "desglose_mensual": desglose_mensual,
    "desglose_bimestral": desglose_bimestral,
    "resumen": {
      "total_cuotas_originales": float(_money(total_original)),
      "total_actualizacion": float(_money(total_actualizacion)),
      "total_recargos": float(_money(total_recargos)),
      "total_a_pagar": float(_money(total_pagar)),
    },
  }

  return ORJSONResponse(response)

 
_SPA_INDEX_PATH = BASE_DIR / "public" / "index.html"


@app.get("/login")
def spa_login() -> FileResponse:
  return FileResponse(_SPA_INDEX_PATH)


@app.get("/ingresos")
@app.get("/ingresos/{path:path}")
def spa_ingresos(path: str = "") -> FileResponse:
  return FileResponse(_SPA_INDEX_PATH)


@app.get("/401")
@app.get("/403")
@app.get("/404")
@app.get("/500")
def spa_errors() -> FileResponse:
  return FileResponse(_SPA_INDEX_PATH)


app.mount("/vendor/react", StaticFiles(directory=BASE_DIR / "node_modules" / "react" / "umd"), name="vendor-react")
app.mount("/vendor/react-dom", StaticFiles(directory=BASE_DIR / "node_modules" / "react-dom" / "umd"), name="vendor-react-dom")
app.mount("/vendor/babel", StaticFiles(directory=BASE_DIR / "node_modules" / "@babel" / "standalone"), name="vendor-babel")
app.mount("/", StaticFiles(directory=BASE_DIR / "public", html=True), name="public")
 
 
if __name__ == "__main__":
  import uvicorn
 
  port_raw = os.getenv("PORT") or "3000"
  port = int(port_raw) if port_raw.isdigit() else 3000
  uvicorn.run(app, host="0.0.0.0", port=port)
